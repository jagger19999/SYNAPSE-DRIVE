import os
import sys

# 自动配置Tcl/Tk环境变量 (解决在venv中运行tkinter报错的问题)
if os.name == 'nt':
    # 尝试定位 venv_appdata 中的 tcl/tk 目录
    # 假设脚本在 .../csv存储.../可视化.../s
    # cript.py
    # venv_appdata 在 .../../../venv_appdata
    # 或者使用绝对路径 C:\Users\85159\正式实验\venv_appdata\tcl
    
    possible_tcl_bases = [
        r"C:\Users\85159\正式实验\venv_appdata\tcl",
        os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "venv_appdata", "tcl")
    ]
    
    for base_dir in possible_tcl_bases:
        tcl_dir = os.path.join(base_dir, "tcl8.6")
        tk_dir = os.path.join(base_dir, "tk8.6")
        if os.path.exists(tcl_dir) and os.path.exists(tk_dir):
            os.environ["TCL_LIBRARY"] = tcl_dir
            os.environ["TK_LIBRARY"] = tk_dir
            print(f"✓ 已自动配置Tcl/Tk路径: {base_dir}")
            break

import tkinter as tk
from tkinter import ttk, scrolledtext, simpledialog, messagebox
import math
import socket
import threading

# Optional imports for plotting
try:
    import matplotlib
    matplotlib.use('TkAgg')
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    from matplotlib.figure import Figure
    import numpy as np
    MATPLOTLIB_AVAILABLE = True
except ImportError as e:
    print(f"Plotting modules missing: {e}")
    MATPLOTLIB_AVAILABLE = False
import time
import csv
import os
import json
import re
import platform
import sys
import traceback
import queue
import collections
from collections import deque
from datetime import datetime
try:
    from gi.repository import GLib, Gst
    Gst.init(None)
    GST_AVAILABLE = True
except Exception:
    GST_AVAILABLE = False

# 尝试导入pylsl
try:
    import pylsl
    from pylsl import StreamInlet, resolve_streams
    PYLSL_AVAILABLE = True
    print("✓ 成功导入pylsl模块")
    
    # 全局LSL解析锁，防止多线程并发调用resolve_streams导致死锁
    LSL_RESOLVE_LOCK = threading.Lock()
    
    # 添加自定义OpenSignals LSL客户端
    import sys
    
    # 尝试多个可能的路径位置
    success_import = False
    
    # 方法1：尝试从项目根目录导入（检查上层目录）
    try:
        # 当前文件路径: LSL同步\主数据采集主文件版本管理\custom_driving_sync - v 8.py
        # opensignals_lsl_client.py可能在: LSL同步\ 或 根目录
        current_dir = os.path.dirname(os.path.abspath(__file__))
        
        # 尝试从LSL同步目录导入
        lsl_sync_dir = os.path.abspath(os.path.join(current_dir, '..'))
        # from opensignals_lsl_client import OpenSignalsLSLClient
        # print(f"✓ 从{lsl_sync_dir}成功导入OpenSignalsLSLClient")
        # success_import = True
        success_import = False # 强制使用本地类以确保线程安全和严格匹配
    except ImportError:
        pass
    
    # 方法2：尝试从根目录导入
    if not success_import:
        try:
            root_dir = os.path.abspath(os.path.join(current_dir, '..', '..'))
            sys.path.append(root_dir)
            # from opensignals_lsl_client import OpenSignalsLSLClient
            # print(f"✓ 从{root_dir}成功导入OpenSignalsLSLClient")
            # success_import = True
            success_import = False # 强制使用本地类
        except ImportError:
            pass
    
    # 方法3：尝试直接从当前目录导入（检查是否有副本）
    if not success_import:
        try:
            if '.' not in sys.path:
                sys.path.append('.')
            # from opensignals_lsl_client import OpenSignalsLSLClient
            # print("✓ 从当前目录成功导入OpenSignalsLSLClient")
            # success_import = True
            success_import = False # 强制使用本地类
        except ImportError:
            pass
    
    # 如果都失败了，创建模拟类
    if not success_import:
        print("✗ 未能导入OpenSignalsLSLClient模块，创建模拟类")
        class OpenSignalsLSLClient:
            def __init__(self, mac_address=""):
                print(f"使用模拟的OpenSignalsLSLClient类，mac_address={mac_address}")
                self.inlet = None
                self.stream_name = "OpenSignals"
                self.mac_address = mac_address
                self.found_stream_info = None
            
            def find_and_connect(self):
                """增强的LSL流查找方法，确保能够可靠地找到生理信号流"""
                try:
                    print("🔍 正在查找LSL流...")
                    # 延长等待时间到3秒，确保有足够时间找到流
                    print("⏱️  等待时间设置为3秒，正在查找所有可用LSL流...")
                    # 使用全局锁防止并发调用导致死锁
                    if 'LSL_RESOLVE_LOCK' in globals():
                        with LSL_RESOLVE_LOCK:
                            streams = resolve_streams(wait_time=3.0)
                    else:
                        streams = resolve_streams(wait_time=3.0)
                    
                    if not streams:
                        print("⚠ 未找到任何LSL流！请确认生理日志系统已正确打开")
                        return False
                    
                    print(f"✓ 成功检测到 {len(streams)} 个LSL流")
                    
                    # 打印所有流的详细信息，便于调试
                    print("📋 所有可用LSL流信息:")
                    for i, stream in enumerate(streams):
                        name = stream.name()
                        stream_type = stream.type()
                        channel_count = stream.channel_count()
                        nominal_srate = stream.nominal_srate()
                        source_id = stream.source_id()
                        print(f"  [{i+1}] 名称: '{name}', 类型: '{stream_type}', 通道数: {channel_count}, 采样率: {nominal_srate}Hz, 源ID: '{source_id}'")
                    
                    # 优先选择可能的生理信号流
                    # 1. 严格匹配：仅连接包含明确生理信号关键词的流
                    # 遵循 "A对A (驾驶), B对B (生理)" 的原则，移除不稳定的兜底逻辑
                    keywords = ['OpenSignals', 'BioSemi', 'BioPac', '生理', '心电', 'EMG', 'ECG', 'EDA', 'GSR', 'PPG', 'Resp']
                    for stream in streams:
                        name_lower = stream.name().lower()
                        type_lower = stream.type().lower()
                        
                        # 检查是否包含关键词
                        if any(keyword.lower() in name_lower or keyword.lower() in type_lower for keyword in keywords):
                            print(f"🚀 找到生理信号流: {stream.name()} (类型: {stream.type()})")
                            self.inlet = StreamInlet(stream)
                            self.found_stream_info = stream
                            print(f"✓ 已成功连接到生理信号流: {stream.name()}")
                            return True
                    
                    print("ℹ️  未找到符合关键词的生理信号流 (OpenSignals/BioSemi/EMG/ECG/EDA...)")
                    print("    请确保生理信号软件 (如OpenSignals) 已开启LSL广播，且名称包含上述关键词。")
                    
                    # 移除原有的 "Option 2" (任意通道数>=2) 和 "Option 3" (兜底任意流)
                    # 避免错误连接到驾驶数据流或其他不相关流
                    
                    return False
                except Exception as e:
                    print(f"✗ 查找LSL流时发生错误: {str(e)}")
                    import traceback
                    traceback.print_exc()  # 打印详细错误信息
                return False
            
            def receive_data(self, blocking=True, timeout=0.005, max_samples=1000):
                """优化的LSL数据接收方法，使用批量获取提高1000Hz采样率性能"""
                if self.inlet is None:
                    return None, None
                
                try:
                    # 对于高频数据，使用pull_chunk批量获取更多样本
                    # 减少超时时间，增加最大样本数，以提高1000Hz采样率性能
                    samples, timestamps = self.inlet.pull_chunk(timeout=timeout, max_samples=max_samples)
                    
                    # 如果获取到多个样本，返回所有样本作为列表（优化为支持批量处理）
                    if samples and len(samples) > 0 and timestamps and len(timestamps) > 0:
                        # 返回所有样本的列表，每个元素是(sample, timestamp)元组
                        return [(samples[i], timestamps[i]) for i in range(len(samples))]
                    
                    # 如果没有获取到数据，返回None
                    return None
                except Exception as e:
                    # 减少错误日志频率，避免影响高频采集
                    if not hasattr(self, '_error_log_counter'):
                        self._error_log_counter = 0
                    self._error_log_counter += 1
                    if self._error_log_counter % 100 == 0:
                        print(f"⚠ LSL数据接收错误: {str(e)}")
                    return None
            
            def connect(self):
                """为了兼容V9版本的方法"""
                return self.find_and_connect()
            
            def start_stream(self):
                """为了兼容V9版本的方法"""
                pass
            
            def stop_stream(self):
                """为了兼容V9版本的方法"""
                pass
            
            def disconnect(self):
                """为了兼容V9版本的方法"""
                self.close()
            
            def get_sample(self):
                """为了兼容V9版本的方法"""
                sample, timestamp = self.receive_data()
                return sample
            
            def close(self):
                """关闭LSL连接"""
                if self.inlet is not None:
                    try:
                        self.inlet.close_stream()
                        print("✓ LSL流已关闭")
                    except Exception:
                        pass
                    self.inlet = None
except ImportError:
    PYLSL_AVAILABLE = False
    print("✗ 未能导入pylsl模块，生物信号功能将不可用")
    
    # 为了兼容性，即使在没有pylsl的情况下也定义模拟类
    class OpenSignalsLSLClient:
        def __init__(self):
            print("使用模拟的OpenSignalsLSLClient类 - 无pylsl模块")
            self.inlet = None
        
        def find_and_connect(self):
            print("⚠ pylsl模块未导入，无法连接到LSL流")
            return False
        
        def receive_data(self, blocking=True, timeout=1.0):
            return None, None
        
        def close(self):
            print("✓ 模拟LSL流已关闭")

class DataSyncBuffer:
    def __init__(self):
        self.lock = threading.Lock()
        self.buffer = []
        self.gaze_buffer = []
    def add_packet(self, pkt):
        with self.lock:
            self.buffer.append(pkt)
    def add_gaze_data(self, pts, gaze_ts, gaze_x, gaze_y):
        with self.lock:
            self.gaze_buffer.append({'pts': pts, 'gaze_ts': gaze_ts, 'gaze_x': gaze_x, 'gaze_y': gaze_y})
    def pop_data_up_to_now(self):
        with self.lock:
            arr = []
            gaze_arr = []
            latest = None
            if not self.buffer:
                return arr, latest, gaze_arr
            
            # 找到最新的视频帧时间戳
            latest_video_pts = self.buffer[-1].get('pts') if self.buffer else None

            # 弹出所有早于或等于最新视频帧时间戳的视频数据包
            i = 0
            while i < len(self.buffer):
                pkt = self.buffer[i]
                if pkt.get('pts') <= latest_video_pts:
                    arr.append(pkt)
                    latest = pkt
                    i += 1
                else:
                    break
            self.buffer = self.buffer[i:]

            # 弹出所有早于或等于最新视频帧时间戳的凝视数据包
            j = 0
            while j < len(self.gaze_buffer):
                gaze_pkt = self.gaze_buffer[j]
                if gaze_pkt.get('pts') <= latest_video_pts:
                    gaze_arr.append(gaze_pkt)
                    j += 1
                else:
                    break
            self.gaze_buffer = self.gaze_buffer[j:]
            
            return arr, latest, gaze_arr

class LSLPlotter:
    def __init__(self, parent_frame, combo_widget=None, stream_combo_widget=None):
        self.parent = parent_frame
        self.combo = combo_widget
        self.stream_combo = stream_combo_widget
        self.channel_idx = 0
        self.channel_names = []
        self.available_streams = [] # Store StreamInfo objects
        self.current_stream_idx = -1
        
        # 即使没有matplotlib，我们也使用Tkinter Canvas绘图作为备选
        if globals().get('MATPLOTLIB_AVAILABLE', False):
            self.use_matplotlib = True
            # Setup Figure
            self.fig = Figure(figsize=(5, 2), dpi=100)
            self.ax = self.fig.add_subplot(111)
            self.ax.set_title("LSL Signal Monitor")
            self.ax.set_xlabel("Time (s)")
            self.ax.grid(True)
            self.line, = self.ax.plot([], [], lw=1)
            self.canvas = FigureCanvasTkAgg(self.fig, master=self.parent)
            self.canvas.draw()
            self.canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=True)
        else:
            self.use_matplotlib = False
            # 使用原生Tkinter Canvas绘图
            self.canvas = tk.Canvas(self.parent, bg='white', height=200)
            self.canvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
            self.canvas.create_text(200, 100, text="LSL Signal Monitor (Tkinter Mode)", fill="gray")
            self.line_id = None
            self.width = 400
            self.height = 200

        self.running = False
        self.inlets = []
        self.plot_duration = 5
        self.update_interval = 50  # ms
        self.max_samples = 5000  # Safety limit
        
        self.buffer_data = [0] * 500  # Default buffer
        self.inlet = None
        self.searching = True  # Control flag for stream search loop
        
        # Start looking for streams if available
        if 'PYLSL_AVAILABLE' in globals() and PYLSL_AVAILABLE:
            self.resolve_thread = threading.Thread(target=self.find_streams, daemon=True)
            self.resolve_thread.start()
        else:
            if self.use_matplotlib:
                self.ax.text(0.5, 0.5, "LSL Not Available", ha='center', va='center')
                self.canvas.draw()
            else:
                self.canvas.create_text(self.width/2, self.height/2, text="LSL Not Available", fill="red")

    def _safe_float(self, x):
        try:
            if x is None:
                return 0.0
            if isinstance(x, (int, float)):
                return float(x)
            s = str(x).strip()
            if "," in s and "." not in s:
                s = s.replace(",", ".")
            return float(s)
        except Exception:
            return 0.0

    def _chunk_to_float_array(self, chunk):
        if 'np' not in globals():
            return None
        if chunk is None:
            return None
        if isinstance(chunk, np.ndarray):
            arr = chunk
        else:
            arr = np.asarray(chunk)
        if arr.size == 0:
            return None
        if arr.dtype.kind in {"f", "i", "u"}:
            return arr.astype(float, copy=False)
        try:
            return arr.astype(float)
        except Exception:
            pass
        if arr.ndim == 1 and len(arr) > 0 and isinstance(arr[0], (list, tuple, np.ndarray)):
            try:
                return np.asarray(chunk, dtype=float)
            except Exception:
                pass
        try:
            rows = []
            for sample in chunk:
                if isinstance(sample, (list, tuple, np.ndarray)):
                    rows.append([self._safe_float(v) for v in sample])
                else:
                    rows.append([self._safe_float(sample)])
            return np.asarray(rows, dtype=float)
        except Exception:
            return None

    def _choose_default_channel_idx(self, inlet, info):
        if not hasattr(info, "channel_count"):
            return 0
        channel_count = info.channel_count()
        if channel_count <= 1:
            return 0
        if self.channel_names:
            keyword_priority = [
                "ecg",
                "ekg",
                "eeg",
                "emg",
                "ppg",
                "bvp",
                "eda",
                "gsr",
                "resp",
                "heart",
                "hr",
            ]
            for kw in keyword_priority:
                for i, name in enumerate(self.channel_names):
                    if kw in str(name).lower():
                        return i

        if 'np' not in globals():
            return 0

        max_samples = int(min(max(self.srate * 0.5, 50), 2000))
        try:
            chunk, _ = inlet.pull_chunk(timeout=0.2, max_samples=max_samples)
        except TypeError:
            chunk, _ = inlet.pull_chunk(timeout=0.2)

        arr = self._chunk_to_float_array(chunk)
        if arr is None or arr.size == 0:
            return 0
        if arr.ndim == 1:
            return 0
        if arr.shape[1] != channel_count:
            channel_count = min(channel_count, arr.shape[1])

        best_idx = 0
        best_score = -1.0
        for i in range(channel_count):
            col = arr[:, i]
            col = col[np.isfinite(col)]
            if col.size < 20:
                continue
            try:
                p10, p90 = np.percentile(col, [10, 90])
                spread = float(p90 - p10)
                uniq = np.unique(np.round(col, 6)).size
                binary_like = uniq <= 3 and spread <= 1.1
                score = spread
                if not binary_like:
                    score *= 1.5
                if score > best_score:
                    best_score = score
                    best_idx = i
            except Exception:
                continue
        return best_idx

    def set_channel(self, idx):
        if idx >= 0:
            self.channel_idx = idx
            # Clear buffer on channel switch to avoid jumps
            if self.use_matplotlib:
                self.buffer_data = np.zeros(self.buffer_size)
            else:
                self.buffer_data = [0] * len(self.buffer_data)
            if self.use_matplotlib:
                 self.line.set_ydata(self.buffer_data)
                 self.canvas.draw_idle()

    def set_stream(self, idx):
        """Switch to a different LSL stream"""
        if 0 <= idx < len(self.available_streams):
            if idx == self.current_stream_idx:
                return
                
            info = self.available_streams[idx]
            print(f"Switching to stream: {info.name()}")
            
            # Close existing inlet if any
            # Note: pylsl.StreamInlet doesn't have a close method, it's handled by GC
            # but we should stop pulling from it
            self.inlet = None
            
            try:
                # Create new inlet
                inlet = StreamInlet(info, max_buflen=self.plot_duration,
                                  processing_flags=pylsl.proc_clocksync | pylsl.proc_dejitter)
                
                self._setup_inlet(inlet, info)
                self.current_stream_idx = idx
            except Exception as e:
                print(f"Error switching stream: {e}")

    def find_streams(self):
        print("Plotter started searching for streams...")
        while self.searching:
            try:
                # Use a short timeout to not block too long
                streams = resolve_streams(wait_time=1.0)
                
                # Update available streams list
                valid_streams = [s for s in streams if s.type() != 'Markers']
                
                # Check if stream list changed
                current_names = [s.name() + " (" + s.type() + ")" for s in self.available_streams]
                new_names = [s.name() + " (" + s.type() + ")" for s in valid_streams]
                
                if new_names != current_names:
                    self.available_streams = valid_streams
                    # Update combo box in main thread
                    if self.stream_combo:
                        def update_combo():
                            self.stream_combo['values'] = new_names
                            # If we have streams but none selected, select the first one
                            if new_names and self.current_stream_idx == -1:
                                self.stream_combo.current(0)
                                self.set_stream(0)
                            # If current selection is invalid (e.g. stream gone), reset
                            elif self.current_stream_idx >= len(new_names):
                                if new_names:
                                    self.stream_combo.current(0)
                                    self.set_stream(0)
                                else:
                                    self.stream_combo.set("无可用数据流")
                                    self.current_stream_idx = -1
                        
                        self.parent.after(0, update_combo)
                
                # Wait before retrying
                time.sleep(2.0)
            except Exception as e:
                print(f"Error resolving streams for plot: {e}")
                time.sleep(1.0)

    def _setup_inlet(self, inlet, info):
        print(f"Setup inlet for {info.name()}. Rate={info.nominal_srate()}, Channels={info.channel_count()}")
        self.inlet = inlet
        self.info = info
        self.srate = info.nominal_srate()
        if self.srate <= 0: self.srate = 100  # Default if irregular

        # Get channel names
        try:
            ch = info.desc().child("channels").child("channel")
            self.channel_names = []
            for k in range(info.channel_count()):
                name = ch.child_value("label")
                if not name: name = f"Ch {k+1}"
                self.channel_names.append(name)
                ch = ch.next_sibling()
        except:
            self.channel_names = [f"Ch {k+1}" for k in range(info.channel_count())]

        # Update combo box
        if self.combo:
            self.combo['values'] = self.channel_names
            if self.channel_names:
                default_idx = self._choose_default_channel_idx(inlet, info)
                try:
                    self.combo.current(default_idx)
                except Exception:
                    pass
                self.channel_idx = default_idx

        # Pre-allocate buffer for display (circular buffer simulation)
        self.buffer_size = int(self.srate * self.plot_duration)
        if self.use_matplotlib:
            self.buffer_data = np.zeros(self.buffer_size)
            # Reset line data if needed
            if hasattr(self, 'line'):
                self.line.set_ydata(self.buffer_data)
                self.line.set_xdata(np.linspace(0, self.plot_duration, self.buffer_size))
        else:
            self.buffer_data = [0.0] * self.buffer_size
        
        self.start()

    def start(self):
        self.running = True
        self.update_plot()

    def stop(self):
        self.running = False
        self.searching = False

    def update_plot(self):
        if not self.running:
            return

        if self.inlet is None:
            self.parent.after(1000, self.update_plot)
            return

        try:
            # Pull chunk
            chunk, ts = self.inlet.pull_chunk(timeout=0.0)
            if ts:
                # Append data
                if self.use_matplotlib:
                    try:
                        arr = self._chunk_to_float_array(chunk)
                        if arr is None or arr.size == 0:
                            data = None
                        elif arr.ndim > 1:
                            ch = self.channel_idx if self.channel_idx < arr.shape[1] else 0
                            data = arr[:, ch]
                        else:
                            data = arr
                        if data is None:
                            raise ValueError("empty_chunk")
                        
                        n = len(data)
                        if n < self.buffer_size:
                            self.buffer_data = np.roll(self.buffer_data, -n)
                            self.buffer_data[-n:] = data
                        else:
                            self.buffer_data = data[-self.buffer_size:]
                        
                        # Update plot
                        x = np.linspace(0, self.plot_duration, len(self.buffer_data))
                        self.line.set_data(x, self.buffer_data)
                        self.ax.set_xlim(0, self.plot_duration)

                        y_data = self.buffer_data[np.isfinite(self.buffer_data)]
                        if y_data.size > 0:
                            p_min, p_max = np.percentile(y_data, [1, 99])
                            if p_max - p_min <= 1e-9:
                                y_min, y_max = float(np.min(y_data)), float(np.max(y_data))
                            else:
                                y_min, y_max = float(p_min), float(p_max)
                            if y_max <= y_min:
                                y_min -= 0.5
                                y_max += 0.5
                            y_margin = (y_max - y_min) * 0.1
                            self.ax.set_ylim(y_min - y_margin, y_max + y_margin)

                        self.canvas.draw_idle()
                    except Exception:
                        pass
                
                else:
                    # Tkinter drawing logic
                    # Process chunk for list
                    for sample in chunk:
                        raw_val = sample[self.channel_idx] if isinstance(sample, (list, tuple)) and len(sample) > self.channel_idx else (sample[0] if isinstance(sample, (list, tuple)) else sample)
                        try:
                            val = float(raw_val)
                        except (ValueError, TypeError):
                            val = 0.0
                        self.buffer_data.pop(0)
                        self.buffer_data.append(val)
                    
                    # Redraw canvas
                    self.width = self.canvas.winfo_width()
                    self.height = self.canvas.winfo_height()
                    if self.width < 10: self.width = 400
                    
                    self.canvas.delete("all")
                    
                    # Calculate scaling
                    data = self.buffer_data
                    try:
                        p_min, p_max = np.percentile(np.asarray(data, dtype=float), [1, 99])
                        min_val = float(p_min)
                        max_val = float(p_max)
                    except Exception:
                        min_val = min(data)
                        max_val = max(data)
                    range_val = max_val - min_val
                    if range_val == 0:
                        range_val = 1.0
                    
                    # Create points for line
                    points = []
                    step_x = self.width / len(data)
                    
                    for i, val in enumerate(data):
                        x = i * step_x
                        # Invert Y (canvas origin is top-left)
                        y = self.height - ((val - min_val) / range_val * (self.height - 20) + 10)
                        points.append(x)
                        points.append(y)
                    
                    if len(points) >= 4:
                        self.canvas.create_line(points, fill="blue", width=1)
                        # Draw stats
                        self.canvas.create_text(10, 10, anchor="nw", text=f"Max: {max_val:.2f}", fill="black")
                        self.canvas.create_text(10, self.height-10, anchor="sw", text=f"Min: {min_val:.2f}", fill="black")
                        # Draw current value
                        current_val = data[-1]
                        self.canvas.create_text(self.width-10, 10, anchor="ne", text=f"Current: {current_val:.2f}", fill="red")

        except Exception as e:
            print(f"Plot update error: {e}")
            # If stream is lost, stop trying to pull from this inlet
            if "lost" in str(e):
                self.inlet = None

        self.parent.after(self.update_interval, self.update_plot)
    
    def reset_buffer(self):
        """重置缓冲区以防止内存溢出"""
        try:
            if self.use_matplotlib:
                self.buffer_data = np.zeros(self.buffer_size)
                self.line.set_ydata(self.buffer_data)
                self.canvas.draw_idle()
            else:
                self.buffer_data = [0.0] * self.buffer_size
                self.canvas.delete("all")
        except Exception:
            pass

class CustomDrivingSyncSystem:
    # 添加datetime作为类属性
    from datetime import datetime
    
    def __init__(self, root):
        self.root = root
        self.root.title("驾驶模拟器数据同步系统")
        self.root.geometry("900x800")
        
        # 创建驾驶指标保存文件夹
        self.create_data_folder()
        
        # 系统状态变量
        self.running = False
        self.current_driving_params = {}
        self.last_driving_params = {}
        self.current_biosig_params = {}  # 生物信号参数
        self.csv_file = None
        self.csv_writer = None
        self.csv_filename = None
        self.udp_socket = None
        self.biosig_inlet = None
        # 为不同数据类型创建独立的锁，避免线程冲突
        self.driving_lock = threading.Lock()
        self.biosig_lock = threading.Lock()
        self.cache_lock = threading.Lock()
        self.last_udp_status_time = time.time()
        self.system_start_time = None  # 记录系统启动时间
        self._biosig_sample_count = 0  # 生物信号样本计数
        self.biosig_connected = False  # 生物信号连接状态
        
        # 数据缓存相关变量 - 修改为分开缓存驾驶数据和生物信号数据
        self.driving_data_cache = []  # 用于缓存驾驶数据
        self.biosig_data_cache = []   # 用于缓存生物信号数据
        self._marker_event_queue = [] # 驾驶/生物信号事件队列
        self._eye_marker_queue = []   # 眼动事件队列 (独立，避免竞态条件)
        try:
            self.marker_udp_enable = True
            self.marker_udp_host = "::1"
            self.marker_udp_port = 50090
        except Exception:
            pass
        try:
            self.msys_python_path = r"C:\\msys64\\mingw64\\bin\\python3.exe"
            self.proc_v16 = None
        except Exception:
            pass
        self.cache_start_time = None  # 缓存开始时间
        self.last_backup_time = 0  # 上次备份时间
        self.backup_interval = 30  # 适当调整备份间隔，平衡安全性和性能
        self.actual_start_time = None  # 记录首次接收到数据的时间
        self.max_cache_size = 50000  # 紧急修复：增加缓存大小以适应1000Hz采样率
        self.cache_flush_batch_size = 10000  # 每次刷新的批处理大小
        self.cache_flush_interval = 900
        
        # 时间同步相关变量
        self.time_sync_enabled = False  # 时间同步功能启用标志
        self.time_offset = 0.0  # LSL时间与系统时间的偏移量
        self.time_sync_threshold = 0.1  # 时间偏差阈值（秒）
        self.sync_samples_collected = 0  # 已收集的同步样本数
        self.sync_samples_required = 10  # 计算时间偏移所需的样本数
        self.sync_offsets = []  # 存储收集的时间偏移样本
        self.driving_received = False  # 驾驶数据是否已接收
        self.biosig_received = False  # 生物信号数据是否已接收
        self.recent_driving_ts = deque(maxlen=2000)
        self.recent_biosig_ts = deque(maxlen=2000)
        self.last_driving_system_ts = None
        self.last_biosig_system_ts = None
        
        # 创建缓存文件夹
        self.cache_folder = os.path.join(self.data_folder, "缓存")
        if not os.path.exists(self.cache_folder):
            os.makedirs(self.cache_folder)
            self.log(f"创建缓存文件夹: {self.cache_folder}")
        
        # UI更新节流控制 - 为不同数据类型使用独立的更新控制
        self.last_ui_update_time_driving = 0
        self.last_ui_update_time_biosig = 0
        self.last_ui_update_time_cache = 0
        self.last_ui_update_time_log = 0  # 日志更新控制
        self.ui_update_interval = 800
        
        # 驾驶参数UI显示优化配置
        self.ui_update_interval_driving = 300  # 驾驶参数UI更新间隔增加到0.3秒（3Hz）
        self.max_driving_display_lines = 8  # 减少到8个最关键参数
        self.last_displayed_driving_params = {}  # 用于检测内容变化的缓存
        
        # 90分钟实验专用配置
        self.data_sampling_rate = 10  # 数据采样率：每10个样本显示1个（100Hz等效）
        self.sample_counter = 0  # 采样计数器
        self.streaming_write_enabled = False
        self.last_flush_time = time.time()
        self.cache_keep_last_n = 5000
        
        # 性能优化相关变量
        self.data_queue = deque()
        self.data_queue_max_size = 5000  # 增大队列容量以处理高频数据
        self.last_queue_process_time = 0
        self.queue_process_interval = 50
        self.flush_batch_size_driving = 2000
        self.flush_batch_size_biosig = 5000
        self.driving_target_rate = 200
        self.driving_min_interval_ms = 1000.0 / self.driving_target_rate
        self._last_driving_emit_time = None
        self.soft_cache_limit = 120000
        self.hard_cache_limit = 200000
        self._high_pressure_mode = False
        
        # 生物信号高频采样保障机制
        self.biosig_batch_size = 100  # 批量处理样本数
        self.biosig_batch_buffer = []  # 批量处理缓冲区
        self.biosig_batch_buffer_max = 500  # 缓冲区最大大小
        self.last_biosig_batch_time = 0  # 上次批量处理时间
        self.biosig_sampling_rate_target = 1000  # 目标采样率
        self.biosig_last_timestamp = None  # 上次样本时间戳
        self.biosig_missed_samples = 0  # 丢失样本计数
        self.biosig_last_check_time = 0  # 上次检查时间
        self.biosig_check_interval = 5.0  # 采样率检查间隔(秒)
        self.lsl_receive_timeout = 0.005  # 生理LSL拉取超时时间（秒），用于目标1000Hz
        self.lsl_max_samples_per_pull = 1000  # 每次pull_chunk最多样本数，确保高频
        
        # 批量处理增强相关变量
        import queue
        self.biosig_batch_queue = queue.Queue(maxsize=100)  # 批量处理队列
        self.biosig_batch_queue_max_size = 100  # 队列最大大小
        self._last_timestamp_update_time = None  # 上次时间戳更新时间
        self._biosig_batch_processing_active = False  # 批量处理标志
        self._flush_in_progress = False  # 缓存刷新是否进行中
        self._biosig_stream_write_queue = queue.Queue(maxsize=100000)  # 90分钟 1000Hz: 需更大缓冲
        self._biosig_stream_writer_thread = None
        self._biosig_stream_writer_stop = False
        self._biosig_stream_writer_file = None
        self._biosig_stream_writer_csv = None
        self._driving_stream_write_queue = queue.Queue(maxsize=20000)
        self._driving_stream_writer_thread = None
        self._driving_stream_writer_stop = False
        self._driving_stream_writer_file = None
        self._driving_stream_writer_csv = None
        self._marker_stream_write_queue = queue.Queue(maxsize=5000)
        self._log_queue = queue.Queue(maxsize=300)  # 日志入队，由主线程定时排出，避免 after(0) 导致点击卡死
        self._marker_stream_writer_thread = None
        self._marker_stream_writer_stop = False
        self._marker_stream_writer_file = None
        self._marker_stream_writer_csv = None

        # 眼动模块配置与状态
        self.tobii_glasses_ip = "fe80::76fe:48ff:fe2e:24d8"
        self.tobii_my_ip = "fe80::74ac:3c53:94a0:8980"
        self.tobii_scope_id = 11
        self.tobii_connected = False
        self._eye_stream_write_queue = queue.Queue(maxsize=20000)
        self._eye_stream_writer_thread = None
        self._eye_stream_writer_stop = False
        self._eye_stream_writer_file = None
        self._eye_stream_writer_csv = None
        self.tobii_thread = None
        self._sync = DataSyncBuffer()
        self.tobii_sock = None
        self._last_tobii_ui_time_ms = 0
        self.tobii_ui_interval_ms = 2000  # 90分钟实验：眼动UI更新间隔从500ms增加到2000ms
        self.tobii_rate_count = 0
        self.tobii_last_rate_time = time.time()
        self.eye_latest = {}
        self._last_gaze_xy = None
        self._last_gaze_time = 0.0
        self._gp_count = 0
        self._pd_count = 0
        self._last_rate_emit = time.time()
        self._eye_structured_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "tobii_structured.jsonl")
        self._eye_structured_fp = None
        self.experiment_started = False
        self.experiment_ended = False
        self.experiment_start_epoch = None
        self.last_driving_epoch = None
        self.experiment_max_duration_s = 0
        self.experiment_inactive_timeout_s = 0
        self.stop_on_unity_end_only = True
        self._eye_video_enabled = False
        self._eye_video_pipe = None
        self._eye_video_textovl = None
        self._eye_video_sock = None
        self._eye_video_keepalive = None
        self._eye_video_buffersync = None
        self._eye_video_loop = None
        self._eye_video_thread = None
        self.ce_coupling_window_s = 10.0
        self.marker_match_window_s = 0.3

        # 美化GUI - 设置样式
        self.style = ttk.Style()
        self._setup_styles()
        
        # 创建UI组件
        self._create_ui()
        
        # 初始化日志
        self.log("系统初始化完成，点击'启动'按钮开始数据采集")
        
        # 移除自动启动系统的代码，改为手动点击启动
        # self.root.after(1000, self._start_system)
    
    def create_data_folder(self):
        """创建驾驶指标保存文件夹"""
        base_dir = "D:\\"
        self.data_folder = os.path.join(base_dir, "多模态驾驶指标保存")
        if not os.path.exists(self.data_folder):
            os.makedirs(self.data_folder)
            self.log(f"创建驾驶指标保存文件夹: {self.data_folder}")
        else:
            self.log(f"使用现有驾驶指标保存文件夹: {self.data_folder}")
    
    def _setup_styles(self):
        """设置UI样式以美化界面"""
        # 定义主题颜色
        bg_color = "#f0f0f0"
        frame_bg = "#ffffff"
        button_active = "#4a7a8c"
        button_inactive = "#6c757d"
        button_start = "#28a745"
        button_stop = "#dc3545"
        
        # 配置ttk主题
        self.style.configure("TFrame", background=frame_bg)
        self.style.configure("TButton", font=("微软雅黑", 12, "bold"), padding=10, foreground="black")
        self.style.configure("TLabel", font=("微软雅黑", 10), background=frame_bg)
        self.style.configure("TEntry", font=("微软雅黑", 10), padding=3)
        
        # 配置特定按钮样式
        self.style.configure("Start.TButton", foreground="black", background=button_start)
        self.style.configure("Stop.TButton", foreground="black", background=button_stop)
        self.style.configure("Normal.TButton", foreground="black", background=button_inactive, padding=12)
        self.style.configure("EyeCtrl.TButton", foreground="black", background=button_inactive, padding=6, font=("微软雅黑", 10, "bold"))
        
        # 悬停效果
        self.style.map("TButton", background=[("active", button_active)], foreground=[("active", "black")])
        self.style.map("Start.TButton", background=[("active", "#218838")], foreground=[("active", "black")])
        self.style.map("Stop.TButton", background=[("active", "#c82333")], foreground=[("active", "black")])
        
        # 配置Notebook标签页
        self.style.configure("TNotebook", background=bg_color)
        self.style.configure("TNotebook.Tab", font=("微软雅黑", 10, "bold"), padding=(10, 5))
        self.style.map("TNotebook.Tab", 
                      background=[("selected", frame_bg), ("!selected", "#e9ecef")],
                      foreground=[("selected", "#007bff"), ("!selected", "#495057")])
        
        # 设置主窗口背景
        self.root.configure(bg=bg_color)
    
    def _create_ui(self):
        """创建用户界面"""
        # 创建主框架
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 创建控制面板
        control_frame = ttk.LabelFrame(main_frame, text="控制面板", padding="10")
        control_frame.pack(fill=tk.X, pady=5)
        
        # 启动按钮
        self.start_button = ttk.Button(control_frame, text="启动", style="Start.TButton", command=self._start_system)
        self.start_button.pack(side=tk.LEFT, padx=5)
        
        # 停止按钮
        self.stop_button = ttk.Button(control_frame, text="停止", style="Stop.TButton", command=self._stop_system, state=tk.DISABLED)
        self.stop_button.pack(side=tk.LEFT, padx=5)
        
        # 文件名设置按钮
        self.filename_button = ttk.Button(control_frame, text="设置文件名", style="Normal.TButton", command=self._set_custom_filename)
        self.filename_button.pack(side=tk.LEFT, padx=5)
        
        # 状态标签
        self.status_var = tk.StringVar(value="状态: 待机")
        status_label = ttk.Label(control_frame, textvariable=self.status_var, font=("微软雅黑", 10, "bold"))
        status_label.pack(side=tk.LEFT, padx=20)
        
        # 当前文件名显示标签
        self.filename_var = tk.StringVar(value="当前文件: driving_simulator_sync_时间戳.csv")
        # 文件名更新功能测试行 - self.filename_var.set用于更新显示
        filename_label = ttk.Label(control_frame, textvariable=self.filename_var, font=("微软雅黑", 9, "italic"), foreground="#007bff")
        filename_label.pack(side=tk.LEFT, padx=20)
        
        # 可拖动分栏：左/右模块区域
        paned = ttk.Panedwindow(main_frame, orient=tk.HORIZONTAL)
        paned.pack(fill=tk.BOTH, expand=True, pady=3)
        pane_left = ttk.Frame(paned)
        pane_right = ttk.Frame(paned)
        paned.add(pane_left, weight=1)
        paned.add(pane_right, weight=1)

        # 第二行：状态与条件（置于左侧，可通过拖动分栏调整大小）
        control_row2 = ttk.Frame(pane_left)
        control_row2.pack(fill=tk.BOTH, expand=True)

        # 标记条件控制 (已移除手动按钮，标记由Unity LSL广播)
        # 恢复 marker_condition 定义以修复 AttributeError
        self.marker_condition = tk.IntVar(value=1)
        # ttk.Button(control_row2, text="实验条件1", style="Normal.TButton", command=lambda: self._set_marker_condition(1)).pack(side=tk.LEFT, padx=5)
        # ttk.Button(control_row2, text="实验条件2", style="Normal.TButton", command=lambda: self._set_marker_condition(2)).pack(side=tk.LEFT, padx=5)
        self.marker_status_var = tk.StringVar(value="标记条件: 由Unity控制")
        ttk.Label(control_row2, textvariable=self.marker_status_var, font=('微软雅黑', 10, 'italic')).pack(side=tk.LEFT, padx=10)
        
        # 时间同步状态标签
        self.time_sync_status_var = tk.StringVar(value="时间同步: 未激活")
        time_sync_label = ttk.Label(control_row2, textvariable=self.time_sync_status_var, font=('微软雅黑', 10, 'italic'))
        time_sync_label.pack(side=tk.LEFT, padx=10)
        
        # 时间精度差标签
        self.time_offset_var = tk.StringVar(value="精度差: -- ms")
        time_offset_label = ttk.Label(control_row2, textvariable=self.time_offset_var, font=('微软雅黑', 10, 'italic'))
        time_offset_label.pack(side=tk.LEFT, padx=10)

        # 刷新状态标签
        self.flush_status_var = tk.StringVar(value="刷新状态: 空闲")
        flush_status_label = ttk.Label(control_row2, textvariable=self.flush_status_var, font=('微软雅黑', 10, 'italic'))
        flush_status_label.pack(side=tk.LEFT, padx=10)

        # 缓存状态标签
        self.cache_status_var = tk.StringVar(value="缓存状态: 0 条记录")
        cache_status_label = ttk.Label(control_row2, textvariable=self.cache_status_var, font=('微软雅黑', 10, 'italic'))
        cache_status_label.pack(side=tk.LEFT, padx=10)
        # 右侧占位模块区域（可用于未来功能面板）
        ttk.Label(pane_right, text="模块区域（可拖动调整大小）", font=('微软雅黑', 10, 'italic')).pack(anchor=tk.NW, padx=10, pady=10)

        # 队列状态标签
        self.queue_status_var = tk.StringVar(value="队列状态: 驾驶Q=0, 写入Q(驾/生)=0/0")
        queue_status_label = ttk.Label(control_row2, textvariable=self.queue_status_var, font=('微软雅黑', 10, 'italic'))
        queue_status_label.pack(side=tk.LEFT, padx=10)

        # 写入线程状态标签
        self.writer_status_var = tk.StringVar(value="写入线程: 驾驶Alive=False, 生理Alive=False, 流式写入=False")
        writer_status_label = ttk.Label(control_row2, textvariable=self.writer_status_var, font=('微软雅黑', 10, 'italic'))
        writer_status_label.pack(side=tk.LEFT, padx=10)

        eye_net_frame = ttk.LabelFrame(main_frame, text="眼动网络配置", padding="10")
        eye_net_frame.pack(fill=tk.X, pady=5)
        eye_ip_row = ttk.Frame(eye_net_frame)
        eye_ip_row.pack(fill=tk.X)
        ttk.Label(eye_ip_row, text="GLASSES_IP:").pack(side=tk.LEFT, padx=5)
        self.tobii_glasses_ip_var = tk.StringVar(value=self.tobii_glasses_ip)
        ttk.Entry(eye_ip_row, textvariable=self.tobii_glasses_ip_var, width=30).pack(side=tk.LEFT, padx=5)
        ttk.Label(eye_ip_row, text="MY_IP:").pack(side=tk.LEFT, padx=5)
        self.tobii_my_ip_var = tk.StringVar(value=self.tobii_my_ip)
        ttk.Entry(eye_ip_row, textvariable=self.tobii_my_ip_var, width=30).pack(side=tk.LEFT, padx=5)
        ttk.Label(eye_ip_row, text="ScopeID:").pack(side=tk.LEFT, padx=5)
        self.tobii_scope_id_var = tk.StringVar(value=str(self.tobii_scope_id))
        ttk.Entry(eye_ip_row, textvariable=self.tobii_scope_id_var, width=10).pack(side=tk.LEFT, padx=5)
        eye_video_ctrl_row = ttk.Frame(eye_net_frame)
        eye_video_ctrl_row.pack(fill=tk.X, pady=2)
        ttk.Button(eye_video_ctrl_row, text="打开监控视频", style="EyeCtrl.TButton", command=self._launch_monitor_only).pack(side=tk.LEFT, padx=3)
        ttk.Button(eye_video_ctrl_row, text="关闭监控视频", style="EyeCtrl.TButton", command=self._close_monitor_only).pack(side=tk.LEFT, padx=3)
        ttk.Button(eye_video_ctrl_row, text="打开注视点视频", style="EyeCtrl.TButton", command=self._launch_v16_mapper).pack(side=tk.LEFT, padx=3)
        ttk.Button(eye_video_ctrl_row, text="关闭注视点视频", style="EyeCtrl.TButton", command=self._close_v16_mapper).pack(side=tk.LEFT, padx=3)
        # 删除监控直播开启/关闭按钮
        
        # 日志与下方模块采用垂直分栏，可拖动调整高度
        v_paned = ttk.Panedwindow(main_frame, orient=tk.VERTICAL)
        v_paned.pack(fill=tk.BOTH, expand=True, pady=5)
        log_frame = ttk.LabelFrame(v_paned, text="系统日志", padding="10")
        v_paned.add(log_frame, weight=1)
        
        # 使用水平分栏将日志区域分为左右两部分
        log_paned = ttk.Panedwindow(log_frame, orient=tk.HORIZONTAL)
        log_paned.pack(fill=tk.BOTH, expand=True)
        
        # 左侧：文本日志
        log_left_frame = ttk.Frame(log_paned)
        log_paned.add(log_left_frame, weight=1)
        
        self.log_text = scrolledtext.ScrolledText(log_left_frame, wrap=tk.WORD, width=40, height=8, font=('微软雅黑', 9))
        self.log_text.pack(fill=tk.BOTH, expand=True)
        self.log_text.yview_moveto(1.0)
        
        # 右侧：LSL信号绘图
        log_right_frame = ttk.Frame(log_paned)
        log_paned.add(log_right_frame, weight=1)
        
        # 控制面板
        control_frame = ttk.Frame(log_right_frame)
        control_frame.pack(side=tk.TOP, fill=tk.X)
        
        ttk.Label(control_frame, text="数据流:").pack(side=tk.LEFT, padx=2)
        self.stream_combo = ttk.Combobox(control_frame, state="readonly", width=25)
        self.stream_combo.pack(side=tk.LEFT, padx=2)
        self.stream_combo.set("扫描中...")
        self.stream_combo.bind("<<ComboboxSelected>>", self._on_stream_selected)

        ttk.Label(control_frame, text="通道:").pack(side=tk.LEFT, padx=2)
        self.channel_combo = ttk.Combobox(control_frame, state="readonly", width=20)
        self.channel_combo.pack(side=tk.LEFT, padx=2)
        self.channel_combo.set("等待连接...")
        self.channel_combo.bind("<<ComboboxSelected>>", self._on_channel_selected)
        
        try:
            self.lsl_plotter = LSLPlotter(log_right_frame, self.channel_combo, self.stream_combo)
        except Exception as e:
            print(f"Failed to initialize LSL Plotter: {e}")
            ttk.Label(log_right_frame, text=f"Plotter Error: {e}").pack()
        
        # 下方三个模块采用水平分栏，可拖动调整宽度
        bottom_paned = ttk.Panedwindow(v_paned, orient=tk.HORIZONTAL)
        v_paned.add(bottom_paned, weight=1)
        
        # 数据显示窗口 - 驾驶参数 - 增加高度
        data_frame = ttk.LabelFrame(bottom_paned, text="最新驾驶参数", padding="10")
        bottom_paned.add(data_frame, weight=1)
        
        self.data_text = scrolledtext.ScrolledText(data_frame, wrap=tk.WORD, width=40, height=15, font=('微软雅黑', 9))
        self.data_text.pack(fill=tk.BOTH, expand=True)
        self.data_text.yview_moveto(1.0)
        
        # 生物信号显示窗口 - 增加高度
        biosig_frame = ttk.LabelFrame(bottom_paned, text="生物信号传输状态", padding="10")
        bottom_paned.add(biosig_frame, weight=1)
        
        # 创建专门的心率和皮电显示框架
        vital_signs_frame = ttk.Frame(biosig_frame)
        vital_signs_frame.pack(fill=tk.X, pady=5)
        
        # 心率显示标签
        ttk.Label(vital_signs_frame, text="💓 实时心率:", font=('微软雅黑', 10, 'bold')).pack(side=tk.LEFT, padx=5)
        self.heart_rate_display_var = tk.StringVar(value="-- BPM")
        self.heart_rate_display = ttk.Label(vital_signs_frame, textvariable=self.heart_rate_display_var, font=('微软雅黑', 12, 'bold'), foreground='#FF5733')
        self.heart_rate_display.pack(side=tk.LEFT, padx=5)
        
        # 皮电显示标签
        ttk.Label(vital_signs_frame, text="⚡ 实时皮电:", font=('微软雅黑', 10, 'bold')).pack(side=tk.LEFT, padx=5)
        self.gsr_display_var = tk.StringVar(value="-- μS")
        self.gsr_display = ttk.Label(vital_signs_frame, textvariable=self.gsr_display_var, font=('微软雅黑', 12, 'bold'), foreground='#9B59B6')
        self.gsr_display.pack(side=tk.LEFT, padx=5)
        
        # 生物信号日志显示区域
        self.biosig_text = scrolledtext.ScrolledText(biosig_frame, wrap=tk.WORD, width=40, height=13, font=('微软雅黑', 9))
        self.biosig_text.pack(fill=tk.BOTH, expand=True)
        self.biosig_text.yview_moveto(1.0)
        eye_frame = ttk.LabelFrame(bottom_paned, text="眼动数据", padding="10")
        bottom_paned.add(eye_frame, weight=1)
        self.eye_gp_var = tk.StringVar(value="gp: --, --")
        self.eye_pd_var = tk.StringVar(value="pd: --")
        self.eye_s_var = tk.StringVar(value="质量: --")
        self.eye_eye_var = tk.StringVar(value="eye: --")
        self.eye_ts_var = tk.StringVar(value="ts/pts: -- / --")
        self.eye_rate_var = tk.StringVar(value="频率: -- Hz")
        eye_grid = ttk.Frame(eye_frame)
        eye_grid.pack(fill=tk.BOTH, expand=True)
        ttk.Label(eye_grid, textvariable=self.eye_gp_var, font=('微软雅黑', 10, 'bold')).grid(row=0, column=0, sticky=tk.W, padx=2, pady=2)
        ttk.Label(eye_grid, textvariable=self.eye_pd_var, font=('微软雅黑', 10)).grid(row=0, column=1, sticky=tk.W, padx=2, pady=2)
        ttk.Label(eye_grid, textvariable=self.eye_s_var, font=('微软雅黑', 10)).grid(row=0, column=2, sticky=tk.W, padx=2, pady=2)
        ttk.Label(eye_grid, textvariable=self.eye_eye_var, font=('微软雅黑', 10)).grid(row=1, column=0, sticky=tk.W, padx=2, pady=2)
        ttk.Label(eye_grid, textvariable=self.eye_ts_var, font=('微软雅黑', 10)).grid(row=1, column=1, sticky=tk.W, padx=2, pady=2)
        ttk.Label(eye_grid, textvariable=self.eye_rate_var, font=('微软雅黑', 10)).grid(row=1, column=2, sticky=tk.W, padx=2, pady=2)
        self.eye_text = scrolledtext.ScrolledText(eye_frame, wrap=tk.WORD, width=40, height=13, font=('微软雅黑', 9))
        self.eye_text.pack(fill=tk.BOTH, expand=True)
        self.eye_text.yview_moveto(1.0)
    def _generate_trial_filenames(self):
        """生成符合规范的Trial文件名，确保不覆盖"""
        # 确保数据文件夹存在
        if not os.path.exists(self.data_folder):
            os.makedirs(self.data_folder)
            
        base_ts = self.__class__.datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # 检查是否已存在同名文件，如果存在则添加后缀
        counter = 0
        while True:
            suffix = f"_{counter:03d}" if counter > 0 else ""
            # 检查主驾驶文件
            check_name = f"driving_{base_ts}{suffix}.csv"
            full_path = os.path.join(self.data_folder, check_name)
            if not os.path.exists(full_path):
                # 同时也检查其他模态文件是否存在，确保整体唯一
                bio_name = f"biosig_{base_ts}{suffix}.csv"
                eye_name = f"eye_{base_ts}{suffix}.csv"
                if not os.path.exists(os.path.join(self.data_folder, bio_name)) and \
                   not os.path.exists(os.path.join(self.data_folder, eye_name)):
                    final_ts = f"{base_ts}{suffix}"
                    break
            counter += 1
            
        # 设置各模态文件名 - 统一格式
        self.current_trial_filenames = {
            'driving': f"driving_{final_ts}.csv",
            'biosig': f"biosig_{final_ts}.csv",
            'eye': f"eye_{final_ts}.csv",
            'markers': f"markers_{final_ts}.csv",
            'event_summary': f"event_summary_{final_ts}.csv"
        }
        
        # 初始化文件写入状态
        self.trial_files_initialized = {
            'driving': False,
            'biosig': False,
            'eye': False,
            'markers': False
        }
        # 更新眼动CSV文件名
        self.eye_csv_filename = self.current_trial_filenames['eye']
        
        self.log(f"✓ Trial文件名已生成: {self.current_trial_filenames['driving']}")
        return True

    def _start_system(self):
        """启动数据接收和记录系统 - 优化版本，确保线程独立运行"""
        if self.running:
            return
        
        # 生成本次Trial的文件名
        self._generate_trial_filenames()
        
        # 更新状态
        self.running = True
        self.start_button.config(state=tk.DISABLED)
        self.stop_button.config(state=tk.NORMAL)
        self.status_var.set("状态: 运行中 (流式写入模式)")
        self.time_sync_status_var = tk.StringVar(value="时间同步: 等待数据...")
        self.time_offset_var = tk.StringVar(value="精度差: -- ms")
        
        # 记录系统启动时间和缓存开始时间
        self.system_start_time = time.time()
        self.cache_start_time = self.__class__.datetime.now()
        self.session_timestamp = self.__class__.datetime.now().strftime("%Y%m%d_%H%M%S")
        self.actual_start_time = None  # 重置实际开始时间
        self._marker_event_queue = deque()
        self.streaming_write_enabled = True
        self.experiment_started = False
        self.experiment_ended = False
        self.experiment_start_epoch = None
        self.last_driving_epoch = None
        
        # 清空当前驾驶参数和数据缓存
        with self.driving_lock:
            self.current_driving_params = {}
            self.last_driving_params = {}
        with self.biosig_lock:
            self.current_biosig_params = {}
            self._biosig_sample_count = 0
            self.biosig_connected = False
        with self.cache_lock:
            self.driving_data_cache = []  # 清空之前的驾驶数据缓存
            self.biosig_data_cache = []   # 清空之前的生物信号数据缓存
        
        # 重置时间同步相关变量
        self.time_sync_enabled = False
        self.time_offset = 0.0
        self.sync_samples_collected = 0
        self.sync_offsets = []
        self.driving_received = False
        self.biosig_received = False
        
        # 启动UDP接收线程
        # 启动驾驶数据LSL接收线程 (替换原有的UDP接收)
        self.udp_thread = threading.Thread(target=self._receive_driving_lsl_data, daemon=True)
        self.udp_thread.daemon = True
        self.udp_thread.start()
        
        # 启动LSL接收线程
        if PYLSL_AVAILABLE:
            self.lsl_thread = threading.Thread(target=self._receive_lsl_data, daemon=True)
            self.lsl_thread.daemon = True
            self.lsl_thread.start()
            
            # 启动标记数据LSL接收线程
            self.marker_thread = threading.Thread(target=self._receive_marker_lsl_data, daemon=True)
            self.marker_thread.daemon = True
            self.marker_thread.start()
        
        # 启动缓存状态自动更新
        self._start_cache_monitoring()
        
        # 驾驶参数 UI 由主线程定时刷新，避免 LSL 线程频繁 after(0) 导致点击窗口卡死
        self._schedule_driving_display_tick()
        # 每 1 秒仅更新“驾驶数据/生物信号条数”与时间，使 GUI 尽快显示与 UXF 同源的数据量
        self._schedule_lightweight_status_tick()
        
        # 启动批量处理定时器 - 使用after方法在UI线程中启动，确保1000Hz采样率稳定
        self.root.after(0, self._biosig_batch_processing)
        # 启动采样率检查
        self.root.after(1000, self._schedule_sampling_rate_check)
        # 启动生理流式写入线程
        try:
            self._biosig_stream_writer_stop = False
            if not self._biosig_stream_writer_thread or not self._biosig_stream_writer_thread.is_alive():
                self._biosig_stream_writer_thread = threading.Thread(target=self._biosig_stream_writer, daemon=True)
                self._biosig_stream_writer_thread.start()
        except Exception as e:
            self.log(f"⚠ 启动生理写入线程失败: {str(e)}")
        # 启动驾驶流式写入线程
        try:
            self._driving_stream_writer_stop = False
            if not self._driving_stream_writer_thread or not self._driving_stream_writer_thread.is_alive():
                self._driving_stream_writer_thread = threading.Thread(target=self._driving_stream_writer, daemon=True)
                self._driving_stream_writer_thread.start()
        except Exception as e:
            self.log(f"⚠ 启动驾驶写入线程失败: {str(e)}")
        try:
            self._marker_stream_writer_stop = False
            if not self._marker_stream_writer_thread or not self._marker_stream_writer_thread.is_alive():
                self._marker_stream_writer_thread = threading.Thread(target=self._marker_stream_writer, daemon=True)
                self._marker_stream_writer_thread.start()
        except Exception as e:
            self.log(f"⚠ 启动标记写入线程失败: {str(e)}")
        
        self.log("✓ 系统已启动，处于数据缓存模式 - 所有数据将在停止时写入CSV")
        self.log("✓ 驾驶数据和生物信号数据各自独立采集，互不干扰")
        self.log("ℹ 等待两种数据都收到后，将自动激活LSL时间同步功能")
        self._init_marker_sets()
        try:
            self._start_tobii_unicast()
        except Exception:
            pass
        try:
            self._start_eye_video_overlay()
        except Exception:
            pass
        try:
            self.root.after(1000, self._monitor_experiment_status)
        except Exception:
            pass

    def _emit_biosig_marker(self, ev: dict):
        try:
            sys_ts = ev.get('event_system_timestamp', 'N/A')
            sync_ts = ev.get('event_synchronized_timestamp', 'N/A')
            heart = 'N/A'
            gsr = 'N/A'
            bio_ts = 'N/A'
            with self.biosig_lock:
                if hasattr(self, 'current_biosig_params') and self.current_biosig_params:
                    heart = self.current_biosig_params.get('heart_rate_bpm', 'N/A')
                    gsr = self.current_biosig_params.get('gsr_uS', 'N/A')
            if hasattr(self, 'biosig_last_timestamp') and self.biosig_last_timestamp:
                try:
                    bio_ts = str(float(self.biosig_last_timestamp))
                except Exception:
                    bio_ts = 'N/A'
            row = {
                'system_timestamp': sys_ts,
                'synchronized_timestamp': sync_ts if self.time_sync_enabled else 'N/A',
                'event_system_timestamp': sys_ts,
                'event_synchronized_timestamp': sync_ts if self.time_sync_enabled else 'N/A',
                'system_lsl_time': 'N/A',
                'biosig_timestamp': bio_ts,
                'time_diff_lsl': 'N/A',
                'heart_rate_bpm': heart if heart is not None else 'N/A',
                'gsr_uS': gsr if gsr is not None else 'N/A',
                'marker_flag': True,
                'marker_type': ev.get('type', 'N/A'),
                'marker_condition': ev.get('cond', 'N/A'),
                'marker_target': ev.get('target', 'N/A'),
                'marker_color': ev.get('color', 'N/A'),
                'marker_label': ev.get('label', 'N/A')
            }
            
            # 独立队列插入，防止高并发下被阻塞
            try:
                self._biosig_stream_write_queue.put_nowait(row)
            except queue.Full:
                # 队列满时尝试丢弃旧数据并重试
                try:
                    _ = self._biosig_stream_write_queue.get_nowait()
                except Exception:
                    pass
                try:
                    self._biosig_stream_write_queue.put_nowait(row)
                except Exception:
                    pass
            
            try:
                self.log(f"✓ 实时同步生理标记 type={row['marker_type']} label={row['marker_label']} ts={row['system_timestamp']}")
            except Exception:
                pass
            try:
                self._broadcast_marker_udp(ev)
            except Exception:
                pass
        except Exception:
            try:
                self.log("⚠ 生理标记实时同步失败")
            except Exception:
                pass

    def _emit_driving_marker(self, ev: dict):
        try:
            sys_ts = ev.get('event_system_timestamp', self.__class__.datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3])
            sync_ts = ev.get('event_synchronized_timestamp', 'N/A')
            row = {
                "system_timestamp": sys_ts,
                "synchronized_timestamp": sync_ts if self.time_sync_enabled else "N/A",
                "lsl_timestamp": ev.get('lsl_timestamp', 'N/A'),
                "marker_flag": True,
                "marker_type": ev.get('type', 'N/A'),
                "marker_condition": ev.get('cond', 'N/A'),
                "marker_target": ev.get('target', 'N/A'),
                "marker_color": ev.get('color', 'N/A'),
                "marker_label": ev.get('label', 'N/A')
            }
            try:
                with self.driving_lock:
                    snap = self.current_driving_params.copy() if getattr(self, 'current_driving_params', None) else {}
                for k, v in snap.items():
                    if k not in row:
                        row[k] = v
            except Exception:
                pass

            try:
                self._driving_stream_write_queue.put_nowait(row.copy())
            except queue.Full:
                try:
                    self._driving_queue_dropped = getattr(self, '_driving_queue_dropped', 0) + 1
                except Exception:
                    pass
                try:
                    _ = self._driving_stream_write_queue.get_nowait()
                except Exception:
                    pass
                try:
                    self._driving_stream_write_queue.put_nowait(row.copy())
                except Exception:
                    pass
            except Exception:
                pass

            try:
                with self.cache_lock:
                    self.driving_data_cache.append(row.copy())
            except Exception:
                pass
        except Exception:
            pass
    
    def _stop_system(self):
        """停止数据接收和记录系统 - 优化版本，确保安全停止和数据保存"""
        if not self.running:
            return
        
        # 更新状态
        self.running = False
        self.status_var.set("状态: 正在停止...")
        if hasattr(self, 'time_sync_status_var'):
            self.time_sync_status_var.set("时间同步: 已停止")
        if hasattr(self, 'time_offset_var'):
            self.time_offset_var.set("精度差: -- ms")
        
        try:
            if hasattr(self, 'udp_thread') and self.udp_thread and self.udp_thread.is_alive():
                self.udp_thread.join(timeout=3.0)
        except Exception:
            pass
        try:
            if hasattr(self, 'lsl_thread') and self.lsl_thread and self.lsl_thread.is_alive():
                self.lsl_thread.join(timeout=3.0)
        except Exception:
            pass
        try:
            if hasattr(self, 'marker_thread') and self.marker_thread and self.marker_thread.is_alive():
                self.marker_thread.join(timeout=3.0)
        except Exception:
            pass

        try:
            while hasattr(self, 'biosig_batch_queue') and not self.biosig_batch_queue.empty():
                self.biosig_batch_queue.get_nowait()
        except Exception:
            pass

        try:
            t0 = time.time()
            while hasattr(self, '_biosig_stream_write_queue') and self._biosig_stream_write_queue.qsize() > 0 and time.time() - t0 < 10.0:
                time.sleep(0.05)
            self._biosig_stream_writer_stop = True
            if self._biosig_stream_writer_thread and self._biosig_stream_writer_thread.is_alive():
                self._biosig_stream_writer_thread.join(timeout=5.0)
        except Exception:
            pass
        try:
            if self._biosig_stream_writer_file:
                self._biosig_stream_writer_file.close()
        except Exception:
            pass
        self._biosig_stream_writer_file = None
        self._biosig_stream_writer_csv = None

        try:
            t0 = time.time()
            while hasattr(self, '_driving_stream_write_queue') and self._driving_stream_write_queue.qsize() > 0 and time.time() - t0 < 30.0:
                time.sleep(0.05)
            self._driving_stream_writer_stop = True
            if self._driving_stream_writer_thread and self._driving_stream_writer_thread.is_alive():
                self._driving_stream_writer_thread.join(timeout=10.0)
        except Exception:
            pass
        try:
            if self._driving_stream_writer_file:
                self._driving_stream_writer_file.close()
        except Exception:
            pass
        self._driving_stream_writer_file = None
        self._driving_stream_writer_csv = None
        try:
            t0 = time.time()
            while hasattr(self, '_marker_stream_write_queue') and self._marker_stream_write_queue.qsize() > 0 and time.time() - t0 < 10.0:
                time.sleep(0.05)
            self._marker_stream_writer_stop = True
            if self._marker_stream_writer_thread and self._marker_stream_writer_thread.is_alive():
                self._marker_stream_writer_thread.join(timeout=5.0)
        except Exception:
            pass
        try:
            if self._marker_stream_writer_file:
                self._marker_stream_writer_file.close()
        except Exception:
            pass
        self._marker_stream_writer_file = None
        self._marker_stream_writer_csv = None

        try:
            if self.udp_socket:
                self.udp_socket.close()
        except Exception:
            pass
        self.udp_socket = None

        try:
            if self.proc_v16:
                self.proc_v16.terminate()
                self.proc_v16 = None
        except Exception:
            pass

        try:
            if self.csv_file:
                self.csv_file.close()
                self.csv_file = None
                self.csv_writer = None
        except Exception:
            pass

        try:
            self._stop_tobii_unicast()
        except Exception:
            pass
        try:
            self._stop_eye_video_overlay()
        except Exception:
            pass

        try:
            self._generate_event_summary()
        except Exception as e:
            try:
                self.log(f"⚠ 生成事件统计报告失败: {str(e)}")
            except Exception:
                pass

        try:
            self.start_button.config(state=tk.NORMAL)
            self.stop_button.config(state=tk.DISABLED)
        except Exception:
            pass
        try:
            self.status_var.set("状态: 已停止")
        except Exception:
            pass

        self.log("✓ 系统已停止")

    def _generate_event_summary(self):
        """生成事件统计报告"""
        if not hasattr(self, 'current_trial_filenames'):
            return
        markers_file = self.current_trial_filenames.get('markers')
        driving_file = self.current_trial_filenames.get('driving')
        file_path = os.path.join(self.data_folder, markers_file) if markers_file else None
        source_name = markers_file
        if not file_path or not os.path.exists(file_path):
            if not driving_file:
                return
            file_path = os.path.join(self.data_folder, driving_file)
            source_name = driving_file
            if not os.path.exists(file_path):
                self.log("⚠ 无法生成事件报告：数据文件不存在")
                return
            
        self.log("📊 正在生成事件统计报告...")
        
        event_counts = collections.defaultdict(int)
        events_details = []
        total_rows = 0
        first_dt = None
        last_dt = None
        max_gap_s = 0.0
        gap_over_200ms = 0
        
        try:
            # 确保文件写入已完成
            time.sleep(0.5)
            
            with open(file_path, 'r', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    total_rows += 1
                    try:
                        ts_str = row.get('system_timestamp', '')
                        if ts_str:
                            try:
                                dt = datetime.strptime(ts_str, "%Y-%m-%d %H:%M:%S.%f")
                            except Exception:
                                dt = datetime.strptime(ts_str, "%Y-%m-%d %H:%M:%S")
                            if first_dt is None:
                                first_dt = dt
                            if last_dt is not None:
                                gap = (dt - last_dt).total_seconds()
                                if gap > max_gap_s:
                                    max_gap_s = gap
                                if gap >= 0.2:
                                    gap_over_200ms += 1
                            last_dt = dt
                    except Exception:
                        pass
                    # 检查marker_flag，处理可能的布尔值字符串
                    flag = str(row.get('marker_flag', '')).lower()
                    if flag in ('true', '1', 'yes'):
                        marker_type = row.get('marker_type', 'N/A')
                        marker_cond = row.get('marker_condition', 'N/A')
                        marker_target = row.get('marker_target', 'N/A')
                        marker_label = row.get('marker_label', 'N/A')
                        
                        key = (marker_type, marker_cond, marker_target, marker_label)
                        event_counts[key] += 1
                        
                        events_details.append({
                            'timestamp': row.get('system_timestamp', 'N/A'),
                            'type': marker_type,
                            'condition': marker_cond,
                            'target': marker_target,
                            'label': marker_label
                        })
            
            # 写入统计报告
            summary_file = self.current_trial_filenames['event_summary']
            summary_path = os.path.join(self.data_folder, summary_file)
            
            with open(summary_path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                writer.writerow(['事件统计报告'])
                writer.writerow(['源文件', source_name])
                writer.writerow(['生成时间', self.__class__.datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
                writer.writerow([])

                duration_s = None
                est_rate = None
                try:
                    if first_dt is not None and last_dt is not None and last_dt >= first_dt:
                        duration_s = (last_dt - first_dt).total_seconds()
                        if duration_s > 0 and total_rows > 0:
                            est_rate = float(total_rows) / duration_s
                except Exception:
                    duration_s = None
                    est_rate = None

                writer.writerow(['数据完整性'])
                writer.writerow(['总行数', total_rows])
                writer.writerow(['起始时间', first_dt.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3] if first_dt else 'N/A'])
                writer.writerow(['结束时间', last_dt.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3] if last_dt else 'N/A'])
                writer.writerow(['持续秒数', f"{duration_s:.3f}" if duration_s is not None else 'N/A'])
                writer.writerow(['估计采样率Hz', f"{est_rate:.3f}" if est_rate is not None else 'N/A'])
                writer.writerow(['最大间隔秒', f"{max_gap_s:.6f}"])
                writer.writerow(['间隔>=200ms次数', gap_over_200ms])
                writer.writerow(['写入行数(线程计数)', getattr(self, '_driving_rows_written', 0)])
                writer.writerow(['队列丢弃次数', getattr(self, '_driving_queue_dropped', 0)])
                writer.writerow([])
                
                writer.writerow(['汇总统计'])
                writer.writerow(['Type', 'Condition', 'Target', 'Label', 'Count'])
                for key, count in sorted(event_counts.items()):
                    writer.writerow([*key, count])
                
                writer.writerow([])
                writer.writerow(['事件明细'])
                writer.writerow(['Timestamp', 'Type', 'Condition', 'Target', 'Label'])
                for ev in events_details:
                    writer.writerow([ev['timestamp'], ev['type'], ev['condition'], ev['target'], ev['label']])
                    
            self.log(f"✓ 事件统计报告已生成: {summary_file}")
            
        except Exception as e:
            self.log(f"⚠ 分析事件数据时出错: {str(e)}")
            # 不抛出异常，以免影响停止流程
        return

    def _open_eye_csv(self):
        if not hasattr(self, 'eye_csv_filename') or not self.eye_csv_filename:
            try:
                self.eye_csv_filename = f"eyedata_{self.session_timestamp}.csv"
            except Exception:
                self.eye_csv_filename = "eyedata.csv"
        p = os.path.join(self.data_folder, self.eye_csv_filename)
        # 如果文件已存在，则追加写入；如果不存在，则创建新文件并写入表头
        created = not os.path.exists(p)
        self._eye_stream_writer_file = open(p, 'a', newline='', encoding='utf-8-sig')
        self._eye_stream_writer_csv = csv.DictWriter(self._eye_stream_writer_file, fieldnames=[
            'system_timestamp','synchronized_timestamp','glassts',
            'gp_x','gp_y','pd',
            'gp3_x','gp3_y','gp3_z',
            'gd_x','gd_y','gd_z',
            'pc_x','pc_y','pc_z',
            'gy_x','gy_y','gy_z',
            'ac_x','ac_y','ac_z',
            'eye','glasses_ts','glasses_pts','s','l','gidx',
            'marker_flag','marker_type','marker_condition','marker_target','marker_color','marker_label',
            'Gaze_TS','Gaze_X','Gaze_Y'
        ])
        if created:
            try:
                self._eye_stream_writer_csv.writeheader()
            except Exception:
                pass

    def _eye_stream_writer(self):
        try:
            self._open_eye_csv()
        except Exception:
            return
        while not self._eye_stream_writer_stop:
            try:
                row = self._eye_stream_write_queue.get(timeout=0.5)
            except Exception:
                row = None
            if row:
                try:
                    self._eye_stream_writer_csv.writerow(row)
                    self._eye_stream_writer_file.flush()
                except Exception:
                    pass
        try:
            if self._eye_stream_writer_file:
                self._eye_stream_writer_file.close()
        except Exception:
            pass
        self._eye_stream_writer_file = None
        self._eye_stream_writer_csv = None

    def _start_tobii_unicast(self):
        if self.tobii_thread and self.tobii_thread.is_alive():
            return
        try:
            self.tobii_glasses_ip = self.tobii_glasses_ip_var.get().strip() if hasattr(self, 'tobii_glasses_ip_var') else self.tobii_glasses_ip
            self.tobii_my_ip = self.tobii_my_ip_var.get().strip() if hasattr(self, 'tobii_my_ip_var') else self.tobii_my_ip
            self.tobii_scope_id = int(self.tobii_scope_id_var.get().strip()) if hasattr(self, 'tobii_scope_id_var') else self.tobii_scope_id
        except Exception:
            pass
        try:
            self.eye_csv_filename = f"eyedata_{self.session_timestamp}.csv"
        except Exception:
            self.eye_csv_filename = "eyedata.csv"
        self._eye_stream_writer_stop = False
        if not self._eye_stream_writer_thread or not self._eye_stream_writer_thread.is_alive():
            self._eye_stream_writer_thread = threading.Thread(target=self._eye_stream_writer, daemon=True)
            self._eye_stream_writer_thread.start()
        self.tobii_connected = False
        self.tobii_thread = threading.Thread(target=self._tobii_unicast_loop, daemon=True)
        self.tobii_thread.start()

    def _open_eye_structured(self):
        try:
            if not self._eye_structured_fp:
                self._eye_structured_fp = open(self._eye_structured_path, 'a', encoding='utf-8')
        except Exception:
            self._eye_structured_fp = None

    def _write_structured_json(self, j):
        try:
            if not self._eye_structured_fp:
                self._open_eye_structured()
            if not self._eye_structured_fp:
                return
            recs = {
                "同步与质量": {
                    "ts": j.get("ts"),
                    "s": j.get("s"),
                    "l": j.get("l"),
                    "pts": j.get("pts"),
                    "gidx": j.get("gidx")
                },
                "核心指标": {
                    "gp": j.get("gp"),
                    "pd": j.get("pd")
                },
                "空间与几何": {
                    "gp3": j.get("gp3"),
                    "gd": j.get("gd"),
                    "pc": j.get("pc")
                },
                "运动传感": {
                    "gy": j.get("gy"),
                    "ac": j.get("ac")
                },
                "其他": {
                    "eye": j.get("eye")
                }
            }
            self._eye_structured_fp.write(json.dumps(recs, ensure_ascii=False) + "\n")
            self._eye_structured_fp.flush()
        except Exception:
            pass

    def _emit_eye_marker(self, ev: dict):
        try:
            sys_ts = ev.get('event_system_timestamp', datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3])
            sync_ts = ev.get('event_synchronized_timestamp', 'N/A')
            j = getattr(self, 'eye_latest', {}) or {}
            gp = j.get('gp')
            pdv = j.get('pd')
            def arr3(a):
                if isinstance(a, (list, tuple)):
                    return (a[0] if len(a)>0 else 'N/A', a[1] if len(a)>1 else 'N/A', a[2] if len(a)>2 else 'N/A')
                return ('N/A','N/A','N/A')
            gp3 = arr3(j.get('gp3'))
            gd = arr3(j.get('gd'))
            pc = arr3(j.get('pc'))
            gy = arr3(j.get('gy'))
            ac = arr3(j.get('ac'))
            row = {
                'system_timestamp': sys_ts,
                'synchronized_timestamp': sync_ts,
                'gp_x': gp[0] if isinstance(gp, (list, tuple)) and len(gp)>0 else 'N/A',
                'gp_y': gp[1] if isinstance(gp, (list, tuple)) and len(gp)>1 else 'N/A',
                'pd': pdv if pdv is not None else 'N/A',
                'gp3_x': gp3[0],'gp3_y': gp3[1],'gp3_z': gp3[2],
                'gd_x': gd[0],'gd_y': gd[1],'gd_z': gd[2],
                'pc_x': pc[0],'pc_y': pc[1],'pc_z': pc[2],
                'gy_x': gy[0],'gy_y': gy[1],'gy_z': gy[2],
                'ac_x': ac[0],'ac_y': ac[1],'ac_z': ac[2],
                'eye': j.get('eye','N/A'),
                'glasses_ts': j.get('ts'),
                'glasses_pts': j.get('pts'),
                's': j.get('s'),
                'l': j.get('l'),
                'gidx': j.get('gidx'),
                'marker_flag': True,
                'marker_type': ev.get('type', 'N/A'),
                'marker_condition': ev.get('cond', 'N/A'),
                'marker_target': ev.get('target', 'N/A'),
                'marker_color': ev.get('color', 'N/A'),
                'marker_label': ev.get('label', 'N/A'),
                'Gaze_TS': j.get('ts', ''),
                'Gaze_X': (gp[0] if isinstance(gp, (list, tuple)) and len(gp)>0 else (self._last_gaze_xy[0] if self._last_gaze_xy and (time.time()-self._last_gaze_time)<=1.0 else '')),
                'Gaze_Y': (gp[1] if isinstance(gp, (list, tuple)) and len(gp)>1 else (self._last_gaze_xy[1] if self._last_gaze_xy and (time.time()-self._last_gaze_time)<=1.0 else ''))
            }
            try:
                self._eye_stream_write_queue.put_nowait(row)
            except Exception:
                try:
                    _ = self._eye_stream_write_queue.get_nowait()
                except Exception:
                    pass
                try:
                    self._eye_stream_write_queue.put_nowait(row)
                except Exception:
                    pass
            try:
                self._broadcast_marker_udp(ev)
            except Exception:
                pass
        except Exception:
            pass

    def _broadcast_marker_udp(self, ev: dict):
        try:
            if not getattr(self, 'marker_udp_enable', False):
                return
            msg = json.dumps({'type': ev.get('type'), 'cond': ev.get('cond'), 'target': ev.get('target'), 'color': ev.get('color'), 'label': ev.get('label'), 'event_system_timestamp': ev.get('event_system_timestamp'), 'event_synchronized_timestamp': ev.get('event_synchronized_timestamp')}, ensure_ascii=False).encode('utf-8')
            s = socket.socket(socket.AF_INET6, socket.SOCK_DGRAM)
            s.settimeout(0.2)
            s.sendto(msg, (self.marker_udp_host, int(self.marker_udp_port), 0, 0))
            s.close()
        except Exception:
            pass

    def _emit_experiment_marker(self, kind: str):
        try:
            st = datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
            try:
                _, sync_ts, _ = self._get_synchronized_timestamp(st)
            except Exception:
                sync_ts = 'N/A'
            ev = {
                'type': 'experiment',
                'cond': kind,
                'target': 'global',
                'color': 'green' if kind == 'start' else 'red',
                'label': 'EXPERIMENT_START' if kind == 'start' else 'EXPERIMENT_END',
                'event_system_timestamp': st,
                'event_synchronized_timestamp': sync_ts
            }
            try:
                self._emit_biosig_marker(ev)
            except Exception:
                pass
            try:
                self._emit_eye_marker(ev)
            except Exception:
                pass
            try:
                self._emit_driving_marker(ev)
            except Exception:
                pass
            try:
                marker_row = {
                    "system_timestamp": st,
                    "synchronized_timestamp": sync_ts,
                    "lsl_timestamp": "N/A",
                    "marker_flag": True,
                    "marker_type": ev.get('type', 'experiment'),
                    "marker_label": ev.get('label', 'N/A'),
                    "marker_condition": ev.get('cond', kind),
                    "marker_target": ev.get('target', 'global'),
                    "marker_color": ev.get('color', 'N/A'),
                    "marker_stream_name": "internal"
                }
                self._marker_stream_write_queue.put_nowait(marker_row)
            except Exception:
                try:
                    self._marker_queue_dropped = getattr(self, '_marker_queue_dropped', 0) + 1
                except Exception:
                    pass
        except Exception:
            pass

    def _is_experiment_end_marker(self, marker_text):
        try:
            text = str(marker_text)
        except Exception:
            return False
        if not text:
            return False
        return ("UXF_RunEnd" in text) or ("EXPERIMENT_END" in text) or ("UXF_SessionEnd" in text)

    def _monitor_experiment_status(self):
        try:
            now = time.time()
            if self.driving_received and not self.experiment_started:
                self.experiment_started = True
                self.experiment_start_epoch = now
                try:
                    self._emit_experiment_marker('start')
                except Exception:
                    pass
            if self.experiment_started and not self.experiment_ended and not getattr(self, 'stop_on_unity_end_only', False):
                inactive = False
                try:
                    if self.last_driving_epoch and (now - self.last_driving_epoch) >= self.experiment_inactive_timeout_s:
                        inactive = True
                except Exception:
                    inactive = False
                timedout = False
                try:
                    if self.experiment_start_epoch and (now - self.experiment_start_epoch) >= self.experiment_max_duration_s:
                        timedout = True
                except Exception:
                    timedout = False
                if inactive or timedout:
                    self.experiment_ended = True
                    try:
                        self._emit_experiment_marker('end')
                    except Exception:
                        pass
            try:
                self.root.after(1000, self._monitor_experiment_status)
            except Exception:
                pass
        except Exception:
            try:
                self.root.after(1000, self._monitor_experiment_status)
            except Exception:
                pass

    def _stop_tobii_unicast(self):
        self._eye_stream_writer_stop = True
        if self._eye_stream_writer_thread and self._eye_stream_writer_thread.is_alive():
            try:
                self._eye_stream_writer_thread.join(timeout=1.0)
            except Exception:
                pass
        if self.tobii_sock:
            try:
                self.tobii_sock.close()
            except Exception:
                pass
            self.tobii_sock = None
        if self.tobii_thread and self.tobii_thread.is_alive():
            try:
                self.tobii_thread.join(timeout=1.0)
            except Exception:
                pass
        self.tobii_connected = False

    def _update_eye_ui(self, rec):
        now_ms = int(time.time()*1000)
        if now_ms - self._last_tobii_ui_time_ms < self.tobii_ui_interval_ms:
            return
        self._last_tobii_ui_time_ms = now_ms
        
        def update_ui_safe():
            gp = rec.get('gp')
            pdv = rec.get('pd')
            s = rec.get('s')
            eye = rec.get('eye')
            ts = rec.get('ts')
            pts = rec.get('pts')
            if isinstance(gp, (list, tuple)) and len(gp) >= 2:
                self.eye_gp_var.set(f"gp: {gp[0]}, {gp[1]}")
            else:
                self.eye_gp_var.set("gp: --, --")
            self.eye_pd_var.set(f"pd: {pdv}" if pdv is not None else "pd: --")
            self.eye_s_var.set(f"质量: {s}" if s is not None else "质量: --")
            self.eye_eye_var.set(f"eye: {eye}" if eye is not None else "eye: --")
            self.eye_ts_var.set(f"ts/pts: {ts} / {pts}" if ts is not None else "ts/pts: -- / --")
            try:
                # 90分钟实验：大幅减少眼动数据文本显示
                # 只显示关键信息，而不是完整的JSON数据
                key_info = f"眼动: gp={rec.get('gp', '--')}, pd={rec.get('pd', '--')}, s={rec.get('s', '--')}"
                self.eye_text.insert(tk.END, key_info + "\n")
                self.eye_text.see(tk.END)
                
                # 更激进的清理：超过50行就删除一半
                lines = int(self.eye_text.index('end-1c').split('.')[0])
                if lines > 50:
                    self.eye_text.delete('1.0', f'{lines//2}.0')
            except Exception:
                pass
        
        # 使用 root.after 确保在主线程更新 UI
        if self.root:
            self.root.after(0, update_ui_safe)

    def _tobii_unicast_loop(self):
        try:
            sock = socket.socket(socket.AF_INET6, socket.SOCK_DGRAM)
            try:
                sock.bind((self.tobii_my_ip, 0, 0, int(self.tobii_scope_id)))
            except Exception:
                return
            self.tobii_sock = sock
            target = (self.tobii_glasses_ip, 49152, 0, int(self.tobii_scope_id))
            try:
                sock.setsockopt(socket.SOL_SOCKET, socket.SO_RCVBUF, 4 * 1024 * 1024)
            except Exception:
                pass
            req = {"type": "live.data.unicast", "key": "manual_test", "op": "start"}
            msg = json.dumps(req).encode('utf-8')
            self.tobii_connected = True
            last_print = time.time()
            last_ka = 0.0
            sock.settimeout(0.02)
            while self.running and self.tobii_connected:
                try:
                    if time.time() - last_ka >= 1.0:
                        try:
                            sock.sendto(msg, target)
                        except Exception:
                            break
                        last_ka = time.time()
                    data, addr = sock.recvfrom(4096)
                    text = data.decode('utf-8', errors='ignore')
                    if 'status' in text:
                        continue
                    try:
                        j = json.loads(text)
                        if isinstance(j, dict) and ('gp' in j or 'pd' in j or 's' in j):
                            pts = j.get('pts')
                            gaze_ts = j.get('ts')
                            gp = j.get('gp')
                            gaze_x = 'N/A'
                            gaze_y = 'N/A'
                            if gp and len(gp) >= 2:
                                gaze_x = gp[0]
                                gaze_y = gp[1]
                                try:
                                    self._last_gaze_xy = [gaze_x, gaze_y]
                                    self._last_gaze_time = time.time()
                                except Exception:
                                    pass
                            if pts is not None and gaze_ts is not None:
                                self._sync.add_gaze_data(pts, gaze_ts, gaze_x, gaze_y)
                    except Exception:
                        continue
                    if isinstance(j, dict):
                        self.eye_latest = j
                        self.tobii_rate_count += 1
                        if isinstance(j.get('gp'), (list, tuple)):
                            self._gp_count += 1
                        if j.get('pd') is not None:
                            self._pd_count += 1
                        try:
                            if self._eye_video_buffersync:
                                self._eye_video_buffersync.add_et(j)
                        except Exception:
                            pass
                        if time.time() - last_print >= 1.0:
                            try:
                                self.eye_rate_var.set(f"频率: gp={self._gp_count} Hz, pd={self._pd_count} Hz")
                            except Exception:
                                pass
                            self.tobii_rate_count = 0
                            self._gp_count = 0
                            self._pd_count = 0
                            last_print = time.time()
                        self._update_eye_ui(j)
                        try:
                            self._write_structured_json(j)
                        except Exception:
                            pass
                        st = datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
                        try:
                            _, sync_ts, _ = self._get_synchronized_timestamp(st)
                        except Exception:
                            sync_ts = 'N/A'
                        gp = j.get('gp')
                        pdv = j.get('pd')
                        def get3(a):
                            if isinstance(a, (list, tuple)):
                                return (a[0] if len(a)>0 else 'N/A', a[1] if len(a)>1 else 'N/A', a[2] if len(a)>2 else 'N/A')
                            return ('N/A','N/A','N/A')
                        gp3 = get3(j.get('gp3'))
                        gd = get3(j.get('gd'))
                        pc = get3(j.get('pc'))
                        gy = get3(j.get('gy'))
                        ac = get3(j.get('ac'))
                        row = {
                            'system_timestamp': st,
                            'synchronized_timestamp': sync_ts,
                            'glassts': j.get('pts'),
                            'gp_x': gp[0] if isinstance(gp, (list, tuple)) and len(gp) > 0 else 'N/A',
                            'gp_y': gp[1] if isinstance(gp, (list, tuple)) and len(gp) > 1 else 'N/A',
                            'pd': pdv if pdv is not None else 'N/A',
                            'gp3_x': gp3[0],'gp3_y': gp3[1],'gp3_z': gp3[2],
                            'gd_x': gd[0],'gd_y': gd[1],'gd_z': gd[2],
                            'pc_x': pc[0],'pc_y': pc[1],'pc_z': pc[2],
                            'gy_x': gy[0],'gy_y': gy[1],'gy_z': gy[2],
                            'ac_x': ac[0],'ac_y': ac[1],'ac_z': ac[2],
                            'eye': j.get('eye','N/A'),
                            'glasses_ts': j.get('ts'),
                            'glasses_pts': j.get('pts'),
                            's': j.get('s'),
                            'l': j.get('l'),
                            'gidx': j.get('gidx'),
                            'marker_flag': False,
                            'marker_type': 'N/A',
                            'marker_condition': 'N/A',
                            'marker_target': 'N/A',
                            'marker_color': 'N/A',
                            'marker_label': 'N/A',
                            'Gaze_TS': j.get('ts', ''),
                'Gaze_X': (gp[0] if isinstance(gp, (list, tuple)) and len(gp)>0 else (self._last_gaze_xy[0] if self._last_gaze_xy and (time.time()-self._last_gaze_time)<=1.0 else '')),
                'Gaze_Y': (gp[1] if isinstance(gp, (list, tuple)) and len(gp)>1 else (self._last_gaze_xy[1] if self._last_gaze_xy and (time.time()-self._last_gaze_time)<=1.0 else ''))
                        }
                        
                        # 尝试附加标记事件到眼动数据
                        try:
                            if hasattr(self, '_eye_marker_queue') and self._eye_marker_queue:
                                # 解析当前行的系统时间为epoch秒
                                row_ts_epoch = None
                                try:
                                    row_ts_epoch = datetime.strptime(st, "%Y-%m-%d %H:%M:%S.%f").timestamp()
                                except Exception:
                                    row_ts_epoch = None
                                win = getattr(self, 'marker_match_window_s', 0.01)
                                if row_ts_epoch is not None:
                                    # 在窗口内找到最早的匹配事件并打标，确保一对一
                                    for idx_ev, ev in enumerate(list(self._eye_marker_queue)):
                                        et = ev.get('time', 0)
                                        # 清理过期事件 (>5秒)
                                        if (time.time() - et) > 5.0:
                                            try:
                                                self._eye_marker_queue.remove(ev)
                                            except Exception:
                                                pass
                                            continue
                                        # 匹配时间窗口
                                        if abs(row_ts_epoch - et) <= win:
                                            row['marker_flag'] = True
                                            row['marker_type'] = ev.get('type', 'N/A')
                                            row['marker_condition'] = ev.get('cond', 'N/A')
                                            row['marker_target'] = ev.get('target', 'N/A')
                                            row['marker_color'] = ev.get('color', 'N/A')
                                            row['marker_label'] = ev.get('label', 'N/A')
                                            try:
                                                self._eye_marker_queue.remove(ev)
                                            except Exception:
                                                pass
                                            break
                        except Exception:
                            pass

                        try:
                            self._eye_stream_write_queue.put_nowait(row)
                        except Exception:
                            try:
                                _ = self._eye_stream_write_queue.get_nowait()
                            except Exception:
                                pass
                            try:
                                self._eye_stream_write_queue.put_nowait(row)
                            except Exception:
                                pass
                except socket.timeout:
                    continue
                except Exception:
                    break
        except Exception:
            pass
        finally:
            try:
                if self.tobii_sock:
                    self.tobii_sock.close()
            except Exception:
                pass
            self.tobii_sock = None
            self.tobii_connected = False

    def _start_eye_video_overlay(self):
        if self._eye_video_enabled:
            return
        if not GST_AVAILABLE:
            return
        try:
            self.tobii_glasses_ip = self.tobii_glasses_ip_var.get().strip() if hasattr(self, 'tobii_glasses_ip_var') else self.tobii_glasses_ip
            self.tobii_my_ip = self.tobii_my_ip_var.get().strip() if hasattr(self, 'tobii_my_ip_var') else self.tobii_my_ip
            self.tobii_scope_id = int(self.tobii_scope_id_var.get().strip()) if hasattr(self, 'tobii_scope_id_var') else self.tobii_scope_id
        except Exception:
            pass
        try:
            self._eye_video_sock = socket.socket(socket.AF_INET6, socket.SOCK_DGRAM)
            try:
                self._eye_video_sock.bind((self.tobii_my_ip, 0, 0, int(self.tobii_scope_id)))
            except Exception:
                return
            try:
                self._eye_video_sock.setsockopt(socket.SOL_SOCKET, socket.SO_RCVBUF, 4 * 1024 * 1024)
            except Exception:
                pass
            self._eye_video_buffersync = self._EyeBufferSync(lambda objs: self._eye_video_draw_gaze(objs))
            self._eye_video_pipe = Gst.parse_launch("udpsrc name=src blocksize=1316 closefd=false buffer-size=5600 ! tsparse ! tsdemux emit-stats=true ! queue ! avdec_h264 ! identity name=decoded ! textoverlay name=textovl text=* halignment=position valignment=position xpad=0 ypad=0 ! autovideosink name=video")
            bus = self._eye_video_pipe.get_bus()
            bus.add_signal_watch()
            bus.connect("message", self._eye_video_bus)
            src = self._eye_video_pipe.get_by_name("src")
            try:
                src.set_property("sockfd", self._eye_video_sock.fileno())
            except Exception:
                pass
            decoded = self._eye_video_pipe.get_by_name("decoded")
            decoded.connect("handoff", self._eye_video_on_decoded)
            self._eye_video_textovl = self._eye_video_pipe.get_by_name("textovl")
            self._eye_video_keepalive = self._EyeKeepAlive(self._eye_video_sock, (self.tobii_glasses_ip, 49152, 0, int(self.tobii_scope_id)), "video")
            self._eye_video_pipe.set_state(Gst.State.PLAYING)
            self._eye_video_loop = GLib.MainLoop()
            def run_loop():
                try:
                    self._eye_video_loop.run()
                except Exception:
                    pass
            self._eye_video_thread = threading.Thread(target=run_loop, daemon=True)
            self._eye_video_thread.start()
            self._eye_video_enabled = True
        except Exception:
            try:
                self._eye_video_enabled = False
            except Exception:
                pass

    def _stop_eye_video_overlay(self):
        try:
            self._eye_video_enabled = False
        except Exception:
            pass
        try:
            if self._eye_video_pipe:
                self._eye_video_pipe.set_state(Gst.State.NULL)
        except Exception:
            pass
        try:
            if self._eye_video_keepalive:
                self._eye_video_keepalive.stop()
        except Exception:
            pass
        try:
            if self._eye_video_sock:
                self._eye_video_sock.close()
        except Exception:
            pass
        try:
            if self._eye_video_loop:
                self._eye_video_loop.quit()
        except Exception:
            pass
        self._eye_video_pipe = None
        self._eye_video_textovl = None
        self._eye_video_sock = None
        self._eye_video_keepalive = None
        self._eye_video_buffersync = None
        self._eye_video_loop = None
        self._eye_video_thread = None

    def _eye_video_bus(self, bus, msg):
        try:
            if msg.type == Gst.MessageType.ELEMENT:
                s = msg.get_structure()
                if s and s.has_name("tsdemux") and s.has_field("pts") and s.has_field("offset"):
                    try:
                        self._eye_video_buffersync.add_pts_offset(s.get_value("offset"), s.get_value("pts"))
                    except Exception:
                        pass
        except Exception:
            pass
        return True

    def _eye_video_on_decoded(self, element, buf):
        try:
            self._eye_video_buffersync.flush_pts(buf.offset, int(buf.pts/1000))
        except Exception:
            pass

    def _eye_video_draw_gaze(self, objs):
        try:
            objs = list(filter(lambda x: "gp" in x, objs))
            if not objs:
                return
            o = objs[-1]
            if "gp" in o and o["gp"][0] != 0 and o["gp"][1] != 0:
                try:
                    if self._eye_video_textovl:
                        self._eye_video_textovl.set_property("xpos", float(o["gp"][0]))
                        self._eye_video_textovl.set_property("ypos", float(o["gp"][1]))
                except Exception:
                    pass
        except Exception:
            pass

    class _EyeKeepAlive:
        def __init__(self, sock, peer, streamtype, timeout=1):
            try:
                j = json.dumps({"op":"start","type":".".join(["live",streamtype,"unicast"]),"key":"py3"}).encode("utf-8")
                sock.sendto(j, peer)
                self._id = GLib.timeout_add_seconds(timeout, self._tick, sock, peer, j)
            except Exception:
                self._id = 0
        def _tick(self, sock, peer, j):
            try:
                sock.sendto(j, peer)
            except Exception:
                pass
            return True
        def stop(self):
            try:
                if self._id:
                    GLib.source_remove(self._id)
            except Exception:
                pass

    class _EyeBufferSync:
        def __init__(self, cb):
            self._cb = cb
            self._et_syncs = []
            self._et_queue = []
            self._et_ts = 0
            self._pts_queue = []
            self._pts_ts = 0
        def add_et(self, o):
            if "pts" in o:
                self._et_syncs.append(o)
            else:
                self._et_queue.append(o)
        def sync_et(self, pts, ts):
            past = list(filter(lambda x: x.get("pts",0) <= pts, self._et_syncs))
            self._et_syncs = list(filter(lambda x: x.get("pts",0) > pts, self._et_syncs))
            if past:
                self._et_ts = past[-1]["ts"]
                self._pts_ts = ts
        def flush_et(self, ts):
            nowts = self._et_ts + (ts - self._pts_ts)
            passed = list(filter(lambda x: x.get("ts",0) <= nowts, self._et_queue))
            self._et_queue = list(filter(lambda x: x.get("ts",0) > nowts, self._et_queue))
            if passed:
                try:
                    self._cb(passed)
                except Exception:
                    pass
        def add_pts_offset(self, off, pts):
            self._pts_queue.append((off, pts))
        def flush_pts(self, off, ts):
            used = list(filter(lambda x: x[0] <= off, self._pts_queue))
            self._pts_queue = list(filter(lambda x: x[0] > off, self._pts_queue))
            if used:
                self.sync_et(used[-1][1], ts)
            self.flush_et(ts)

    def _set_marker_condition(self, cond: int):
        self.marker_condition.set(cond)
        self.marker_status_var.set(f"标记条件: {cond}")
        self.log(f"切换标记条件为 {cond}")

    def _on_channel_selected(self, event=None):
        if hasattr(self, 'lsl_plotter') and self.lsl_plotter:
            self.lsl_plotter.set_channel(self.channel_combo.current())

    def _on_stream_selected(self, event=None):
        if hasattr(self, 'lsl_plotter') and self.lsl_plotter:
            self.lsl_plotter.set_stream(self.stream_combo.current())

    def _init_marker_sets(self):
        # 定义事件数据
        # 格式: (distance, road, type, label)
        # Condition 1
        self.cond1_fv_data = [
            # BE Road 1
            (7416, 'Road 1', 'BE', '仅刹车事件'), (14693, 'Road 1', 'BE', '仅刹车事件'),
            (21749, 'Road 1', 'BE', '仅刹车事件'), (28777, 'Road 1', 'BE', '仅刹车事件'),
            (35749, 'Road 1', 'BE', '仅刹车事件'), (42804, 'Road 1', 'BE', '仅刹车事件'),
            (49749, 'Road 1', 'BE', '仅刹车事件'),
            # BE Road 2
            (1666, 'Road 2', 'BE', '仅刹车事件'), (8582, 'Road 2', 'BE', '仅刹车事件'),
            (14999, 'Road 2', 'BE', '仅刹车事件'), (22304, 'Road 2', 'BE', '仅刹车事件'),
            (28360, 'Road 2', 'BE', '仅刹车事件'),
            # CE-2 Road 1
            (4809, 'Road 1', 'CE-2', '耦合事件-前车刹车'), (10003, 'Road 1', 'CE-2', '耦合事件-前车刹车'),
            (19392, 'Road 1', 'CE-2', '耦合事件-前车刹车'), (26476, 'Road 1', 'CE-2', '耦合事件-前车刹车'),
            (31226, 'Road 1', 'CE-2', '耦合事件-前车刹车'), (38170, 'Road 1', 'CE-2', '耦合事件-前车刹车'),
            (45142, 'Road 1', 'CE-2', '耦合事件-前车刹车'), (52087, 'Road 1', 'CE-2', '耦合事件-前车刹车'),
            # CE-2 Road 2
            (4003, 'Road 2', 'CE-2', '耦合事件-前车刹车'), (10920, 'Road 2', 'CE-2', '耦合事件-前车刹车'),
            (17837, 'Road 2', 'CE-2', '耦合事件-前车刹车'), (24448, 'Road 2', 'CE-2', '耦合事件-前车刹车')
        ]
        
        self.cond1_uv_data = [
            # OE Road 1
            (2500, 'Road 1', 'OE', '仅对向来车事件'), (12278, 'Road 1', 'OE', '仅对向来车事件'),
            (16944, 'Road 1', 'OE', '仅对向来车事件'), (23917, 'Road 1', 'OE', '仅对向来车事件'),
            (33306, 'Road 1', 'OE', '仅对向来车事件'), (40333, 'Road 1', 'OE', '仅对向来车事件'),
            (47333, 'Road 1', 'OE', '仅对向来车事件'), (54250, 'Road 1', 'OE', '仅对向来车事件'),
            # OE Road 2
            (6167, 'Road 2', 'OE', '仅对向来车事件'), (12722, 'Road 2', 'OE', '仅对向来车事件'),
            (19833, 'Road 2', 'OE', '仅对向来车事件'), (26306, 'Road 2', 'OE', '仅对向来车事件'),
            # CE-1 Road 1
            (4639, 'Road 1', 'CE-1', '耦合事件-对向来车'), (9833, 'Road 1', 'CE-1', '耦合事件-对向来车'),
            (19222, 'Road 1', 'CE-1', '耦合事件-对向来车'), (26306, 'Road 1', 'CE-1', '耦合事件-对向来车'),
            (31056, 'Road 1', 'CE-1', '耦合事件-对向来车'), (38000, 'Road 1', 'CE-1', '耦合事件-对向来车'),
            (44972, 'Road 1', 'CE-1', '耦合事件-对向来车'), (51917, 'Road 1', 'CE-1', '耦合事件-对向来车'),
            # CE-1 Road 2
            (3833, 'Road 2', 'CE-1', '耦合事件-对向来车'), (10750, 'Road 2', 'CE-1', '耦合事件-对向来车'),
            (17667, 'Road 2', 'CE-1', '耦合事件-对向来车'), (24278, 'Road 2', 'CE-1', '耦合事件-对向来车'),
            # Scale
            (29333, 'Road 1', 'Scale', '量表2'),
            (1100, 'Road 2', 'Scale', '量表3'),
            (28750, 'Road 2', 'Scale', '量表4')
        ]

        # Condition 2
        self.cond2_fv_data = [
            # BE Road 1
            (2193, 'Road 1', 'BE', '仅刹车事件'), (11499, 'Road 1', 'BE', '仅刹车事件'),
            (16332, 'Road 1', 'BE', '仅刹车事件'), (25416, 'Road 1', 'BE', '仅刹车事件'),
            (32388, 'Road 1', 'BE', '仅刹车事件'), (36777, 'Road 1', 'BE', '仅刹车事件'),
            (45832, 'Road 1', 'BE', '仅刹车事件'), (50610, 'Road 1', 'BE', '仅刹车事件'),
            # BE Road 2
            (5054, 'Road 2', 'BE', '仅刹车事件'), (9554, 'Road 2', 'BE', '仅刹车事件'),
            (19582, 'Road 2', 'BE', '仅刹车事件'), (23610, 'Road 2', 'BE', '仅刹车事件'),
            # CE-2 Road 1
            (6871, 'Road 1', 'CE-2', '耦合事件-前车刹车'), (14177, 'Road 1', 'CE-2', '耦合事件-前车刹车'),
            (21177, 'Road 1', 'CE-2', '耦合事件-前车刹车'), (28066, 'Road 1', 'CE-2', '耦合事件-前车刹车'),
            (34593, 'Road 1', 'CE-2', '耦合事件-前车刹车'), (41454, 'Road 1', 'CE-2', '耦合事件-前车刹车'),
            (48427, 'Road 1', 'CE-2', '耦合事件-前车刹车'),
            # CE-2 Road 2
            (538, 'Road 2', 'CE-2', '耦合事件-前车刹车'), (7593, 'Road 2', 'CE-2', '耦合事件-前车刹车'),
            (14649, 'Road 2', 'CE-2', '耦合事件-前车刹车'), (21732, 'Road 2', 'CE-2', '耦合事件-前车刹车'),
            (28260, 'Road 2', 'CE-2', '耦合事件-前车刹车')
        ]

        self.cond2_uv_data = [
            # OE Road 1
            (4389, 'Road 1', 'OE', '仅对向来车事件'), (8972, 'Road 1', 'OE', '仅对向来车事件'),
            (18667, 'Road 1', 'OE', '仅对向来车事件'), (23361, 'Road 1', 'OE', '仅对向来车事件'),
            (30139, 'Road 1', 'OE', '仅对向来车事件'), (39167, 'Road 1', 'OE', '仅对向来车事件'),
            (43639, 'Road 1', 'OE', '仅对向来车事件'), (52722, 'Road 1', 'OE', '仅对向来车事件'),
            # OE Road 2
            (2861, 'Road 2', 'OE', '仅对向来车事件'), (11833, 'Road 2', 'OE', '仅对向来车事件'),
            (16778, 'Road 2', 'OE', '仅对向来车事件'), (25833, 'Road 2', 'OE', '仅对向来车事件'),
            # CE-1 Road 1
            (6611, 'Road 1', 'CE-1', '耦合事件-对向来车'), (13917, 'Road 1', 'CE-1', '耦合事件-对向来车'),
            (20917, 'Road 1', 'CE-1', '耦合事件-对向来车'), (27806, 'Road 1', 'CE-1', '耦合事件-对向来车'),
            (34333, 'Road 1', 'CE-1', '耦合事件-对向来车'), (41194, 'Road 1', 'CE-1', '耦合事件-对向来车'),
            (48167, 'Road 1', 'CE-1', '耦合事件-对向来车'),
            # CE-1 Road 2
            (278, 'Road 2', 'CE-1', '耦合事件-对向来车'), (7333, 'Road 2', 'CE-1', '耦合事件-对向来车'),
            (14389, 'Road 2', 'CE-1', '耦合事件-对向来车'), (21472, 'Road 2', 'CE-1', '耦合事件-对向来车'),
            (28000, 'Road 2', 'CE-1', '耦合事件-对向来车'),
            # Scale
            (29333, 'Road 1', 'Scale', '量表2'),
            (1100, 'Road 2', 'Scale', '量表3'),
            (28750, 'Road 2', 'Scale', '量表4')
        ]

        # Helper to process and sort data
        def process_data(data):
            # Sort by Road (Road 1 < Road 2) then by distance
            # Assuming 'Road 1' < 'Road 2' string comparison works
            return sorted(data, key=lambda x: (x[1], x[0]))

        self.cond1_fv_events = process_data(self.cond1_fv_data)
        self.cond1_uv_events = process_data(self.cond1_uv_data)
        self.cond2_fv_events = process_data(self.cond2_fv_data)
        self.cond2_uv_events = process_data(self.cond2_uv_data)

        self._marked_fv = set()
        self._marked_uv = set()
        self._next_idx_fv = {} # cond -> index (0-based)
        self._next_idx_uv = {} # cond -> index (0-based)
        # Split event streams by type for independent sequencing
        def filter_by_type(data, t):
            return process_data([x for x in data if x[2] == t])
        # Condition 1 type-separated events
        self.cond1_fv_BE_events = filter_by_type(self.cond1_fv_data, 'BE')
        self.cond1_fv_CE2_events = filter_by_type(self.cond1_fv_data, 'CE-2')
        self.cond1_uv_OE_events = filter_by_type(self.cond1_uv_data, 'OE')
        self.cond1_uv_CE1_events = filter_by_type(self.cond1_uv_data, 'CE-1')
        self.cond1_uv_Scale_events = filter_by_type(self.cond1_uv_data, 'Scale')
        # Condition 2 type-separated events
        self.cond2_fv_BE_events = filter_by_type(self.cond2_fv_data, 'BE')
        self.cond2_fv_CE2_events = filter_by_type(self.cond2_fv_data, 'CE-2')
        self.cond2_uv_OE_events = filter_by_type(self.cond2_uv_data, 'OE')
        self.cond2_uv_CE1_events = filter_by_type(self.cond2_uv_data, 'CE-1')
        self.cond2_uv_Scale_events = filter_by_type(self.cond2_uv_data, 'Scale')
        # Independent next indices per stream
        self._next_idx_fv_BE = {}    # cond -> next idx (1-based)
        self._next_idx_fv_CE2 = {}   # cond -> next idx (1-based)
        self._next_idx_uv_OE = {}    # cond -> next idx (1-based)
        self._next_idx_uv_CE1 = {}   # cond -> next idx (1-based)
        self._next_idx_uv_Scale = {} # cond -> next idx (1-based)
        # Track last CE-1 time per condition for coupling constraint
        self._last_ce1_time = {1: None, 2: None}

        self._last_marker_event = None

    def _emit_marker_event(self, row, event_type, cond, target_dist, color, label):
        try:
            event_data = {
                'type': event_type,
                'cond': cond,
                'target': target_dist,
                'color': color,
                'label': label,
                'time': time.time(),
                'event_system_timestamp': row.get('system_timestamp'),
                'event_synchronized_timestamp': row.get('synchronized_timestamp')
            }
            # 分发到各模块的独立队列，避免竞态条件
            self._marker_event_queue.append(event_data)
            try:
                self._eye_marker_queue.append(event_data)
            except Exception:
                pass

            try:
                # Reuse existing emit methods which take a dict
                # They expect: type, cond, target, color, label
                self._emit_biosig_marker(event_data)
                self._emit_eye_marker(event_data)
            except Exception:
                pass
            try:
                marker_row = {
                    "system_timestamp": event_data.get('event_system_timestamp', row.get('system_timestamp')),
                    "synchronized_timestamp": event_data.get('event_synchronized_timestamp', row.get('synchronized_timestamp')),
                    "lsl_timestamp": "N/A",
                    "marker_flag": True,
                    "marker_type": event_type,
                    "marker_label": label,
                    "marker_condition": cond,
                    "marker_target": target_dist,
                    "marker_color": color,
                    "marker_stream_name": "internal"
                }
                self._marker_stream_write_queue.put_nowait(marker_row)
            except Exception:
                try:
                    self._marker_queue_dropped = getattr(self, '_marker_queue_dropped', 0) + 1
                except Exception:
                    pass
            try:
                if event_type == 'CE-1':
                    self._last_ce1_time[cond] = event_data.get('time')
            except Exception:
                pass
        except Exception:
            self._last_marker_event = {
                'type': event_type,
                'cond': cond,
                'target': target_dist,
                'color': color,
                'label': label,
                'time': time.time()
            }

    def _check_and_attach_marker(self, row: dict):
        cond_raw = None
        try:
            cond_raw = self.marker_condition.get()
        except Exception:
            cond_raw = 1
        try:
            cond = int(cond_raw)
        except Exception:
            cond = cond_raw
        try:
            uv_travel_raw = self.current_driving_params.get('uv_distanceTravelled', None)
            uv_travel = float(uv_travel_raw) if uv_travel_raw not in (None, 'N/A', '') else None
        except Exception:
            uv_travel = None
        if uv_travel is None:
            try:
                uv_along_raw = self.current_driving_params.get('uv_distanceAlongRoad', None)
                uv_travel = float(uv_along_raw) if uv_along_raw not in (None, 'N/A', '') else None
            except Exception:
                uv_travel = None
        if uv_travel is None:
            return
        dist_norm = uv_travel

        # Select event lists based on condition, separated by type
        if cond == 1:
            fv_be_events = self.cond1_fv_BE_events
            fv_ce2_events = self.cond1_fv_CE2_events
            uv_oe_events = self.cond1_uv_OE_events
            uv_ce1_events = self.cond1_uv_CE1_events
            uv_scale_events = self.cond1_uv_Scale_events
        else:
            fv_be_events = self.cond2_fv_BE_events
            fv_ce2_events = self.cond2_fv_CE2_events
            uv_oe_events = self.cond2_uv_OE_events
            uv_ce1_events = self.cond2_uv_CE1_events
            uv_scale_events = self.cond2_uv_Scale_events

        road_fv = 'Road 1'
        road_uv = 'Road 1'
        try:
            row['uv_road'] = road_uv
        except Exception:
            pass
        try:
            row['fv_road'] = road_fv
        except Exception:
            pass

        # --- FV Events Processing ---
        fv_val = dist_norm
            
        # Get fv_brake for BE/CE-2 check with enhanced brake intensity detection
        fv_brake = 0.0
        fv_brake_intensity = 0.0
        try:
            # Check if fv_brake is in current_driving_params (it should be)
            fv_brake_raw = self.current_driving_params.get('fv_brake', 0)
            fv_brake = float(fv_brake_raw) if fv_brake_raw not in (None,'N/A','') else 0.0
            
            # Enhanced brake intensity detection for better TTC calculation
            if fv_brake > 0:
                # Calculate brake intensity based on deceleration rate
                # Strong brake: > 6.0 m/s², Medium brake: 3.0-6.0 m/s², Light brake: < 3.0 m/s²
                fv_speed_ms = float(self.current_driving_params.get('fv_speedInMetresPerSecond', 0))
                if fv_speed_ms > 0:
                    # Estimate deceleration based on speed change (simplified)
                    fv_brake_intensity = min(fv_brake * 8.0, 10.0)  # Scale to reasonable deceleration values
                    
                    # Log significant brake events for analysis
                    if fv_brake_intensity >= 6.0:
                        self.log(f"🚨 强刹车事件检测: 强度={fv_brake_intensity:.1f} m/s², 速度={fv_speed_ms:.1f} m/s")
                    elif fv_brake_intensity >= 3.0:
                        self.log(f"⚠️ 中刹车事件检测: 强度={fv_brake_intensity:.1f} m/s², 速度={fv_speed_ms:.1f} m/s")
                        
        except Exception:
            fv_brake = 0.0
            fv_brake_intensity = 0.0
        uv_brake = 0.0
        try:
            uv_brake_raw = self.current_driving_params.get('uv_brake', 0)
            uv_brake = float(uv_brake_raw) if uv_brake_raw not in (None,'N/A','') else 0.0
        except Exception:
            uv_brake = 0.0

        if fv_val is not None:
            fv_norm = fv_val

            try:
                row['fv_road'] = road_fv
            except Exception:
                pass

            # BE stream
            be_idx = self._next_idx_fv_BE.get(cond, 1)
            if 1 <= be_idx <= len(fv_be_events):
                t_dist, t_road, t_type, t_label = fv_be_events[be_idx - 1]
                
                # Check if passed
                dist_diff = fv_norm - t_dist
                if dist_diff > 600.0:
                    self._next_idx_fv_BE[cond] = be_idx + 1
                    self.log(f"⚠ 跳过BE事件 {be_idx} {t_label} (passed by {dist_diff:.1f}m)")
                
                # Use larger tolerance (200m) with enhanced brake intensity check
                if abs(fv_norm - t_dist) <= 200.0 and fv_brake > 0:
                    # Enhanced brake event detection based on intensity
                    brake_intensity_level = "轻" if fv_brake_intensity < 3.0 else ("中" if fv_brake_intensity < 6.0 else "强")
                    
                    key = (cond, 'fv', t_dist, t_type)
                    if key not in self._marked_fv:
                        row['marker_flag'] = True
                        row['marker_type'] = t_type
                        row['marker_condition'] = cond
                        row['marker_target'] = t_dist
                        row['marker_color'] = 'blue'
                        row['marker_label'] = f"{be_idx} {t_label}（{t_type}）[{brake_intensity_level}刹车]"
                        row['brake_intensity'] = fv_brake_intensity  # Store for TTC calculation
                        self._marked_fv.add(key)
                        self._next_idx_fv_BE[cond] = be_idx + 1
                        self._emit_marker_event(row, t_type, cond, t_dist, 'blue', row['marker_label'])
                        self.log(f"✓ 触发FV标记 条件{cond} 类型={t_type} 目标={t_dist} (road={road_fv})")
                        try:
                            print(f"✓ 触发FV标记 条件{cond} 类型={t_type} 目标={t_dist} (road={road_fv})", flush=True)
                        except:
                            pass
                        self._log_marker_progress(cond)
            # CE-2 stream with coupling constraint (CE-1 within 5s)
            ce2_idx = self._next_idx_fv_CE2.get(cond, 1)
            if 1 <= ce2_idx <= len(fv_ce2_events):
                t_dist, t_road, t_type, t_label = fv_ce2_events[ce2_idx - 1]
                
                # Check if passed
                dist_diff = fv_norm - t_dist
                if dist_diff > 600.0:
                    self._next_idx_fv_CE2[cond] = ce2_idx + 1
                    self.log(f"⚠ 跳过CE-2事件 {ce2_idx} {t_label} (passed by {dist_diff:.1f}m)")

                ce1_ok = False
                try:
                    last = self._last_ce1_time.get(cond)
                    if last is not None and (time.time() - last) <= getattr(self, 'ce_coupling_window_s', 10.0):
                        ce1_ok = True
                except Exception:
                    ce1_ok = False
                # Use larger tolerance (200m)
                if abs(fv_norm - t_dist) <= 200.0 and fv_brake > 0 and ce1_ok:
                    key = (cond, 'fv', t_dist, t_type)
                    if key not in self._marked_fv:
                        row['marker_flag'] = True
                        row['marker_type'] = t_type
                        row['marker_condition'] = cond
                        row['marker_target'] = t_dist
                        row['marker_color'] = 'blue'
                        row['marker_label'] = f"{ce2_idx} {t_label}（{t_type}）"
                        self._marked_fv.add(key)
                        self._next_idx_fv_CE2[cond] = ce2_idx + 1
                        self._emit_marker_event(row, t_type, cond, t_dist, 'blue', row['marker_label'])
                        self.log(f"✓ 触发FV标记 条件{cond} 类型={t_type} 目标={t_dist} (road={road_fv})")
                        try:
                            print(f"✓ 触发FV标记 条件{cond} 类型={t_type} 目标={t_dist} (road={road_fv})", flush=True)
                        except:
                            pass
                        self._log_marker_progress(cond)

        # --- UV Events Processing ---
        uv_val = dist_norm

        if uv_val is not None:
            # Normalize distance
            uv_norm = uv_val

            try:
                row['uv_road'] = road_uv
            except Exception:
                pass

            # OE stream
            oe_idx = self._next_idx_uv_OE.get(cond, 1)
            if 1 <= oe_idx <= len(uv_oe_events):
                t_dist, t_road, t_type, t_label = uv_oe_events[oe_idx - 1]
                if abs(uv_norm - t_dist) <= 3.0:
                    key = (cond, 'uv', t_dist, t_type)
                    if key not in self._marked_uv:
                        row['marker_flag'] = True
                        row['marker_type'] = t_type
                        row['marker_condition'] = cond
                        row['marker_target'] = t_dist
                        row['marker_color'] = 'yellow'
                        row['marker_label'] = f"{oe_idx} {t_label}（{t_type}）"
                        self._marked_uv.add(key)
                        self._next_idx_uv_OE[cond] = oe_idx + 1
                        self._emit_marker_event(row, t_type, cond, t_dist, 'yellow', row['marker_label'])
                        self.log(f"✓ 触发UV标记 条件{cond} 类型={t_type} 目标={t_dist} (road={road_uv})")
                        try:
                            print(f"✓ 触发UV标记 条件{cond} 类型={t_type} 目标={t_dist} (road={road_uv})", flush=True)
                        except:
                            pass
                        self._log_marker_progress(cond)
            # CE-1 stream
            ce1_idx = self._next_idx_uv_CE1.get(cond, 1)
            if 1 <= ce1_idx <= len(uv_ce1_events):
                t_dist, t_road, t_type, t_label = uv_ce1_events[ce1_idx - 1]
                if abs(uv_norm - t_dist) <= 3.0:
                    key = (cond, 'uv', t_dist, t_type)
                    if key not in self._marked_uv:
                        row['marker_flag'] = True
                        row['marker_type'] = t_type
                        row['marker_condition'] = cond
                        row['marker_target'] = t_dist
                        row['marker_color'] = 'yellow'
                        row['marker_label'] = f"{ce1_idx} {t_label}（{t_type}）"
                        self._marked_uv.add(key)
                        self._next_idx_uv_CE1[cond] = ce1_idx + 1
                        self._emit_marker_event(row, t_type, cond, t_dist, 'yellow', row['marker_label'])
                        self.log(f"✓ 触发UV标记 条件{cond} 类型={t_type} 目标={t_dist} (road={road_uv})")
                        try:
                            print(f"✓ 触发UV标记 条件{cond} 类型={t_type} 目标={t_dist} (road={road_uv})", flush=True)
                        except:
                            pass
                        self._log_marker_progress(cond)
            # Scale stream
            sc_idx = self._next_idx_uv_Scale.get(cond, 1)
            if 1 <= sc_idx <= len(uv_scale_events):
                t_dist, t_road, t_type, t_label = uv_scale_events[sc_idx - 1]
                if abs(uv_norm - t_dist) <= 3.0:
                    key = (cond, 'uv', t_dist, t_type)
                    if key not in self._marked_uv:
                        row['marker_flag'] = True
                        row['marker_type'] = t_type
                        row['marker_condition'] = cond
                        row['marker_target'] = t_dist
                        row['marker_color'] = 'yellow'
                        row['marker_label'] = f"{sc_idx} {t_label}（{t_type}）"
                        self._marked_uv.add(key)
                        self._next_idx_uv_Scale[cond] = sc_idx + 1
                        self._emit_marker_event(row, t_type, cond, t_dist, 'yellow', row['marker_label'])
                        self.log(f"✓ 触发UV标记 条件{cond} 类型={t_type} 目标={t_dist} (road={road_uv})")
                        try:
                            print(f"✓ 触发UV标记 条件{cond} 类型={t_type} 目标={t_dist} (road={road_uv})", flush=True)
                        except:
                            pass
                        self._log_marker_progress(cond)

    def _log_marker_progress(self, cond):
        try:
            if cond == 1:
                fv_events = self.cond1_fv_events
                uv_events = self.cond1_uv_events
            else:
                fv_events = self.cond2_fv_events
                uv_events = self.cond2_uv_events
                
            total_expected = len(fv_events) + len(uv_events)
            fv_count = sum(1 for k in self._marked_fv if k[0] == cond)
            uv_count = sum(1 for k in self._marked_uv if k[0] == cond)
            total_count = fv_count + uv_count
            
            msg = f"[标记统计] FV: {fv_count}/{len(fv_events)}, UV: {uv_count}/{len(uv_events)}, 总计: {total_count}/{total_expected}"
            self.log(msg)
            try:
                print(msg, flush=True)
            except:
                pass
                
            if total_count == total_expected:
                done = f"[标记完成] 条件{cond}全部{total_count}个标记已触发"
                self.log(done)
                try:
                    print(done, flush=True)
                except:
                    pass
        except Exception:
            pass
    
    def _set_custom_filename(self):
        """设置自定义CSV文件名 - 改进版，更易用且支持实时预览"""
        dialog = tk.Toplevel(self.root)
        dialog.title("设置文件名")
        dialog.geometry("600x300")  # 增大对话框尺寸
        dialog.resizable(False, False)
        dialog.transient(self.root)  # 设置为主窗口的子窗口
        dialog.grab_set()  # 模态对话框
        
        # 创建样式
        dialog_style = ttk.Style(dialog)
        dialog_style.configure("TFrame", background="#f0f0f0")
        dialog_style.configure("TLabel", font=("微软雅黑", 10), background="#f0f0f0")
        dialog_style.configure("TEntry", font=("微软雅黑", 10), padding=5, width=40)
        dialog_style.configure("TButton", font=("微软雅黑", 10, "bold"), padding=6)
        dialog_style.configure("Preview.TLabel", font=("微软雅黑", 11, "italic"), foreground="#28a745", background="#f0f0f0")
        
        # 创建主框架
        main_frame = ttk.Frame(dialog, padding=20)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 当前时间戳
        try:
            base_ts = self.session_timestamp
        except Exception:
            base_ts = self.__class__.datetime.now().strftime("%Y%m%d_%H%M%S")
        default_name = f"driving_simulator_sync_{base_ts}"
        
        # 变量
        name_var = tk.StringVar(value=default_name)
        preview_var = tk.StringVar(value=f"预览: {default_name}.csv")
        
        # 文件名输入标签
        ttk.Label(main_frame, text="文件名 (无需输入.csv后缀):", anchor="w").pack(fill=tk.X, pady=(0, 5))
        
        # 文件名输入框 - 增大尺寸，更容易修改
        name_entry = ttk.Entry(main_frame, textvariable=name_var, width=50)
        name_entry.pack(fill=tk.X, pady=(0, 15), ipady=5)
        name_entry.focus()  # 自动聚焦
        name_entry.select_range(0, tk.END)  # 自动选中全部文本
        
        # 实时预览标签
        ttk.Label(main_frame, text="生成的完整文件名:", anchor="w").pack(fill=tk.X, pady=(0, 5))
        ttk.Label(main_frame, textvariable=preview_var, style="Preview.TLabel").pack(fill=tk.X, pady=(0, 20))
        
        # 错误消息标签
        error_var = tk.StringVar(value="")
        ttk.Label(main_frame, textvariable=error_var, foreground="#dc3545", anchor="w").pack(fill=tk.X, pady=(0, 10))
        
        # 分隔线
        separator = ttk.Separator(main_frame, orient="horizontal")
        separator.pack(fill=tk.X, pady=15)
        
        # 按钮框架
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, side=tk.BOTTOM)
        
        # 居中按钮
        button_frame.grid_columnconfigure(0, weight=1)
        button_frame.grid_columnconfigure(1, weight=1)
        button_frame.grid_columnconfigure(2, weight=1)
        
        # 取消按钮
        cancel_btn = ttk.Button(button_frame, text="取消", command=lambda: [setattr(self, "custom_filename", None), dialog.destroy()])
        cancel_btn.grid(row=0, column=0, padx=5)
        
        # 重置按钮
        def reset_filename():
            name_var.set(default_name)
            preview_var.set(f"预览: {default_name}.csv")
            error_var.set("")
        
        reset_btn = ttk.Button(button_frame, text="重置", command=reset_filename)
        reset_btn.grid(row=0, column=1, padx=5)
        
        # 确认按钮
        def confirm_filename():
            name = name_var.get().strip()
            if not name:
                error_var.set("文件名不能为空!")
                return
            
            # 检查文件名合法性
            invalid_chars = '<>:"/\\|?*'
            if any(char in name for char in invalid_chars):
                error_var.set(f"文件名不能包含以下字符: {invalid_chars}")
                return
            
            # 检查文件是否已存在
            full_path = os.path.join(self.data_folder, f"{name}.csv")
            if os.path.exists(full_path):
                if messagebox.askyesno("文件已存在", f"文件 '{name}.csv' 已存在，是否覆盖?"):
                    self.custom_filename = f"{name}.csv"
                    try:
                        self.filename_var.set(f"当前文件: {self.custom_filename}")
                    except Exception:
                        pass
                    dialog.destroy()
            else:
                self.custom_filename = f"{name}.csv"
                try:
                    self.filename_var.set(f"当前文件: {self.custom_filename}")
                except Exception:
                    pass
                dialog.destroy()
        
        confirm_btn = ttk.Button(button_frame, text="确认", style="Start.TButton", command=confirm_filename)
        confirm_btn.grid(row=0, column=2, padx=5)
        
        # 实时更新预览
        def update_preview(*args):
            name = name_var.get().strip()
            if name:
                preview_var.set(f"预览: {name}.csv")
                error_var.set("")
            else:
                preview_var.set("预览: (文件名不能为空)")
        
        name_var.trace_add("write", update_preview)
        
        # 按Enter确认，按Esc取消
        dialog.bind("<Return>", lambda event: confirm_filename())
        dialog.bind("<Escape>", lambda event: [setattr(self, "custom_filename", None), dialog.destroy()])
        
        # 等待对话框关闭
        self.root.wait_window(dialog)
        
        if hasattr(self, 'custom_filename') and self.custom_filename:
            self.log(f"✓ CSV文件名已设置为: {self.custom_filename}")
        return self.custom_filename

    def _launch_v16_mapper(self):
        try:
            import re, os, subprocess, sys
            sid = "S001"
            try:
                if hasattr(self, 'custom_filename') and self.custom_filename:
                    m = re.search(r"(S\d{2,})", self.custom_filename, re.I)
                    if m:
                        sid = m.group(1).upper()
                else:
                    s = self.filename_var.get()
                    m = re.search(r"(S\d{2,})", s, re.I)
                    if m:
                        sid = m.group(1).upper()
            except Exception:
                pass
            base = None
            try:
                if hasattr(self, 'custom_filename') and self.custom_filename:
                    base = os.path.splitext(self.custom_filename)[0]
                else:
                    s = self.filename_var.get()
                    if ":" in s:
                        s = s.split(":",1)[1].strip()
                    base = os.path.splitext(s)[0]
            except Exception:
                base = None
            try:
                x = self.root.winfo_rootx()
                y = self.root.winfo_rooty()
                w = max(self.root.winfo_width(), 800)
                h = max(self.root.winfo_height(), 600)
                h_top = h // 2
                vx = x
                vy = y + h_top
                vw = w
                vh = h - h_top
            except Exception:
                vx, vy, vw, vh = 0, 480, 960, 540
            env = os.environ.copy()
            env['GI_TYPELIB_PATH'] = env.get('GI_TYPELIB_PATH', r'C:\msys64\mingw64\lib\girepository-1.0')
            env['GST_PLUGIN_PATH'] = env.get('GST_PLUGIN_PATH', r'C:\msys64\mingw64\lib\gstreamer-1.0')
            env['GST_PLUGIN_SYSTEM_PATH'] = env.get('GST_PLUGIN_SYSTEM_PATH', r'C:\msys64\mingw64\lib\gstreamer-1.0')
            env['GST_PLUGIN_SYSTEM_PATH_1_0'] = env.get('GST_PLUGIN_SYSTEM_PATH_1_0', r'C:\msys64\mingw64\lib\gstreamer-1.0')
            try:
                scanner = r'C:\msys64\mingw64\lib\gstreamer-1.0\gst-plugin-scanner.exe'
                if os.path.exists(scanner):
                    env['GST_PLUGIN_SCANNER'] = scanner
            except Exception:
                pass
            # 过滤可能引入 MSVC GStreamer/GLib 的路径，避免与 MSYS 混用
            old_path = env.get('Path','')
            parts = [p for p in old_path.split(';') if p]
            msys_bin = r'C:\msys64\mingw64\bin'
            def keep(p):
                low = p.lower()
                if low.startswith(msys_bin.lower()):
                    return True
                if ('gstreamer' in low or 'glib-2.0' in low) and ('msvc' in low or 'msvc_x86_64' in low):
                    return False
                return True
            parts = [p for p in parts if keep(p)]
            env['Path'] = msys_bin + ';' + ';'.join(parts)
            # 仅保留 MSYS 的 GI_TYPELIB 路径
            gi_parts = [p for p in env.get('GI_TYPELIB_PATH','').split(';') if p]
            gi_ms = r'C:\msys64\mingw64\lib\girepository-1.0'
            gi_parts = [gi_ms] + [p for p in gi_parts if p.lower().startswith(gi_ms.lower())]
            env['GI_TYPELIB_PATH'] = ';'.join(dict.fromkeys(gi_parts))
            msys_py = getattr(self, 'msys_python_path', r'C:\msys64\mingw64\bin\python3.exe')
            py = msys_py if os.path.exists(msys_py) else sys.executable
            vdual = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'TobiiSyncRecorder_v16.py')
            args = [py, vdual, '--x', str(vx), '--y', str(vy), '--w', str(vw), '--h', str(vh)]
            if base:
                args.extend(['--base', base])
            try:
                self.log(f"✓ 启动V16眼动窗口: {py}")
                self.log(f"  Args: {args}")
                self.proc_v16_dual = subprocess.Popen(args, env=env)
            except Exception as e:
                self.log(f"✗ 启动V16眼动窗口失败: {e}")
                self.proc_v16_dual = None
            try:
                self.log("✓ 已启动注视点映射视频模块(V16)")
            except Exception:
                pass
        except Exception as e:
            self.log(f"✗ _launch_v16_mapper 异常: {e}")

    def _close_v16_mapper(self):
        try:
            if self.proc_v16:
                self.proc_v16.terminate()
                self.proc_v16 = None
                try:
                    self.log("✓ 已关闭注视点映射视频模块(V16)")
                except Exception:
                    pass
            if hasattr(self, 'proc_v16_dual') and self.proc_v16_dual:
                try:
                    self.proc_v16_dual.terminate()
                except Exception:
                    pass
                self.proc_v16_dual = None
        except Exception:
            pass

    def _launch_monitor_live(self):
        try:
            import os, subprocess, sys
            env = os.environ.copy()
            env['GI_TYPELIB_PATH'] = env.get('GI_TYPELIB_PATH', r'C:\msys64\mingw64\lib\girepository-1.0')
            env['GST_PLUGIN_PATH'] = env.get('GST_PLUGIN_PATH', r'C:\msys64\mingw64\lib\gstreamer-1.0')
            env['GST_PLUGIN_SYSTEM_PATH'] = env.get('GST_PLUGIN_SYSTEM_PATH', r'C:\msys64\mingw64\lib\gstreamer-1.0')
            env['GST_PLUGIN_SYSTEM_PATH_1_0'] = env.get('GST_PLUGIN_SYSTEM_PATH_1_0', r'C:\msys64\mingw64\lib\gstreamer-1.0')
            try:
                scanner = r'C:\msys64\mingw64\lib\gstreamer-1.0\gst-plugin-scanner.exe'
                if os.path.exists(scanner):
                    env['GST_PLUGIN_SCANNER'] = scanner
            except Exception:
                pass
            # 过滤可能引入 MSVC GStreamer/GLib 的路径，避免与 MSYS 混用
            old_path = env.get('Path','')
            parts = [p for p in old_path.split(';') if p]
            msys_bin = r'C:\msys64\mingw64\bin'
            def keep(p):
                low = p.lower()
                if low.startswith(msys_bin.lower()):
                    return True
                if ('gstreamer' in low or 'glib-2.0' in low) and ('msvc' in low or 'msvc_x86_64' in low):
                    return False
                return True
            parts = [p for p in parts if keep(p)]
            env['Path'] = msys_bin + ';' + ';'.join(parts)
            msys_py = getattr(self, 'msys_python_path', r'C:\msys64\mingw64\bin\python3.exe')
            py = msys_py if os.path.exists(msys_py) else sys.executable
            base = None
            try:
                if hasattr(self, 'custom_filename') and self.custom_filename:
                    base = os.path.splitext(self.custom_filename)[0]
                else:
                    s = self.filename_var.get()
                    if ":" in s:
                        s = s.split(":",1)[1].strip()
                    base = os.path.splitext(s)[0]
            except Exception:
                base = None
            vdual = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'v16_dual_screen_sync.py')
            # 计算顶部窗口几何
            try:
                x = self.root.winfo_rootx()
                y = self.root.winfo_rooty()
                w = max(self.root.winfo_width(), 800)
                h = max(self.root.winfo_height(), 600)
                h_top = h // 2
                vx, vy = x, y
                vw, vh = w, h_top
            except Exception:
                vx, vy, vw, vh = 0, 0, 960, 480
            args = [py, vdual, '--mode', 'monitor_only', '--x', str(vx), '--y', str(vy), '--w', str(vw), '--h', str(vh)]
            if base:
                args.extend(['--base', base])
            try:
                self.log(f"✓ 启动监控直播模块... {py}")
                self.log(f"  Args: {args}")
                self.proc_monitor_only = subprocess.Popen(args, env=env)
                try:
                    self.log("✓ 已启动监控直播模块")
                except Exception:
                    pass
            except Exception as e:
                self.log(f"✗ 启动监控直播模块失败: {e}")
                self.proc_monitor_only = None
        except Exception as e:
            self.log(f"✗ _launch_monitor_live 异常: {e}")

    def _close_monitor_live(self):
        try:
            if hasattr(self, 'proc_monitor_only') and self.proc_monitor_only:
                try:
                    self.proc_monitor_only.terminate()
                except Exception:
                    pass
                self.proc_monitor_only = None
                try:
                    self.log("✓ 已关闭监控直播模块")
                except Exception:
                    pass
        except Exception:
            pass

    def _launch_monitor_only(self):
        return self._launch_monitor_live()
    def _close_monitor_only(self):
        return self._close_monitor_live()
        
    def _open_csv_file(self):
        """打开CSV文件进行写入"""
        # 使用自定义文件名或生成默认文件名
        if hasattr(self, 'custom_filename') and self.custom_filename:
            self.csv_filename = self.custom_filename
        else:
            try:
                base_ts = self.session_timestamp
            except Exception:
                base_ts = self.__class__.datetime.now().strftime("%Y%m%d_%H%M%S")
            self.csv_filename = f"driving_simulator_sync_{base_ts}.csv"
        
        # 创建完整的文件路径，将文件保存到驾驶指标保存文件夹
        full_csv_path = os.path.join(self.data_folder, self.csv_filename)
        
        # 创建文件，使用utf-8-sig编码确保中文在Excel中正常显示
        self.csv_file = open(full_csv_path, 'w', newline='', encoding='utf-8-sig')
        
        # 构建CSV表头：按照用户要求的顺序
        csv_columns = [
            "system_timestamp", "Model", "ID", "description", "uv_Time", "laneWidth", "fv_distanceTravelled", "speedInKmPerHour", 
            "speedInMetresPerSecond", "distanceTravelled", "steering", "brake", "uv_throttle", "lightState", "distanceAlongRoad", 
            "fv_ID", "fv_Time", "fv_Model", "fv_description", "fv_distanceToLeftBorder", 
            "fv_distanceToRightBorder", "fv_speedInKmPerHour", "fv_steering", "fv_throttle", 
            "fv_lightState", "fv_automaticControl", "fv_wheelBase", "fv_road", "fv_distanceAlongRoad", 
            "fv_laneNumber", "distanceToLeftBorder", "distanceToRightBorder", "offsetFromRoadCenter", 
            "fv_position X", "fv_position Y", "fv_position Z", "offsetFromLaneCenter",
            "TTC", "DistanceToLead",
            "biosig_timestamp", "heart_rate_bpm", "gsr_uS", "system_lsl_time", "time_diff_lsl"
        ]
        
        # 创建CSV写入器
        self.csv_writer = csv.DictWriter(self.csv_file, fieldnames=csv_columns)
        self.csv_writer.writeheader()
        
        self.log(f"✓ CSV文件已创建: {full_csv_path}")
        self.log(f"✓ CSV表头包含 {len(csv_columns)} 列")
    
    def _receive_driving_lsl_data(self):
        """通过LSL接收驾驶数据 (替代UDP)"""
        if not globals().get('PYLSL_AVAILABLE', False):
            self.log("✗ 未检测到 pylsl 模块，无法使用 LSL 接收驾驶数据。")
            self.log("  请在终端运行: pip install pylsl")
            return

        self.log("🔍 正在查找驾驶数据LSL流 (Type='Driving' 或 Name='UnityDriving')...")
        
        driving_inlet = None
        channel_names = []
        
        # 增加看门狗定时器
        last_data_time = time.time()
        
        while self.running:
            try:
                # 1. 连接逻辑
                if driving_inlet is None:
                    # 使用全局锁防止并发调用导致死锁
                    if 'LSL_RESOLVE_LOCK' in globals():
                        with LSL_RESOLVE_LOCK:
                             streams = resolve_streams(wait_time=1.0)
                    else:
                         streams = resolve_streams(wait_time=1.0)
                    
                    # 调试：打印所有发现的流 (已移除详细日志以避免UI卡顿)
                    if streams:
                         pass


                    for stream in streams:
                        # 兼容多种常见的流名称
                        # 扩展匹配规则：包含 'Driving' 的名称或类型，或者包含 'Unity' 且通道数较多的流
                        name_lower = stream.name().lower()
                        type_lower = stream.type().lower()
                        
                        is_target = (
                            stream.type() == 'Driving' or 
                            stream.name() in ['UnityDrivingData', 'UnityDriving'] or
                            'driving' in name_lower or
                            ('unity' in name_lower and stream.channel_count() > 10)
                        )
                        
                        if is_target:
                            driving_inlet = StreamInlet(stream)
                            self.log(f"✓ 已连接到驾驶数据流: {stream.name()} (Type: {stream.type()}) — 数据来源: Unity 驾驶模拟器 LSL")
                            
                            # [Fix] Dynamic Channel Mapping
                            # Instead of hardcoding, we try to read channel names from LSL metadata first.
                            # If metadata is empty (old Unity plugin), we fall back to hardcoded list.
                            
                            dynamic_mapping = []
                            try:
                                info = driving_inlet.info()
                                ch = info.desc().child("channels").child("channel")
                                for k in range(info.channel_count()):
                                    name = ch.child_value("label")
                                    if name:
                                        # Map specific Unity names to Python internal names if needed
                                        # For now, we trust the Unity names match our keys (e.g. 'speedInKmPerHour')
                                        # But we need to handle 'uv_' prefix if Unity doesn't send it
                                        
                                        # Simple mapping logic:
                                        # If name is 'speedInKmPerHour', map to 'uv_speedInKmPerHour'
                                        # If name starts with 'fv_', keep it.
                                        
                                        if not name.startswith('fv_') and not name.startswith('uv_') and name != 'Event_Marker':
                                            # Assume it is a uv_ property
                                            # Special cases
                                            if name == 'speedInKmPerHour': dynamic_mapping.append('uv_speedInKmPerHour')
                                            elif name == 'steering': dynamic_mapping.append('uv_steering')
                                            elif name == 'brake': dynamic_mapping.append('uv_brake')
                                            elif name == 'laneWidth': dynamic_mapping.append('uv_laneWidth')
                                            elif name == 'distanceTravelled': dynamic_mapping.append('uv_distanceTravelled')
                                            elif name == 'lightState': dynamic_mapping.append('uv_lightState')
                                            elif name == 'distanceAlongRoad': dynamic_mapping.append('uv_distanceAlongRoad')
                                            elif name == 'distanceToLeftBorder': dynamic_mapping.append('uv_distanceToLeftBorder')
                                            elif name == 'distanceToRightBorder': dynamic_mapping.append('uv_distanceToRightBorder')
                                            elif name == 'offsetFromRoadCenter': dynamic_mapping.append('uv_offsetFromRoadCenter')
                                            elif name == 'offsetFromLaneCenter': dynamic_mapping.append('uv_offsetFromLaneCenter')
                                            elif name == 'speedInMetresPerSecond': dynamic_mapping.append('uv_speedInMetresPerSecond')
                                            else: dynamic_mapping.append(name) # Keep as is (e.g. TTC, DistanceToLead)
                                        else:
                                            dynamic_mapping.append(name)
                                    else:
                                        dynamic_mapping.append(f"Channel_{k}")
                                    ch = ch.next_sibling()
                            except Exception as e:
                                self.log(f"⚠ 读取LSL元数据失败: {e}")
                            
                            if len(dynamic_mapping) > 0:
                                self.driving_channel_mapping = dynamic_mapping
                                self.log(f"  已应用动态通道映射 (共{len(dynamic_mapping)}个通道):")
                                self.log(f"  {dynamic_mapping}")
                            else:
                                # Fallback：与 Unity LSLDrivingOutlet 的 35 通道顺序严格一致（索引 0=Event_Marker，5=本车里程）
                                self.log("⚠ 未找到通道元数据，使用默认硬编码映射(与Unity 35通道一致)")
                                self.driving_channel_mapping = [
                                    "Event_Marker",                 # 0
                                    "uv_laneWidth",                 # 1 laneWidth
                                    "fv_distanceTravelled",         # 2
                                    "uv_speedInKmPerHour",          # 3 speedInKmPerHour
                                    "uv_speedInMetresPerSecond",    # 4 speedInMetresPerSecond
                                    "uv_distanceTravelled",         # 5 distanceTravelled 本车里程
                                    "uv_steering",                  # 6 steering
                                    "uv_brake",                     # 7 brake
                                    "uv_lightState",                # 8 lightState
                                    "uv_distanceAlongRoad",         # 9 distanceAlongRoad
                                    "fv_ID_Hash",                   # 10
                                    "fv_Time",                      # 11
                                    "fv_Model_Hash",                # 12
                                    "fv_description_Hash",          # 13
                                    "fv_distanceToLeftBorder",      # 14
                                    "fv_distanceToRightBorder",     # 15
                                    "fv_speedInKmPerHour",          # 16
                                    "fv_steering",                  # 17
                                    "fv_throttle",                  # 18
                                    "fv_brake",                     # 19
                                    "fv_lightState",                # 20
                                    "fv_automaticControl",          # 21
                                    "fv_wheelBase",                 # 22
                                    "fv_road_Hash",                 # 23
                                    "fv_distanceAlongRoad",         # 24
                                    "fv_laneNumber",                # 25
                                    "uv_distanceToLeftBorder",      # 26 distanceToLeftBorder
                                    "uv_distanceToRightBorder",     # 27 distanceToRightBorder
                                    "uv_offsetFromRoadCenter",      # 28 offsetFromRoadCenter
                                    "fv_position X",                # 29
                                    "fv_position Y",                 # 30
                                    "fv_position Z",                # 31
                                    "uv_offsetFromLaneCenter",      # 32 offsetFromLaneCenter
                                    "TTC",                          # 33
                                    "DistanceToLead"                 # 34
                                ]
                            
                            channel_names = self.driving_channel_mapping
                            
                            # 验证通道数
                            try:
                                info = driving_inlet.info()
                                if info.channel_count() != len(channel_names):
                                    self.log(f"⚠ 通道数不匹配: Unity发送 {info.channel_count()} vs 映射表 {len(channel_names)}")
                            except:
                                pass
                            
                            break
                    
                    if driving_inlet is None:
                        # 没找到，打印当前检测到的流并等待后重试
                        stream_names = [s.name() for s in streams] if streams else []
                        if not hasattr(self, '_last_driving_stream_log_t') or (time.time() - getattr(self, '_last_driving_stream_log_t', 0)) > 10.0:
                            self._last_driving_stream_log_t = time.time()
                            self.log("⚠ 未找到驾驶数据LSL流 (Type=Driving 或 Name=UnityDrivingData)。当前检测到: " + (", ".join(stream_names) if stream_names else "无") + "。请确认Unity场景中已开始试次并发送驾驶数据。")
                        time.sleep(2.0)
                        continue

                # 2. 数据接收逻辑
                sample, timestamp = driving_inlet.pull_sample(timeout=0.05)
                if sample:
                    last_data_time = time.time()
                    # print("Got Data")
                    # 1. 解析并更新数据 (需加锁)
                    parsed_data = {}
                    with self.driving_lock:
                        for i, val in enumerate(sample):
                            if i < len(channel_names):
                                key = channel_names[i]
                                parsed_data[key] = val
                        
                        self.current_driving_params.update(parsed_data)
                        self.current_driving_params['lsl_timestamp'] = timestamp
                        self.last_driving_epoch = time.time()
                    
                    # 2. 后续处理 (释放锁后执行，避免死锁)
                    # UI 更新改为由主线程定时器 _tick_driving_display 执行，避免 LSL 线程频繁调度导致点击窗口卡死
                    # 此处仅更新数据，不调用 _display_driving_params()
                    
                    # 标记已接收到驾驶数据
                    self.driving_received = True
                    
                    # 添加驾驶数据心跳日志
                    current_time = time.time()
                    if not hasattr(self, '_last_driving_heartbeat_time'):
                        self._last_driving_heartbeat_time = 0
                    
                    if current_time - self._last_driving_heartbeat_time > 5.0:
                         # 使用局部变量 parsed_data 避免再次加锁读取
                         speed = parsed_data.get('uv_speedInKmPerHour', 0)
                         self.log(f"🚙 驾驶数据接收正常 (来源: Unity LSL) | LSL时间: {timestamp:.3f} | 速度: {speed:.1f} km/h")
                         self._last_driving_heartbeat_time = current_time

                    # 手动触发写入
                    if self.experiment_started:
                        self._write_driving_params_to_csv()
                        
                        # 高压模式下的缓存保护
                        if getattr(self, '_high_pressure_mode', False):
                            # 1. 先获取数据副本 (避免嵌套锁导致的死锁)
                            dummy_record = None
                            with self.driving_lock:
                                if self.current_driving_params:
                                    dummy_record = self.current_driving_params.copy()
                            
                            # 2. 再检查缓存并追加 (仅当有数据且缓存为空时)
                            if dummy_record:
                                with self.cache_lock:
                                    if not self.driving_data_cache:
                                        dummy_record['system_timestamp'] = self.__class__.datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
                                        self.driving_data_cache.append(dummy_record)
                                        
                else:
                    # 没数据，短暂休眠避免死循环占用CPU
                    # 如果连接断开 pull_sample 可能会抛出异常或者返回 None
                    
                    # 检查是否超时（例如5秒无数据），如果超时则尝试重置连接
                    if time.time() - last_data_time > 5.0 and driving_inlet is not None:
                        self.log("⚠ 驾驶数据接收超时 (5s)，尝试重置连接...")
                        driving_inlet = None
                        last_data_time = time.time()
                    pass
                            
            except Exception as e:
                # 如果出错（比如流断开），重置inlet
                self.log(f"⚠ LSL接收循环异常 (将尝试重连): {e}")
                driving_inlet = None
                time.sleep(1.0)

    def _receive_marker_lsl_data(self):
        """接收LSL标记数据 (Markers)"""
        if not globals().get('PYLSL_AVAILABLE', False):
            return

        self.log("🔍 正在查找LSL标记流 (Type='Markers')...")
        marker_inlets = []
        
        while self.running:
            try:
                if not marker_inlets:
                    if 'LSL_RESOLVE_LOCK' in globals():
                        with LSL_RESOLVE_LOCK:
                            streams = resolve_streams(wait_time=1.0)
                    else:
                        streams = resolve_streams(wait_time=1.0)

                    candidates = []
                    for stream in streams:
                        try:
                            name = stream.name()
                            stype = stream.type()
                        except Exception:
                            continue
                        if stype == 'Markers' or 'marker' in str(name).lower():
                            candidates.append(stream)

                    if not candidates:
                        time.sleep(2.0)
                        continue

                    marker_inlets = []
                    seen_names = set()
                    for stream in candidates:
                        try:
                            name = stream.name()
                            if name in seen_names:
                                continue
                            seen_names.add(name)
                            marker_inlets.append((StreamInlet(stream), name))
                            self.log(f"✓ 已连接到标记流: {name} (Type: {stream.type()})")
                        except Exception:
                            continue

                    if not marker_inlets:
                        time.sleep(2.0)
                        continue

                got_any = False
                for inlet, stream_name in marker_inlets:
                    sample = None
                    timestamp = None
                    try:
                        sample, timestamp = inlet.pull_sample(timeout=0.0)
                    except Exception:
                        raise
                    if not sample:
                        continue
                    got_any = True

                    try:
                        marker_text = sample[0] if isinstance(sample, list) and len(sample) == 1 else json.dumps(sample, ensure_ascii=False)
                    except Exception:
                        marker_text = str(sample)

                    try:
                        ts_key = None
                        try:
                            ts_key = round(float(timestamp), 6) if timestamp is not None else None
                        except Exception:
                            ts_key = str(timestamp)
                        key = (stream_name, str(marker_text), ts_key)
                        recent = getattr(self, '_recent_marker_keys', None)
                        if recent is None:
                            recent = deque(maxlen=200)
                            self._recent_marker_keys = recent
                        if key in recent:
                            continue
                        recent.append(key)
                    except Exception:
                        pass

                    try:
                        self.log(f"📍 收到Unity忠实标记: {marker_text}")
                    except Exception:
                        pass

                    try:
                        if marker_text == "EXPERIMENT_END":
                            self.log("收到实验结束标记，准备自动停止...")
                            try:
                                if self.root:
                                    self.root.after(0, self._stop_system)
                            except Exception:
                                pass
                    except Exception:
                        pass

                    try:
                        current_time = self.__class__.datetime.now()
                        system_timestamp = current_time.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
                        _, synchronized_timestamp, _ = self._get_synchronized_timestamp(system_timestamp)

                        # 1. 构造标准事件对象（原样保留 marker_text）
                        marker_ev = {
                            'time': time.time(),
                            'type': 'LSL_Event',        # 类型统一为 LSL_Event
                            'label': str(marker_text),  # <--- 这里原样保存 Unity 发来的 _LeadCar_Brake_Start 等所有内容！
                            'cond': 'Unity_LSL',
                            'target': 'N/A',
                            'color': 'purple',
                            'event_system_timestamp': system_timestamp,
                            'event_synchronized_timestamp': synchronized_timestamp,
                            'lsl_timestamp': timestamp,
                            'stream_name': stream_name
                        }

                        # 2. 【核心】：推入全局队列，让生理和眼动线程抓取
                        if not hasattr(self, '_marker_event_queue'):
                            self._marker_event_queue = deque()
                        self._marker_event_queue.append(marker_ev)

                        # 3. 立即强制写入生理和眼动文件各一行（作为脉冲事件行）
                        try:
                            self._emit_biosig_marker(marker_ev)
                            self._emit_eye_marker(marker_ev)
                        except Exception as e:
                            self.log(f"写入外设标记失败: {e}")

                        # 4. 作为独立行写入驾驶数据的缓存流
                        row = {
                            "system_timestamp": system_timestamp,
                            "synchronized_timestamp": synchronized_timestamp,
                            "lsl_timestamp": timestamp,
                            "marker_flag": True,
                            "marker_type": "LSL_Event",
                            "marker_label": str(marker_text), # 原样透传
                            "marker_condition": "Unity_LSL",
                            "marker_target": "N/A",
                            "marker_color": "purple",
                            "marker_stream_name": stream_name
                        }
                        
                        # 写入驾驶数据缓存
                        if hasattr(self, 'cache_lock') and hasattr(self, 'driving_data_cache'):
                            with self.cache_lock:
                                self.driving_data_cache.append(row.copy())
                        
                        # 同时写入驾驶CSV队列，使“仅标记、无驾驶流”时驾驶文件也有行数据
                        try:
                            self._driving_stream_write_queue.put_nowait(row.copy())
                        except queue.Full:
                            try:
                                _ = self._driving_stream_write_queue.get_nowait()
                            except Exception:
                                pass
                            try:
                                self._driving_stream_write_queue.put_nowait(row.copy())
                            except Exception:
                                pass
                        except Exception:
                            pass
                        
                        # 写入标记流式队列
                        try:
                            self._marker_stream_write_queue.put_nowait(row.copy())
                        except Exception:
                            try:
                                self._marker_queue_dropped = getattr(self, '_marker_queue_dropped', 0) + 1
                            except Exception:
                                pass
                        
                        # 驾驶参数/占位由主线程 _tick_driving_display 定时刷新，此处不调用避免卡死
                    except Exception as e:
                        self.log(f"⚠ 处理标记数据时出错: {e}")

                if not got_any:
                    time.sleep(0.002)
                             
            except Exception as e:
                self.log(f"⚠ 标记流接收异常: {e}")
                marker_inlets = []
                time.sleep(1.0)

    # _receive_udp_data 已被完全移除，仅使用 LSL 协议
    
    # _process_data_queue 已被移除
    
    # _parse_driving_params 已被移除
    
    def _parse_driving_params(self, message):
        """解析驾驶参数数据，忽略指定的UDP参数"""
        try:
            # 标记是否解析到有效数据
            has_parsed_data = False
            
            # 要忽略的UDP参数列表
            ignored_params = [
                'uv_yawAngle', 'uv_pitchAngle', 'uv_rollAngle', 'uv_bodyPitchAngle', 'uv_bodyRollAngle',
                'uv_RPM', 'uv_gearNumber', 'uv_speedVectInMetresPerSecond X', 'uv_speedVectInMetresPerSecond Y',
                'uv_speedVectInMetresPerSecond Z', 'uv_rotSpeedInRadsPerSecond Yaw', 'uv_rotSpeedInRadsPerSecond Pitch',
                'uv_rotSpeedInRadsPerSecond Roll', 'uv_rotAccelInRadsPerSecond Yaw', 'uv_rotAccelInRadsPerSecond Pitch',
                'uv_rotAccelInRadsPerSecond Roll', 'uv_steeringVelocity', 'uv_turningCurvature',
                'uv_localAccelInMetresPerSecond2 X', 'uv_localAccelInMetresPerSecond2 Y', 'uv_localAccelInMetresPerSecond2 Z',
                'uv_bodyRotSpeedInRadsPerSecond Yaw', 'uv_bodyRotSpeedInRadsPerSecond Pitch', 'uv_bodyRotSpeedInRadsPerSecond Roll',
                'uv_bodyRotAccelInRadsPerSecond Yaw', 'uv_bodyRotAccelInRadsPerSecond Pitch', 'uv_bodyRotAccelInRadsPerSecond Roll',
                'uv_dragForce', 'uv_mass', 'uv_centerOfGravityHeight', 'uv_centerOfGravityPosition',
                'uv_rollAxisHeight', 'uv_trailer', 'uv_trailerAngle', 'uv_trailerPitchAngle', 'uv_trailerWheelbase',
                'uv_inIntersection', 'uv_laneDirection X', 'uv_laneDirection Y', 'uv_laneDirection Z', 'uv_laneCurvature',
                'fv_yawAngle', 'fv_pitchAngle', 'fv_rollAngle', 'fv_bodyPitchAngle',
                'fv_bodyRollAngle', 'fv_RPM', 'fv_bodyRotSpeedInRadsPerSecond Yaw', 'fv_bodyRotSpeedInRadsPerSecond Pitch',
                'fv_bodyRotSpeedInRadsPerSecond Roll', 'fv_bodyRotAccelInRadsPerSecond Yaw', 'fv_bodyRotAccelInRadsPerSecond Pitch',
                'fv_bodyRotAccelInRadsPerSecond Roll', 'fv_rotSpeedInRadsPerSecond Yaw', 'fv_rotSpeedInRadsPerSecond Pitch',
                'fv_rotSpeedInRadsPerSecond Roll', 'fv_rotAccelInRadsPerSecond Yaw', 'fv_gearNumber',
                'fv_speedVectInMetresPerSecond X', 'fv_speedVectInMetresPerSecond Y', 'fv_speedVectInMetresPerSecond Z',
                'fv_localAccelInMetresPerSecond2 X', 'fv_localAccelInMetresPerSecond2 Y', 'fv_localAccelInMetresPerSecond2 Z',
                'fv_rotAccelInRadsPerSecond Pitch', 'fv_rotAccelInRadsPerSecond Roll', 'fv_turningCurvature',
                'fv_dragForce', 'fv_mass', 'fv_centerOfGravityHeight', 'fv_centerOfGravityPosition',
                'fv_rollAxisHeight', 'fv_trailer', 'fv_trailerAngle', 'fv_trailerPitchAngle', 'fv_trailerWheelbase',
                'fv_inIntersection', 'fv_laneDirection X', 'fv_laneDirection Y', 'fv_laneDirection Z', 'fv_laneCurvature',
                'fv_roadLongitudinalSlope', 'fv_roadLateralSlope',
                'fv_rightLaneOverLap'
            ]
            
            # 快速判断消息类型并选择解析方法
            if '=' in message and len(message.split('=')) == 2:
                try:
                    key, value = message.split('=', 1)
                    key = key.strip()
                    value = value.strip()
                    
                    # 忽略指定的UDP参数
                    if key in ignored_params:
                        return
                    
                    # 保留原始参数名（包括uv_和fv_开头的参数）
                    # 同时也添加映射后的常用名称，便于向后兼容
                    param_mapping = {
                        'uv_ID': 'ID', 'uv_Model': 'Model', 'uv_description': 'description', 
                        'uv_speed': 'speedInKmPerHour', 'uv_speedInKmPerHour': 'speedInKmPerHour',
                        'uv_distanceTravelled': 'distanceTravelled', 'uv_steering': 'steering', 'uv_brake': 'brake',
                        'uv_throttle': 'uv_throttle', 'uv_laneNumber': 'laneNumber'
                    }
                    
                    # 保存原始参数名
                    parsed_value = value
                    # 尝试转换数值类型
                    try:
                        if value.lower() not in ['true', 'false', 'n/a', '']:
                            parsed_value = float(value)
                    except (ValueError, TypeError):
                        pass
                    
                    # 保存原始键名
                    with self.driving_lock:
                        self.current_driving_params[key] = parsed_value
                    
                    # 如果有映射关系，也保存映射后的键名（向后兼容）
                    if key in param_mapping:
                        target_key = param_mapping[key]
                        with self.driving_lock:
                            self.current_driving_params[target_key] = parsed_value
                    
                    # 特殊处理速度字段，自动计算m/s
                    if key in ['uv_speed', 'uv_speedInKmPerHour'] and isinstance(parsed_value, (int, float)):
                        with self.driving_lock:
                            self.current_driving_params['speedInMetresPerSecond'] = parsed_value * 1000 / 3600
                    
                    has_parsed_data = True
                    
                except Exception:
                    # 静默处理，避免频繁错误日志
                    pass
            
            # 处理JSON格式
            elif '{' in message and '}' in message and ':' in message:
                try:
                    data = json.loads(message)
                    # 过滤掉要忽略的参数
                    filtered_data = {param: value for param, value in data.items() if param not in ignored_params}
                    with self.driving_lock:
                        for param, value in filtered_data.items():
                            self.current_driving_params[param] = value
                    has_parsed_data = has_parsed_data or len(filtered_data) > 0
                except (json.JSONDecodeError, TypeError):
                    # 不是有效的JSON，静默处理
                    pass
            
            # 处理SIMULATOR_LOG格式
            elif message.startswith('SIMULATOR_LOG'):
                try:
                    if ': ' in message:
                        parts = message.split(': ')[1].split(', ')
                        for part in parts:
                            if '=' in part:
                                try:
                                    key, value = part.split('=', 1)
                                    # 忽略指定的UDP参数
                                    if key in ignored_params:
                                        continue
                                    
                                    with self.driving_lock:
                                        self.current_driving_params[key] = value
                                        # 特殊处理已知的关键字段
                                        if 'Steering' in key:
                                            try:
                                                self.current_driving_params['steering'] = float(value)
                                            except ValueError:
                                                pass
                                        elif 'Brake' in key:
                                            try:
                                                self.current_driving_params['brake'] = float(value)
                                            except ValueError:
                                                pass
                                        elif 'Speed' in key:
                                            try:
                                                speed = float(value)
                                                self.current_driving_params['speedInKmPerHour'] = speed
                                                self.current_driving_params['speedInMetresPerSecond'] = speed * 1000 / 3600
                                            except ValueError:
                                                pass
                                    has_parsed_data = True
                                except Exception:
                                    pass
                except Exception:
                    # 静默处理，避免频繁错误日志
                    pass
            
            # 如果解析到了有用的参数，更新和处理
            if has_parsed_data:
                # 数据采样：每10个样本处理一次（100Hz等效）
                self.sample_counter += 1
                if self.sample_counter >= self.data_sampling_rate:
                    self.sample_counter = 0
                    
                    # 保存最后一组参数的副本
                    with self.driving_lock:
                        self.last_driving_params = self.current_driving_params.copy()
                    
                    # 显示由主线程 _tick_driving_display 定时刷新
                
                # 写入CSV缓存（保持原始频率）
                self._write_driving_params_to_csv()
        except Exception as e:
            # 记录错误但不频繁输出
            if time.time() - getattr(self, '_last_error_time', 0) > 5:  # 错误日志节流
                self._last_error_time = time.time()
                self.log(f"⚠ 解析驾驶参数时发生异常: {str(e)}")
    
    def _schedule_driving_display_tick(self):
        """在主线程中启动驾驶参数 UI 定时刷新，避免 LSL 线程直接调度导致点击卡死"""
        if getattr(self, 'root', None) and self.running:
            self.root.after(250, self._tick_driving_display)
    
    def _schedule_lightweight_status_tick(self):
        """每 1 秒仅刷新运行时间/数据时间/驾驶与生物信号条数，标明来源与 UXF 同源"""
        if getattr(self, 'root', None) and self.running:
            self.root.after(1000, self._tick_lightweight_status)
    
    def _tick_lightweight_status(self):
        """仅更新状态栏：运行时间、数据时间、驾驶数据 N 条 (来源:Unity LSL)、生物信号 N 条"""
        if not self.running or not getattr(self, 'root', None) or not hasattr(self, 'cache_status_var'):
            return
        try:
            with self.cache_lock:
                driving_count = len(self.driving_data_cache)
                biosig_count = len(self.biosig_data_cache)
            total_records = driving_count + biosig_count
            if self.system_start_time:
                system_runtime = time.time() - self.system_start_time
                h, r = divmod(system_runtime, 3600)
                m, s = divmod(r, 60)
                system_time_str = f"{int(h):02d}:{int(m):02d}:{int(s):02d}"
            else:
                system_time_str = "00:00:00"
            if getattr(self, 'actual_start_time', None):
                data_runtime = time.time() - self.actual_start_time
                h, r = divmod(data_runtime, 3600)
                m, s = divmod(r, 60)
                data_time_str = f"{int(h):02d}:{int(m):02d}:{int(s):02d}"
            else:
                data_time_str = "00:00:00"
            source_hint = " (来源:Unity LSL)" if driving_count > 0 else ""
            self.cache_status_var.set(
                f"运行时间: {system_time_str} | 数据时间: {data_time_str} | "
                f"驾驶数据: {driving_count} 条{source_hint} | 生物信号: {biosig_count} 条 | 总计: {total_records} 条"
            )
        except Exception:
            pass
        if self.running and getattr(self, 'root', None):
            self.root.after(1000, self._tick_lightweight_status)
    
    def _tick_driving_display(self):
        """主线程定时器：刷新驾驶参数面板（由 Unity LSL 数据驱动）+ 排出日志队列，避免 after(0) 堆积卡死"""
        if not self.running or not getattr(self, 'root', None):
            return
        try:
            self._display_driving_params()
        except Exception:
            pass
        try:
            self._drain_log_queue()
        except Exception:
            pass
        if self.running and getattr(self, 'root', None):
            self.root.after(250, self._tick_driving_display)
    
    def _drain_log_queue(self):
        """主线程调用：从 _log_queue 取出最多 20 条写入日志框，避免子线程频繁 after(0)"""
        if not getattr(self, 'log_text', None) or not getattr(self, '_log_queue', None):
            return
        n = 0
        while n < 20:
            try:
                log_message = self._log_queue.get_nowait()
            except queue.Empty:
                break
            n += 1
            try:
                self.log_text.insert(tk.END, log_message + "\n")
                self.log_text.see(tk.END)
                lines = int(self.log_text.index('end-1c').split('.')[0])
                if lines > 500:
                    self.log_text.delete(1.0, f"{lines-500}.0")
            except tk.TclError:
                break
    
    def _display_driving_params(self):
        """在UI中仅显示指定的驾驶参数 - 性能优化版"""
        # 1. 节流：使用更长的更新间隔，减少UI刷新频率
        now_ms = int(time.time() * 1000)
        if now_ms - getattr(self, 'last_ui_update_time_driving', 0) < self.ui_update_interval_driving:
            return
        self.last_ui_update_time_driving = now_ms
        
        # 2. 在锁的保护下，快速复制一份数据
        # 这样做可以避免在读取数据的同时，数据被其他线程修改
        params_copy = None
        with self.driving_lock:
            if self.current_driving_params:
                params_copy = self.current_driving_params.copy()
        
        # 主车里程：以 Unity LSL 发来的值为准（与 Inspector/UXF CSV 同源），不本地从 0 计算
        try:
            raw_dist = params_copy.get('uv_distanceTravelled') or params_copy.get('distanceTravelled')
            raw_speed = params_copy.get('uv_speedInKmPerHour') or params_copy.get('speedInKmPerHour') or 0
            if raw_dist is not None:
                raw_dist = float(raw_dist)
            else:
                raw_dist = None
            if raw_speed is not None:
                raw_speed = float(raw_speed)
            else:
                raw_speed = 0.0
            now_t = time.time()
            if not hasattr(self, '_display_odometer_last_raw'):
                self._display_odometer_last_raw = -1.0
                self._display_odometer = 0.0
                self._display_odometer_time = now_t
            # 有 LSL 里程时一律直接采用，保证“先跑 7km 再开采集”时显示与 Unity 一致
            if raw_dist is not None:
                self._display_odometer_last_raw = raw_dist
                self._display_odometer = raw_dist
                params_copy = params_copy.copy()
                params_copy['uv_distanceTravelled'] = raw_dist
            elif raw_speed > 0.5:
                # 仅当 LSL 未提供里程时用速度积分兜底
                dt = now_t - getattr(self, '_display_odometer_time', now_t)
                if dt > 0 and dt < 5.0:
                    self._display_odometer = getattr(self, '_display_odometer', 0) + (raw_speed / 3.6) * dt
                self._display_odometer_time = now_t
                params_copy = params_copy.copy()
                params_copy['uv_distanceTravelled'] = self._display_odometer
        except Exception:
            pass

        # 如果没有驾驶数据但有标记/缓存条数，显示占位提示（驾驶流未连接时）
        if not params_copy:
            cache_n = 0
            try:
                with self.cache_lock:
                    cache_n = len(self.driving_data_cache)
            except Exception:
                pass
            if cache_n > 0 and hasattr(self, 'root') and self.root and getattr(self, 'data_text', None) and self.data_text:
                def _show_placeholder():
                    try:
                        msg = (
                            "驾驶数据流未连接。\n\n"
                            "请确认：\n"
                            "1. Unity 场景中已开始试次（开始驾驶）\n"
                            "2. Unity 正在发送 UnityDrivingData 流\n\n"
                            f"已收到 {cache_n} 条事件标记（仅标记，无连续驾驶数据）。"
                        )
                        self.data_text.delete(1.0, tk.END)
                        self.data_text.insert(tk.END, msg)
                    except Exception:
                        pass
                self.root.after(0, _show_placeholder)
            return

        # 3. 智能更新检测：只在内容真正变化时更新UI
        # 比较当前参数与上次显示的参数是否有显著变化
        significant_changes = False
        current_params_summary = {}
        
        # 关键参数列表 - 只监控这些重要参数的变化
        key_params = [
            'uv_speedInKmPerHour', 'uv_distanceTravelled', 'uv_brake', 'uv_steering',
            'fv_speedInKmPerHour', 'fv_distanceAlongRoad', 'fv_throttle', 'fv_laneNumber'
        ]
        
        for key in key_params:
            val = None
            if key in params_copy:
                val = params_copy[key]
            elif key.replace('uv_', '') in params_copy:
                val = params_copy[key.replace('uv_', '')]
            elif 'uv_' + key in params_copy:
                val = params_copy['uv_' + key]
            
            if val is not None:
                try:
                    fv = float(val)
                    # 主车/前车里程：保留 1 位小数以便大数值时仍能检测到变化（避免约 33km 后显示卡住）
                    if key in ('uv_distanceTravelled', 'uv_distanceAlongRoad', 'fv_distanceAlongRoad'):
                        current_params_summary[key] = round(fv, 1)
                    else:
                        current_params_summary[key] = round(fv, 2)
                except Exception:
                    current_params_summary[key] = val
        
        # 检查是否有显著变化（使用缓存避免频繁比较）
        if not hasattr(self, '_last_params_summary'):
            self._last_params_summary = {}
            significant_changes = True
        else:
            # 比较关键参数是否发生变化
            for key, value in current_params_summary.items():
                if key not in self._last_params_summary or self._last_params_summary[key] != value:
                    significant_changes = True
                    break
        
        # 如果没有显著变化，跳过此次更新
        if not significant_changes:
            return
            
        self._last_params_summary = current_params_summary.copy()
        
        # 4. 定义一个在主线程运行的更新函数 - 优化版本
        def update_ui_safe():
            try:
                # 检查控件是否存在
                if not hasattr(self, 'data_text') or not self.data_text:
                    return

                display_lines = []
                
                # 添加系统时间（只在需要时获取）
                try:
                    display_lines.append(f"系统时间: {self.__class__.datetime.now().strftime('%H:%M:%S')}")
                except Exception:
                    pass
                
                # 尝试显示LSL时间戳
                if 'lsl_timestamp' in params_copy:
                    try:
                        display_lines.append(f"LSL时间: {float(params_copy['lsl_timestamp']):.3f}")
                    except:
                        pass
                
                # 优化的关键参数显示 - 只显示最重要的参数
                important_params = [
                    ('uv_speedInKmPerHour', '主车速度(km/h)'),
                    ('uv_distanceTravelled', '主车里程(m)'),
                    ('uv_brake', '主车刹车'),
                    ('uv_steering', '主车转向'),
                    ('fv_speedInKmPerHour', '前车速度(km/h)'),
                    ('fv_distanceAlongRoad', '前车距离(m)'),
                    ('fv_throttle', '前车油门'),
                    ('fv_laneNumber', '前车车道')
                ]
                
                display_count = 0
                for key, display_name in important_params:
                    val = None
                    # 尝试多种键名匹配方式
                    if key in params_copy:
                        val = params_copy[key]
                    elif key.replace('uv_', '') in params_copy:
                        val = params_copy[key.replace('uv_', '')]
                    elif 'uv_' + key in params_copy:
                        val = params_copy['uv_' + key]
                    
                    if val is not None and display_count < self.max_driving_display_lines:
                        try:
                            # 格式化数值，避免过长的小数
                            if isinstance(val, (float, int)):
                                if abs(float(val)) < 10:
                                    display_lines.append(f"{display_name}: {float(val):.2f}")
                                else:
                                    display_lines.append(f"{display_name}: {float(val):.1f}")
                            else:
                                display_lines.append(f"{display_name}: {val}")
                            display_count += 1
                        except:
                            pass
                
                # 智能更新：只在内容真正变化时执行昂贵的UI操作
                new_content = '\n'.join(display_lines)
                if hasattr(self, '_last_display_content') and self._last_display_content == new_content:
                    return  # 内容没有变化，跳过更新
                    
                self._last_display_content = new_content
                
                # 优化的文本更新：使用更高效的replace方法
                try:
                    # 如果文本框为空，直接插入；否则使用更高效的更新方式
                    current_content = self.data_text.get(1.0, tk.END).strip()
                    if not current_content:
                        self.data_text.insert(tk.END, new_content)
                    else:
                        # 使用delete和insert的组合，但尽量减少操作
                        self.data_text.delete(1.0, tk.END)
                        self.data_text.insert(tk.END, new_content)
                except tk.TclError:
                    # 控件可能已被销毁
                    return
                    
                # 只在必要时滚动到底部
                if display_lines:
                    try:
                        self.data_text.see(tk.END)
                    except tk.TclError:
                        pass
                
            except Exception as e:
                print(f"UI Update Error: {e}")

        # 4. 关键：使用 root.after 将任务调度到主线程执行
        if hasattr(self, 'root') and self.root:
            self.root.after(0, update_ui_safe)
    
    def _calculate_lsl_system_offset(self, biosig_timestamp):
        """计算LSL时间戳与系统时间的偏移量 - 修改为毫秒级精度显示"""
        if not PYLSL_AVAILABLE:
            return 0.0
            
        try:
            if self.last_driving_system_ts and self.last_biosig_system_ts:
                try:
                    ddt = datetime.strptime(self.last_driving_system_ts, "%Y-%m-%d %H:%M:%S.%f")
                    bdt = datetime.strptime(self.last_biosig_system_ts, "%Y-%m-%d %H:%M:%S.%f")
                    diff_ms_ss = (bdt.timestamp() - ddt.timestamp()) * 1000.0
                    if abs(diff_ms_ss) < 100.0:
                        self.sync_offsets.append(diff_ms_ss)
                        self.sync_samples_collected += 1
                        if hasattr(self, 'time_offset_var'):
                            self.time_offset_var.set(f"精度差: 校准中({self.sync_samples_collected}/{self.sync_samples_required})")
                        if self.sync_samples_collected >= self.sync_samples_required:
                            filtered_offsets = [offset for offset in self.sync_offsets if abs(offset) < 1000]
                            if len(filtered_offsets) > 0:
                                avg_offset_ms = sum(filtered_offsets) / len(filtered_offsets)
                                mean = avg_offset_ms
                                var = sum((o - mean) * (o - mean) for o in filtered_offsets) / len(filtered_offsets)
                                std_ms = var ** 0.5
                                self.time_offset = max(-5.0, min(5.0, avg_offset_ms / 1000.0))
                                self.time_sync_enabled = True
                                if hasattr(self, 'time_sync_status_var'):
                                    self.time_sync_status_var.set("时间同步: 已完成")
                                if hasattr(self, 'time_offset_var'):
                                    self.time_offset_var.set(f"精度差: {abs(avg_offset_ms):.3f} ms (jitter {std_ms:.3f} ms)")
                            else:
                                self.time_offset = 0.0
                                self.time_sync_enabled = True
                                if hasattr(self, 'time_sync_status_var'):
                                    self.time_sync_status_var.set("时间同步: 已完成")
                                if hasattr(self, 'time_offset_var'):
                                    self.time_offset_var.set(f"精度差: 0.000 ms (jitter 0.000 ms)")
                            return diff_ms_ss
                except Exception:
                    pass
            from pylsl import local_clock
            # 获取当前LSL时钟时间作为参考
            current_lsl_time = local_clock()
            # 获取当前系统时间
            current_system_time = time.time()
            
            # 重要：LSL时间和系统时间使用完全不同的时间基准
            # 绝对时间差异会非常大（可能达数百万秒），但我们只关心相对同步精度
            # 我们需要计算的是：在同一时间基准下，系统时间和LSL时间的相对偏移
            
            # 计算时间差（秒）
            raw_time_diff = current_lsl_time - current_system_time
            
            # 只关注近期的小范围时间差异（毫秒级）
            # 对于数据同步，我们需要的是两个时间流之间的相对偏移，而不是绝对基准差异
            # 这里我们将使用固定基数来表示相对偏移，避免显示异常大的数值
            
            # 转换为毫秒级精度
            time_diff_ms = raw_time_diff * 1000
            
            # 如果有生物信号时间戳（通常是LSL格式），计算两个LSL时间流的时间差
            if biosig_timestamp is not None:
                try:
                    # 两个LSL时间流的时间差：当前LSL时间 - 生物信号LSL时间
                    lsl_streams_time_diff = current_lsl_time - float(biosig_timestamp)
                    # 转换为毫秒显示
                    lsl_streams_time_diff_ms = lsl_streams_time_diff * 1000
                    self.log(f"ℹ LSL流时间差: {lsl_streams_time_diff_ms:.3f} ms")
                except Exception:
                    pass
            
            # 收集毫秒级偏移样本
            self.sync_offsets.append(time_diff_ms)
            self.sync_samples_collected += 1
            
            # 更新UI显示样本收集进度
            if hasattr(self, 'time_offset_var'):
                self.time_offset_var.set(f"精度差: 校准中({self.sync_samples_collected}/{self.sync_samples_required})")
            
            # 当收集足够样本时，计算平均偏移量
            if self.sync_samples_collected >= self.sync_samples_required:
                # 过滤异常大的偏移量值，只保留合理范围内的样本（±1000毫秒）
                # 这是因为LSL时间和系统时间基准不同，异常大的值不代表实际同步精度
                filtered_offsets = [offset for offset in self.sync_offsets if abs(offset) < 1000]
                
                # 如果过滤后还有足够的样本，使用过滤后的样本计算；否则安全回退为0
                if len(filtered_offsets) > 0:
                    avg_offset_ms = sum(filtered_offsets) / len(filtered_offsets)
                    self.log(f"✓ 时间同步已完成: 平均偏移量 = {avg_offset_ms:.3f} ms (已过滤异常值，使用 {len(filtered_offsets)}/{len(self.sync_offsets)} 个有效样本)")
                else:
                    avg_offset_ms = 0.0
                    self.log("⚠ 时间同步样本全部异常，已回退为0偏移")
                
                # 存储为秒用于计算，并进行安全夹取（±5秒）
                self.time_offset = max(-5.0, min(5.0, avg_offset_ms / 1000.0))
                self.time_sync_enabled = True
                
                # 更新UI显示最终时间精度差（毫秒显示）
                if hasattr(self, 'time_sync_status_var'):
                    self.time_sync_status_var.set("时间同步: 已完成")
                if hasattr(self, 'time_offset_var'):
                    self.time_offset_var.set(f"精度差: {abs(avg_offset_ms):.3f} ms")
            
            return time_diff_ms
        except Exception as e:
            self.log(f"⚠ 计算时间偏移量失败: {str(e)}")
            return 0.0
    
    def _synchronize_timestamps(self):
        """当两种数据都收到时，激活时间同步功能 - 使用统一的LSL时间格式"""
        # 只有当两种数据都收到时，才激活时间同步
        if not self.driving_received or not self.biosig_received:
            return
            
        # 时间同步已经激活，不需要重复激活
        if self.time_sync_enabled:
            return
            
        self.log("✓ 两种数据都已收到，开始激活LSL时间同步功能")
        
        # 更新UI状态
        if hasattr(self, 'time_sync_status_var'):
            self.time_sync_status_var.set("时间同步: 校准中")
        if hasattr(self, 'time_offset_var'):
            self.time_offset_var.set(f"精度差: 初始化")
        
        # 初始化时间同步过程
        if PYLSL_AVAILABLE:
            try:
                from pylsl import local_clock
                # 获取初始LSL时间作为参考
                current_lsl_time = local_clock()
                # 获取当前系统时间
                current_system_time = time.time()
                
                # 计算LSL与系统时间的真实偏移量 (LSL = System + Offset)
                # Offset = LSL - System
                self.time_offset = current_lsl_time - current_system_time
                
                # 初始时设置一个合理的小值，让系统开始收集样本
                initial_offset_ms = self.time_offset * 1000.0
                
                # 存储初始偏移量样本（毫秒级）
                self.sync_offsets.append(initial_offset_ms)
                self.sync_samples_collected = 1
                
                # 显示初始化信息，避免显示异常大的初始偏移量
                self.log(f"ℹ 时间同步初始化: 开始收集时间偏移样本")
                self.log(f"ℹ LSL时间基准和系统时间基准不同，将通过多次采样计算精确的相对偏移")
                self.log(f"ℹ 使用毫秒级精度进行时间同步校准")
                self.log(f"ℹ 收集 {self.sync_samples_required} 个样本以计算平均偏移量...")
                
                # 关键修复：初始化完成后设置为True，防止重复初始化导致卡顿
                self.time_sync_enabled = True
            except Exception as e:
                self.log(f"⚠ 初始化时间同步失败: {str(e)}")
    
    def _get_synchronized_timestamp(self, system_timestamp_str):
        """将系统时间戳统一转换为LSL时间格式并以毫秒级精度显示时间差"""
        if not self.time_sync_enabled:
            return system_timestamp_str, "N/A", "N/A"
            
        try:
            if not system_timestamp_str or system_timestamp_str == "N/A":
                return system_timestamp_str, "N/A", "N/A"
            # 将系统时间戳字符串转换为datetime对象
            dt = datetime.strptime(system_timestamp_str, "%Y-%m-%d %H:%M:%S.%f")
            # 转换为系统时间戳（秒）
            system_time = dt.timestamp()
            
            from datetime import timedelta
            lsl_time = system_time + (self.time_offset if isinstance(self.time_offset, (int, float)) else 0.0)
            synchronized_dt = dt + timedelta(seconds=(self.time_offset if isinstance(self.time_offset, (int, float)) else 0.0))
            synchronized_str = synchronized_dt.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
            
            # 计算并返回LSL时间格式的原始值（用于时间差计算）
            lsl_time_str = f"{lsl_time:.6f}"
            
            # 优化：减少日志记录频率，只在偏移量变化较大时记录
            time_diff_ms = self.time_offset * 1000
            # 只在偏移量超过1ms或每隔一定数量的样本才记录日志，减少日志开销
            if hasattr(self, '_last_sync_log_time'):
                current_time = time.time()
                # 暂时完全禁用时间戳同步日志，避免影响1000Hz高频采样性能
                # if current_time - self._last_sync_log_time > 5.0 or abs(time_diff_ms) > 10.0:
                #     self.log(f"ℹ 时间戳同步: 系统时间 → LSL时间 (偏移 {time_diff_ms:.3f} ms)")
                #     self._last_sync_log_time = current_time
            else:
                # 初始化日志记录时间
                self._last_sync_log_time = time.time()
                # 只在初始化时记录一次日志，避免高频调用
            
            return system_timestamp_str, synchronized_str, lsl_time_str
        except Exception as e:
            if isinstance(e, OSError) and getattr(e, "errno", None) == 22:
                return system_timestamp_str, "N/A", "N/A"
            self.log(f"⚠ 时间戳同步失败: {str(e)}")
            return system_timestamp_str, "N/A", "N/A"
    
    def _write_driving_params_to_csv(self):
        """将驾驶参数缓存到内存，不立即写入文件"""
        if not self.running:
            return
        
        try:
            # 使用独立的驾驶数据锁，避免与生物信号线程冲突
            with self.driving_lock:
                # 记录首次接收到数据的时间
                if self.actual_start_time is None:
                    self.actual_start_time = time.time()
                    self.log("✓ 开始接收和缓存数据")
                
                # 标记驾驶数据已收到
                self.driving_received = True
                
                # 尝试激活时间同步
                self._synchronize_timestamps()
                
                # 限速与平滑：仅当满足时间间隔才输出，并使用严格等间隔的平滑时间戳
                now_dt = self.__class__.datetime.now()
                if self._last_driving_emit_time is None:
                    scheduled_dt = now_dt
                    self._last_driving_emit_time = scheduled_dt
                else:
                    elapsed_ms = (now_dt - self._last_driving_emit_time).total_seconds() * 1000.0
                    if elapsed_ms < self.driving_min_interval_ms:
                        return
                    # 使用上次时间 + 固定步长，保证时间连续与平滑
                    # 修复：如果滞后超过100ms，则重置为当前时间，防止长时间运行导致的累积误差
                    from datetime import timedelta
                    if elapsed_ms > 100:  # 允许最大100ms的滞后
                         scheduled_dt = now_dt
                    else:
                         scheduled_dt = self._last_driving_emit_time + timedelta(milliseconds=self.driving_min_interval_ms)
                    self._last_driving_emit_time = scheduled_dt
                system_timestamp = scheduled_dt.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
                self.last_driving_system_ts = system_timestamp
                try:
                    self.recent_driving_ts.append(system_timestamp)
                except Exception:
                    pass
                
                # 获取同步后的时间戳（如果启用）
                original_timestamp, synchronized_timestamp, lsl_time_str = self._get_synchronized_timestamp(system_timestamp)
                
                # 创建数据行，使用uv_/fv_列头
                target_cols = [
                    "uv_ID","uv_Time","uv_Model","uv_description","uv_position X","uv_position Y","uv_position Z",
                    "uv_distanceTravelled","uv_distanceAlongRoad","uv_steering","uv_throttle","uv_brake","uv_lightState",
                    "uv_automaticControl","uv_wheelBase","uv_road","uv_distanceToLeftBorder","uv_distanceToRightBorder",
                    "uv_offsetFromRoadCenter","uv_offsetFromLaneCenter","uv_laneNumber","uv_laneWidth","uv_drivingForwards",
                    "uv_speedLimit","uv_speedOver",
                    "fv_ID","fv_Time","fv_Model","fv_description","fv_position X","fv_position Y","fv_position Z",
                    "fv_distanceTravelled","fv_steering","fv_throttle","fv_brake","fv_lightState","fv_automaticControl",
                    "fv_wheelBase","fv_road","fv_distanceAlongRoad","fv_distanceToLeftBorder","fv_distanceToRightBorder",
                    "fv_offsetFromRoadCenter","fv_offsetFromLaneCenter","fv_laneNumber","fv_laneWidth","fv_drivingForwards",
                    "fv_speedLimit","fv_speedOver"
                ]
                row = {k: "N/A" for k in target_cols}
                row["system_timestamp"] = original_timestamp
                row["synchronized_timestamp"] = synchronized_timestamp if self.time_sync_enabled else "N/A"
                # 关键修复：添加LSL原始时间戳，用于与生理数据毫秒级对齐
                row["lsl_timestamp"] = self.current_driving_params.get('lsl_timestamp', 'N/A')
                
                # 标记列默认
                row['marker_flag'] = False
                row['marker_type'] = "N/A"
                row['marker_condition'] = "N/A"
                row['marker_target'] = "N/A"
                row['marker_color'] = "N/A"
                row['marker_label'] = "N/A"
                
                # 遍历所有当前驾驶参数
                for param, value in self.current_driving_params.items():
                    if param in row:
                        try:
                            row[param] = float(value)
                        except (ValueError, TypeError):
                            row[param] = value
                # 基于非前缀键的回填映射到uv_
                fallback_map = {
                    'ID':'uv_ID','Model':'uv_Model','description':'uv_description','position X':'uv_position X','position Y':'uv_position Y','position Z':'uv_position Z',
                    'distanceTravelled':'uv_distanceTravelled','distanceAlongRoad':'uv_distanceAlongRoad','steering':'uv_steering','throttle':'uv_throttle','brake':'uv_brake',
                    'lightState':'uv_lightState','automaticControl':'uv_automaticControl','wheelBase':'uv_wheelBase','road':'uv_road','distanceToLeftBorder':'uv_distanceToLeftBorder',
                    'distanceToRightBorder':'uv_distanceToRightBorder','offsetFromRoadCenter':'uv_offsetFromRoadCenter','offsetFromLaneCenter':'uv_offsetFromLaneCenter',
                    'laneNumber':'uv_laneNumber','laneWidth':'uv_laneWidth','drivingForwards':'uv_drivingForwards','speedLimit':'uv_speedLimit','speedOver':'uv_speedOver'
                }
                for k_src,k_dst in fallback_map.items():
                    if k_src in self.current_driving_params and k_dst in row:
                        v = self.current_driving_params[k_src]
                        try:
                            row[k_dst] = float(v)
                        except (ValueError, TypeError):
                            row[k_dst] = v

                # 移除距离计算标记逻辑，仅使用LSL透传标记

                # 优化：将所有浮点数格式化为3位小数，减少内存和文件大小
                for k, v in row.items():
                    if isinstance(v, float):
                        row[k] = f"{v:.3f}"
                
                # 决定是否流式写入：如果启用了流式写入 或者 实验已经开始，则写入队列
                should_stream_write = getattr(self, "streaming_write_enabled", False) or self.experiment_started
                
                if should_stream_write:
                    try:
                        self._driving_stream_write_queue.put_nowait(row.copy())
                    except queue.Full:
                        try:
                            self._driving_queue_dropped = getattr(self, '_driving_queue_dropped', 0) + 1
                        except Exception:
                            pass
                        try:
                            _ = self._driving_stream_write_queue.get_nowait()
                        except Exception:
                            pass
                        try:
                            self._driving_stream_write_queue.put_nowait(row.copy())
                        except Exception:
                            pass
                
                # 无论是否流式写入，都同时写入内存缓存，确保点击停止时能可靠保存
                # （流式只写队列时，若停止时队列未排空或文件未创建，会导致“没有驾驶数据需要保存”）
                with self.cache_lock:
                    self.driving_data_cache.append(row.copy())
                    try:
                        self._driving_total_samples = getattr(self, '_driving_total_samples', 0) + 1
                    except Exception:
                        pass
                    try:
                        keep_n = int(getattr(self, 'cache_keep_last_n', 0) or 0)
                        if getattr(self, "streaming_write_enabled", False) and keep_n > 0 and len(self.driving_data_cache) > keep_n:
                            del self.driving_data_cache[:-keep_n]
                    except Exception:
                        pass
                
                # 不再在 LSL 线程内调用 _update_cache_status，避免主线程被大量 after(0) 拖垮导致点击卡死；状态由 _start_cache_monitoring 定时更新
                
                # 2秒节流：避免在“只保留最近N条缓存”时缓存长度恒定导致疯狂刷屏
                now_t = time.time()
                try:
                    last_t = getattr(self, '_last_driving_cache_status_log_t', 0.0)
                except Exception:
                    last_t = 0.0
                try:
                    total_samples = getattr(self, '_driving_total_samples', 0)
                except Exception:
                    total_samples = 0
                if (now_t - last_t) >= 2.0:
                    try:
                        self._last_driving_cache_status_log_t = now_t
                    except Exception:
                        pass
                    if self.actual_start_time:
                        actual_runtime = now_t - self.actual_start_time
                        sync_status = "已同步" if self.time_sync_enabled else "未同步"
                        self.log(f"[缓存状态] 驾驶累计{total_samples}条 | 内存保留{len(self.driving_data_cache)}条 | 实际记录{actual_runtime:.1f}秒 | {sync_status}", is_debug=True)
                    else:
                        self.log(f"[缓存状态] 驾驶累计{total_samples}条 | 内存保留{len(self.driving_data_cache)}条", is_debug=True)
                    
        except Exception as e:
            error_info = traceback.format_exc()
            self.log(f"⚠ 缓存驾驶参数时发生错误: {str(e)}")
            self.log(f"  错误详情: {error_info.splitlines()[-1]}")
    
    def _start_cache_monitoring(self):
        """启动缓存状态监控，支持长时间稳定运行"""
        if not self.running:
            return
            
        # 更新缓存状态
        self._update_cache_status()
        
        current_time = time.time()
        
        # 检查是否需要备份
        if current_time - self.last_backup_time > self.backup_interval:
            self._backup_cached_data()
            self.last_backup_time = current_time
        
        # 检查是否需要自动刷新缓存到磁盘（防止内存溢出）
        # self.streaming_write_enabled = True
        # self._high_pressure_mode = True
        
        # 检查缓存大小，如果超过最大限制，执行强制刷新
        if hasattr(self, 'max_cache_size'):
            total_cache_size = len(self.driving_data_cache) + len(self.biosig_data_cache)
            if total_cache_size > self.max_cache_size:
                # 避免频繁触发刷新，设置冷却时间
                if not hasattr(self, '_last_cache_flush_time'):
                    self._last_cache_flush_time = 0
                
                current_time = time.time()
                if current_time - self._last_cache_flush_time >= 5:  # 5秒冷却，加快刷新频率
                    self._last_cache_flush_time = current_time
                    self.log(f"[内存保护] 缓存大小({total_cache_size})超过最大限制，执行强制刷新（异步）")
                    # 异步刷新，避免主线程卡顿
                    self._flush_cache_data_async()
                else:
                    self.log(f"[内存保护] 缓存({total_cache_size})超限制，但处于冷却期，跳过刷新", is_debug=True)
        
        # 定期UI内存清理 - 每15秒执行一次（90分钟实验专用）
        if not hasattr(self, '_last_ui_cleanup_time'):
            self._last_ui_cleanup_time = current_time
        
        if current_time - self._last_ui_cleanup_time >= 15:
            self._perform_ui_cleanup()
            self._last_ui_cleanup_time = current_time
            
        # 90分钟实验：每5分钟强制垃圾回收
        if not hasattr(self, '_last_aggressive_gc_time'):
            self._last_aggressive_gc_time = current_time
            
        if current_time - self._last_aggressive_gc_time >= 300:  # 5分钟
            self._perform_aggressive_cleanup()
            self._last_aggressive_gc_time = current_time
        
        # 继续监控，降低频率到每5秒一次
        self.root.after(5000, self._start_cache_monitoring)
    
    def _perform_ui_cleanup(self):
        """定期清理UI内存，防止长时间运行导致内存泄漏"""
        try:
            # 清理驾驶参数UI缓存
            if hasattr(self, '_last_params_summary'):
                # 保留最近的数据，但清理临时变量
                if len(self._last_params_summary) > 50:
                    # 如果缓存过大，清理一半的旧数据
                    keys_to_keep = list(self._last_params_summary.keys())[-25:]
                    self._last_params_summary = {k: self._last_params_summary[k] for k in keys_to_keep}
            
            # 清理显示内容缓存
            if hasattr(self, '_last_display_content') and len(getattr(self, '_last_display_content', '')) > 1000:
                self._last_display_content = ""
            
            # 清理ScrolledText的历史记录（如果支持）
            if hasattr(self, 'data_text') and self.data_text:
                try:
                    # 获取当前行数
                    total_lines = int(self.data_text.index('end-1c').split('.')[0])
                    if total_lines > 100:  # 如果超过100行
                        # 删除前50%的旧内容，保留最新的
                        lines_to_keep = total_lines // 2
                        self.data_text.delete(f"1.0", f"{lines_to_keep}.0")
                        self.log(f"[UI清理] 清理了 {lines_to_keep} 行旧数据，当前行数: {total_lines - lines_to_keep}", is_debug=True)
                except (tk.TclError, AttributeError):
                    # 忽略控件相关的错误
                    pass
            
            # 强制垃圾回收（谨慎使用，每几分钟执行一次）
            if not hasattr(self, '_last_gc_time'):
                self._last_gc_time = time.time()
            
            current_time = time.time()
            if current_time - self._last_gc_time >= 300:  # 每5分钟执行一次GC
                import gc
                gc.collect()
                self._last_gc_time = current_time
                self.log("[内存管理] 执行了垃圾回收", is_debug=True)
                
        except Exception as e:
             self.log(f"⚠ UI清理时发生错误: {str(e)}", is_debug=True)
    
    def _perform_aggressive_cleanup(self):
        """90分钟实验专用：激进的内存清理"""
        try:
            self.log("[90分钟实验] 执行激进内存清理...", is_debug=True)
            
            # 1. 强制清理所有临时缓存
            if hasattr(self, '_last_params_summary'):
                self._last_params_summary.clear()
            if hasattr(self, '_last_display_content'):
                self._last_display_content = ""
            if hasattr(self, '_last_biosig_summary'):
                self._last_biosig_summary.clear()
            
            # 2. 清理ScrolledText的所有历史记录
            if hasattr(self, 'data_text') and self.data_text:
                try:
                    # 完全清空并重新创建内容
                    self.data_text.delete(1.0, tk.END)
                    self.data_text.insert(tk.END, "=== 内存清理完成 ===\n")
                except tk.TclError:
                    pass
            
            # 3. 清理日志文本（保留最后10行）
            if hasattr(self, 'log_text') and self.log_text:
                try:
                    total_lines = int(self.log_text.index('end-1c').split('.')[0])
                    if total_lines > 50:
                        # 只保留最后10行
                        self.log_text.delete(f"1.0", f"{total_lines-10}.0")
                except (tk.TclError, AttributeError):
                    pass
            
            # 4. 强制垃圾回收和内存整理
            import gc
            gc.collect()
            
            # 5. 清理生物信号显示
            if hasattr(self, 'biosig_text') and self.biosig_text:
                try:
                    self.biosig_text.delete(1.0, tk.END)
                    self.biosig_text.insert(tk.END, "=== 生物信号显示清理 ===\n")
                except tk.TclError:
                    pass
            
            # 6. 清理眼动数据显示
            if hasattr(self, 'eye_text') and self.eye_text:
                try:
                    self.eye_text.delete(1.0, tk.END)
                    self.eye_text.insert(tk.END, "=== 眼动数据显示清理 ===\n")
                except tk.TclError:
                    pass
            
            # 7. 清理 LSL Plotter
            if hasattr(self, 'lsl_plotter') and self.lsl_plotter:
                 try:
                     self.lsl_plotter.reset_buffer()
                 except Exception:
                     pass

            self.log("[90分钟实验] 激进内存清理完成", is_debug=True)
            
        except Exception as e:
            self.log(f"⚠ 激进清理失败: {str(e)}", is_debug=True)
        
    def _flush_cache_data(self):
        """安全地将缓存数据刷新到磁盘并重置缓存，用于长时间运行时的内存管理"""
        try:
            # 获取当前时间戳作为批处理标识
            batch_timestamp = self.__class__.datetime.now().strftime("%Y%m%d_%H%M%S")
            
            # 使用缓存锁确保数据一致性
            with self.cache_lock:
                # 先备份当前缓存数据，确保安全
                self._backup_cached_data()
                
                # 检查是否有驾驶数据需要刷新
                if self.driving_data_cache:
                    self.log(f"正在刷新驾驶数据缓存...")
                    # 创建临时文件名
                    temp_filename = f"driving_batch_{batch_timestamp}.csv"
                    temp_file_path = os.path.join(self.data_folder, temp_filename)
                    
                    # 获取CSV列名（uv_/fv_版）
                    csv_columns = [
                        "system_timestamp","synchronized_timestamp","lsl_timestamp",
                        "marker_flag","marker_type","marker_condition","marker_target","marker_color","marker_label",
                        "uv_ID","uv_Time","uv_Model","uv_description","uv_position X","uv_position Y","uv_position Z",
                        "uv_distanceTravelled","uv_distanceAlongRoad","uv_steering","uv_throttle","uv_brake","uv_lightState",
                        "uv_automaticControl","uv_wheelBase","uv_road","uv_distanceToLeftBorder","uv_distanceToRightBorder",
                        "uv_offsetFromRoadCenter","uv_offsetFromLaneCenter","uv_laneNumber","uv_laneWidth","uv_drivingForwards",
                        "uv_speedLimit","uv_speedOver",
                        "fv_ID","fv_Time","fv_Model","fv_description","fv_position X","fv_position Y","fv_position Z",
                        "fv_distanceTravelled","fv_steering","fv_throttle","fv_brake","fv_lightState","fv_automaticControl",
                        "fv_wheelBase","fv_road","fv_distanceAlongRoad","fv_distanceToLeftBorder","fv_distanceToRightBorder",
                        "fv_offsetFromRoadCenter","fv_offsetFromLaneCenter","fv_laneNumber","fv_laneWidth","fv_drivingForwards",
                        "fv_speedLimit","fv_speedOver"
                    ]
                    
                    # 写入临时文件
                    with open(temp_file_path, 'w', newline='', encoding='utf-8-sig') as temp_file:
                        writer = csv.DictWriter(temp_file, fieldnames=csv_columns, quoting=csv.QUOTE_NONNUMERIC)
                        writer.writeheader()
                        count = 0
                        for row in self.driving_data_cache:
                            formatted_row = {col: row.get(col, "N/A") for col in csv_columns}
                            writer.writerow(formatted_row)
                            count += 1
                            if count % self.flush_batch_size_driving == 0:
                                temp_file.flush()
                    
                    self.log(f"已将 {len(self.driving_data_cache)} 条驾驶数据刷新到: {temp_filename}")
                    # 清空已处理的缓存
                    self.driving_data_cache = []
                
                # 检查是否有生物信号数据需要刷新
                if self.biosig_data_cache:
                    self.log(f"正在刷新生物信号数据缓存...")
                    # 创建临时文件名
                    temp_filename = f"biosig_batch_{batch_timestamp}.csv"
                    temp_file_path = os.path.join(self.data_folder, temp_filename)
                    
                    # 写入临时文件
                    with open(temp_file_path, 'w', newline='', encoding='utf-8-sig') as temp_file:
                        if self.biosig_data_cache:
                            fieldnames = list(self.biosig_data_cache[0].keys())
                            writer = csv.DictWriter(temp_file, fieldnames=fieldnames, quoting=csv.QUOTE_NONNUMERIC)
                            writer.writeheader()
                            count = 0
                            for row in self.biosig_data_cache:
                                writer.writerow(row)
                                count += 1
                                if count % self.flush_batch_size_biosig == 0:
                                    temp_file.flush()
                    
                    self.log(f"已将 {len(self.biosig_data_cache)} 条生物信号数据刷新到: {temp_filename}")
                    # 清空已处理的缓存
                    self.biosig_data_cache = []
                    
                self.log("✓ 缓存刷新完成")

        except Exception as e:
            error_info = traceback.format_exc()
            self.log(f"⚠ 刷新缓存数据时发生错误: {str(e)}")
            self.log(f"  错误详情: {error_info.splitlines()[-1]}")
            # 发生错误时不清除缓存，以免数据丢失
    
    def _flush_cache_data_progressive(self):
        """渐进式缓存刷新，避免一次性处理大量数据导致卡死"""
        try:
            self.log(f"[渐进式刷新] 开始处理缓存数据...")
            
            # 获取当前缓存大小
            with self.cache_lock:
                driving_count = len(self.driving_data_cache)
                biosig_count = len(self.biosig_data_cache)
                total_count = driving_count + biosig_count
            
            if total_count == 0:
                return
            
            self.log(f"[渐进式刷新] 待处理数据: 驾驶{driving_count}条, 生物{biosig_count}条")
            
            # 分批次处理，每批最多5000条
            batch_size = 5000
            processed_total = 0
            
            # 处理驾驶数据
            if driving_count > 0:
                batches = (driving_count + batch_size - 1) // batch_size
                for batch_idx in range(batches):
                    start_idx = batch_idx * batch_size
                    end_idx = min(start_idx + batch_size, driving_count)
                    
                    self.log(f"[渐进式刷新] 处理驾驶数据批次 {batch_idx+1}/{batches}")
                    
                    # 分批处理驾驶数据
                    with self.cache_lock:
                        batch_data = self.driving_data_cache[start_idx:end_idx]
                        # 立即从缓存中移除已处理的数据
                        self.driving_data_cache = self.driving_data_cache[end_idx:]
                    
                    # 处理这一批数据
                    self._process_driving_batch(batch_data)
                    processed_total += len(batch_data)
                    
                    # 给系统一点时间处理其他任务
                    time.sleep(0.01)  # 10ms延迟
            
            # 处理生物信号数据
            if biosig_count > 0:
                batches = (biosig_count + batch_size - 1) // batch_size
                for batch_idx in range(batches):
                    start_idx = batch_idx * batch_size
                    end_idx = min(start_idx + batch_size, biosig_count)
                    
                    self.log(f"[渐进式刷新] 处理生物信号批次 {batch_idx+1}/{batches}")
                    
                    # 分批处理生物信号数据
                    with self.cache_lock:
                        batch_data = self.biosig_data_cache[start_idx:end_idx]
                        # 立即从缓存中移除已处理的数据
                        self.biosig_data_cache = self.biosig_data_cache[end_idx:]
                    
                    # 处理这一批数据
                    self._write_biosig_batch(batch_data)
                    processed_total += len(batch_data)
                    
                    # 给系统一点时间处理其他任务
                    time.sleep(0.01)  # 10ms延迟
            
            self.log(f"[渐进式刷新] 完成，共处理 {processed_total} 条数据")
            
            # 更新缓存状态
            self._update_cache_status()
            
        except Exception as e:
            self.log(f"⚠ 渐进式刷新失败: {str(e)}")
            # 出错时等待一段时间后重试
            time.sleep(1)
            # 递归重试，但限制重试次数
            if not hasattr(self, '_flush_retry_count'):
                self._flush_retry_count = 0
    
    def _process_driving_batch(self, batch_data):
        """处理一批驾驶数据并写入文件"""
        if not batch_data:
            return
            
        try:
            # 使用Trial文件名
            filename = self.current_trial_filenames['driving']
            file_path = os.path.join(self.data_folder, filename)
            
            # 获取CSV列名
            csv_columns = [
                "system_timestamp","synchronized_timestamp","lsl_timestamp",
                "marker_flag","marker_type","marker_condition","marker_target","marker_color","marker_label",
                "uv_ID","uv_Time","uv_Model","uv_description","uv_position X","uv_position Y","uv_position Z",
                "uv_distanceTravelled","uv_distanceAlongRoad","uv_steering","uv_throttle","uv_brake","uv_lightState",
                "uv_automaticControl","uv_wheelBase","uv_road","uv_distanceToLeftBorder","uv_distanceToRightBorder",
                "uv_offsetFromRoadCenter","uv_offsetFromLaneCenter","uv_laneNumber","uv_laneWidth","uv_drivingForwards",
                "uv_speedLimit","uv_speedOver",
                "fv_ID","fv_Time","fv_Model","fv_description","fv_position X","fv_position Y","fv_position Z",
                "fv_distanceTravelled","fv_steering","fv_throttle","fv_brake","fv_lightState","fv_automaticControl",
                "fv_wheelBase","fv_road","fv_distanceAlongRoad","fv_distanceToLeftBorder","fv_distanceToRightBorder",
                "fv_offsetFromRoadCenter","fv_offsetFromLaneCenter","fv_laneNumber","fv_laneWidth","fv_drivingForwards",
                "fv_speedLimit","fv_speedOver"
            ]
            
            # 写入文件 - 始终使用追加模式 'a'
            with open(file_path, 'a', newline='', encoding='utf-8-sig') as f:
                writer = csv.DictWriter(f, fieldnames=csv_columns, quoting=csv.QUOTE_NONNUMERIC)
                
                # 如果是该Trial的第一次写入，先写入表头
                if not self.trial_files_initialized['driving']:
                    writer.writeheader()
                    self.trial_files_initialized['driving'] = True
                
                # 写入数据
                for row in batch_data:
                    formatted_row = {col: row.get(col, "N/A") for col in csv_columns}
                    writer.writerow(formatted_row)
            
            self.log(f"[批次写入] 已写入 {len(batch_data)} 条驾驶数据到 {filename}", is_debug=True)
            
        except Exception as e:
            self.log(f"⚠ 驾驶数据批次写入失败: {str(e)}")
            raise
    
    def _write_biosig_batch(self, batch_data):
        """处理一批生物信号数据并写入文件"""
        if not batch_data:
            return
            
        try:
            # 使用Trial文件名
            filename = self.current_trial_filenames['biosig']
            file_path = os.path.join(self.data_folder, filename)
            
            csv_columns = ["system_timestamp","synchronized_timestamp","biosig_timestamp","heart_rate_bpm","gsr_uS","event_system_timestamp","event_synchronized_timestamp","marker_flag","marker_type","marker_condition","marker_target","marker_color","marker_label"]
            with open(file_path, 'a', newline='', encoding='utf-8-sig') as f:
                writer = csv.DictWriter(f, fieldnames=csv_columns, quoting=csv.QUOTE_NONNUMERIC)
                if not self.trial_files_initialized['biosig']:
                    writer.writeheader()
                    self.trial_files_initialized['biosig'] = True
                for row in batch_data:
                    writer.writerow({k: row.get(k, "N/A") for k in csv_columns})
            
            self.log(f"[批次写入] 已写入 {len(batch_data)} 条生物信号数据到 {filename}", is_debug=True)
            
        except Exception as e:
            self.log(f"⚠ 生物信号数据批次写入失败: {str(e)}")
            raise
            self._flush_retry_count += 1
            if self._flush_retry_count <= 3:
                self.log(f"[渐进式刷新] 第{self._flush_retry_count}次重试...")
                self._flush_cache_data_progressive()
            else:
                self.log("[渐进式刷新] 达到最大重试次数，放弃刷新")
                self._flush_retry_count = 0

    def _flush_cache_data_async(self):
        """优化版异步缓存刷新，防止UI卡死"""
        if self._flush_in_progress:
            return
        self._flush_in_progress = True
        
        # 立即返回，不等待任何UI更新
        def _run():
            try:
                # 分批处理，避免一次性处理过多数据
                self._flush_cache_data_progressive()
            except Exception as e:
                # 在后台线程中记录错误，不影响UI
                print(f"异步刷新错误: {e}")
            finally:
                self._flush_in_progress = False
        
        # 使用守护线程，确保不会阻塞主程序退出
        flush_thread = threading.Thread(target=_run, daemon=True)
        flush_thread.start()
    
    def _backup_cached_data(self):
        """创建缓存数据的临时备份，防止程序崩溃数据丢失"""
        if not self.driving_data_cache and not self.biosig_data_cache:
            return
            
        try:
            backup_time = self.__class__.datetime.now().strftime("%Y%m%d_%H%M%S")
            # 将备份文件保存在缓存文件夹中
            backup_filename = os.path.join(self.cache_folder, f"cache_backup_{backup_time}.json")
            
            # 使用缓存锁读取缓存数据
            with self.cache_lock:
                driving_count = len(self.driving_data_cache)
                biosig_count = len(self.biosig_data_cache)
                # 只备份数据的摘要信息，避免占用过多磁盘空间
                backup_data = {
                    "backup_time": backup_time,
                    "cache_start_time": str(self.cache_start_time),
                    "driving_record_count": driving_count,
                    "biosig_record_count": biosig_count,
                    "total_record_count": driving_count + biosig_count
                }
                
                # 添加驾驶数据样本
                if self.driving_data_cache:
                    backup_data["driving_sample_records"] = self.driving_data_cache[:5] if driving_count > 5 else self.driving_data_cache
                    backup_data["latest_driving_record"] = self.driving_data_cache[-1] if driving_count > 0 else None
                
                # 添加生物信号数据样本
                if self.biosig_data_cache:
                    backup_data["biosig_sample_records"] = self.biosig_data_cache[:5] if biosig_count > 5 else self.biosig_data_cache
                    backup_data["latest_biosig_record"] = self.biosig_data_cache[-1] if biosig_count > 0 else None
            
            # 写入备份文件
            with open(backup_filename, 'w', encoding='utf-8') as f:
                json.dump(backup_data, f, ensure_ascii=False, indent=2)
            
            self.log(f"✓ 已创建缓存数据备份: {os.path.basename(backup_filename)}")
        except Exception as e:
            self.log(f"⚠ 备份缓存数据失败: {str(e)}")
            
    def _update_cache_status(self):
        """更新缓存状态显示 - 高频数据优化版本"""
        current_time = time.time()
        current_time_ms = current_time * 1000  # 转换为毫秒
        
        # 1. 检查是否需要备份缓存数据（独立于UI更新）
        if current_time - self.last_backup_time >= self.backup_interval:
            self._backup_cached_data()
            self.last_backup_time = current_time
        
        # 2. 节流控制：使用独立的缓存状态UI更新控制
        if current_time_ms - self.last_ui_update_time_cache >= self.ui_update_interval:
            # 快速获取缓存状态，最小化锁持有时间
            driving_count = 0
            biosig_count = 0
            with self.cache_lock:
                driving_count = len(self.driving_data_cache)
                biosig_count = len(self.biosig_data_cache)
            total_records = driving_count + biosig_count
            # 背压与自动切换写入策略
            if not getattr(self, '_flush_in_progress', False) and total_records > self.soft_cache_limit:
                try:
                    self._flush_cache_data_async()
                except Exception:
                    pass
            if total_records > self.hard_cache_limit:
                self.streaming_write_enabled = True
                self._high_pressure_mode = True
            elif total_records < (self.soft_cache_limit // 2) and self.streaming_write_enabled and not getattr(self, '_flush_in_progress', False):
                pass
            
            # 减少时间计算频率，每2次UI更新才计算一次时间
            if hasattr(self, '_last_time_calculation') and \
               current_time_ms - self._last_time_calculation < self.ui_update_interval * 2:
                # 复用上次计算的时间字符串，减少计算开销
                pass
            else:
                # 计算系统运行时间
                if self.system_start_time:
                    system_runtime = current_time - self.system_start_time
                    hours, remainder = divmod(system_runtime, 3600)
                    minutes, seconds = divmod(remainder, 60)
                    self.system_time_str = f"{int(hours):02d}:{int(minutes):02d}:{int(seconds):02d}"
                else:
                    self.system_time_str = "00:00:00"
                
                # 计算数据收集时间
                if self.actual_start_time:
                    data_runtime = current_time - self.actual_start_time
                    data_hours, data_remainder = divmod(data_runtime, 3600)
                    data_minutes, data_seconds = divmod(data_remainder, 60)
                    self.data_time_str = f"{int(data_hours):02d}:{int(data_minutes):02d}:{int(data_seconds):02d}"
                else:
                    self.data_time_str = "00:00:00"
                
                self._last_time_calculation = current_time_ms
            
            # 4. 检查LSL Plotter是否造成阻塞 (如果绘图缓冲区过大)
            if hasattr(self, 'lsl_plotter') and self.lsl_plotter and self.lsl_plotter.running:
                # 简单检查：如果队列积压严重，尝试暂停绘图
                bwq = self._biosig_stream_write_queue.qsize() if hasattr(self, '_biosig_stream_write_queue') else 0
                if bwq > 5000:
                    # self.log("⚠ 警告：生物信号队列积压严重，暂时降低绘图频率")
                    self.lsl_plotter.update_interval = 200 # 降低到5Hz
                elif bwq < 100:
                    self.lsl_plotter.update_interval = 50  # 恢复20Hz
            # 简化UI更新函数（驾驶数据>0 时标明与 UXF 同源）
            source_hint = " (来源:Unity LSL)" if driving_count > 0 else ""
            def update_ui():
                self.cache_status_var.set(
                    f"运行时间: {self.system_time_str} | 数据时间: {self.data_time_str} | "
                    f"驾驶数据: {driving_count} 条{source_hint} | 生物信号: {biosig_count} 条 | 总计: {total_records} 条"
                )
                if hasattr(self, 'driving_cache_status_var'):
                    self.driving_cache_status_var.set(f"驾驶数据缓存: {driving_count} 条")
                if hasattr(self, 'biosig_cache_status_var'):
                    self.biosig_cache_status_var.set(f"生物信号缓存: {biosig_count} 条")
                try:
                    dq_len = len(self.data_queue) if hasattr(self, 'data_queue') else 0
                    dwq = self._driving_stream_write_queue.qsize() if hasattr(self, '_driving_stream_write_queue') else 0
                    bwq = self._biosig_stream_write_queue.qsize() if hasattr(self, '_biosig_stream_write_queue') else 0
                    ewq = self._eye_stream_write_queue.qsize() if hasattr(self, '_eye_stream_write_queue') else 0
                    self.queue_status_var.set(f"队列状态: 驾驶Q={dq_len}, 写入Q(驾/生/眼)={dwq}/{bwq}/{ewq}")
                    da = self._driving_stream_writer_thread.is_alive() if hasattr(self, '_driving_stream_writer_thread') and self._driving_stream_writer_thread else False
                    ba = self._biosig_stream_writer_thread.is_alive() if hasattr(self, '_biosig_stream_writer_thread') and self._biosig_stream_writer_thread else False
                    ea = self._eye_stream_writer_thread.is_alive() if hasattr(self, '_eye_stream_writer_thread') and self._eye_stream_writer_thread else False
                    self.writer_status_var.set(f"写入线程: 驾驶Alive={da}, 生理Alive={ba}, 眼动Alive={ea}, 流式写入={self.streaming_write_enabled}")
                except Exception:
                    pass
            
            # 在主线程中更新UI
            self.root.after(0, update_ui)
            
            # 更新最后UI更新时间
            self.last_ui_update_time_cache = current_time_ms

    def _write_biosig_to_csv(self, biosig_data):
        """处理生物信号数据，支持批量处理以提高高频数据性能"""
        if not self.running:
            return
        
        try:
            # 减少时间戳计算次数，只在必要时更新系统时间戳
            if self._last_timestamp_update_time is None or \
                (self.__class__.datetime.now() - self._last_timestamp_update_time).total_seconds() > 0.001:  # 最多每秒更新1000次
                self._last_timestamp_update_time = self.__class__.datetime.now()
                system_timestamp = self._last_timestamp_update_time.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
            else:
                # 复用之前的系统时间戳，减少datetime调用开销
                system_timestamp = self._last_timestamp_update_time.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
            
            # 使用独立的生物信号锁，避免与驾驶数据线程冲突
            with self.biosig_lock:
                # 更新当前生物信号参数
                self.current_biosig_params = biosig_data.copy()
                # 添加系统时间戳
                self.current_biosig_params['system_timestamp'] = system_timestamp
                self.last_biosig_system_ts = system_timestamp
                try:
                    self.recent_biosig_ts.append(system_timestamp)
                except Exception:
                    pass
                self.current_biosig_params['sample_count'] = self._biosig_sample_count
                
                # 标记生物信号数据已收到
                self.biosig_received = True
                
                # 尝试激活时间同步
                self._synchronize_timestamps()
                
                # 如果时间同步正在进行，收集样本
                if not self.time_sync_enabled and self.sync_samples_collected < self.sync_samples_required:
                    self._calculate_lsl_system_offset(biosig_data["biosig_timestamp"])
                
                # 显示最新生物信号参数 - 直接调用，内部有节流控制
                self._display_biosig_params()
                
                # 获取同步后的时间戳（如果启用）
                original_timestamp, synchronized_timestamp, lsl_time_str = self._get_synchronized_timestamp(system_timestamp)
                
                # 创建数据行，只包含必要的字段
                row = {
                    "system_timestamp": original_timestamp,
                    "synchronized_timestamp": synchronized_timestamp if self.time_sync_enabled else "N/A",
                    "event_system_timestamp": "N/A",
                    "event_synchronized_timestamp": "N/A",
                    "biosig_timestamp": biosig_data["biosig_timestamp"],
                    "heart_rate_bpm": float(biosig_data["heart_rate_bpm"]) if biosig_data["heart_rate_bpm"] != "N/A" else "N/A",
                    "gsr_uS": float(biosig_data["gsr_uS"]) if biosig_data["gsr_uS"] != "N/A" else "N/A",
                    "marker_flag": False,
                    "marker_type": "N/A",
                    "marker_condition": "N/A",
                    "marker_target": "N/A",
                    "marker_color": "N/A",
                    "marker_label": "N/A"
                }

                # 若刚触发过标记，将其同步到生理行（一次性）
                try:
                    q = getattr(self, '_marker_event_queue', None)
                    if q:
                        try:
                            tsf = None
                            try:
                                tsf = datetime.strptime(original_timestamp, "%Y-%m-%d %H:%M:%S.%f").timestamp()
                            except Exception:
                                tsf = None
                            newq = deque()
                            matched = False
                            now = time.time()
                            win = getattr(self, 'marker_match_window_s', 0.01)
                            while q:
                                ev = q.popleft()
                                et = ev.get('time', 0)
                                if not matched and tsf is not None and abs(tsf - et) <= win:
                                    row['marker_flag'] = True
                                    row['marker_type'] = ev.get('type', 'N/A')
                                    row['marker_condition'] = ev.get('cond', 'N/A')
                                    row['marker_target'] = ev.get('target', 'N/A')
                                    row['marker_color'] = ev.get('color', 'N/A')
                                    row['marker_label'] = ev.get('label', 'N/A')
                                    row['event_system_timestamp'] = ev.get('event_system_timestamp', 'N/A')
                                    row['event_synchronized_timestamp'] = ev.get('event_synchronized_timestamp', 'N/A')
                                    matched = True
                                else:
                                    if now - et < 5.0:
                                        newq.append(ev)
                            self._marker_event_queue = newq
                        except Exception:
                            pass
                except Exception:
                    pass
                
                try:
                    self._biosig_stream_write_queue.put_nowait([row])
                except queue.Full:
                    try:
                        _ = self._biosig_stream_write_queue.get_nowait()
                    except Exception:
                        pass
                    try:
                        self._biosig_stream_write_queue.put_nowait([row])
                    except Exception:
                        pass
                if not getattr(self, 'streaming_write_enabled', False):
                    with self.cache_lock:
                        self.biosig_data_cache.append(row.copy())
            
            # 更新缓存状态显示 - 使用单独的函数避免锁嵌套
            self._update_cache_status()
                
            # 每500条生物信号数据记录一次缓存状态，减少日志频率
            if len(self.biosig_data_cache) % 500 == 0:
                # 计算实际数据收集时间
                if self.actual_start_time:
                    actual_runtime = time.time() - self.actual_start_time
                    sync_status = "已同步" if self.time_sync_enabled else "未同步"
                    self.log(f"[缓存状态] 已缓存 {len(self.biosig_data_cache)} 条生物信号记录，实际记录时间: {actual_runtime:.1f}秒，时间状态: {sync_status}", is_debug=True)
                else:
                    self.log(f"[缓存状态] 已缓存 {len(self.biosig_data_cache)} 条生物信号记录", is_debug=True)
                    
        except Exception as e:
            error_info = traceback.format_exc()
            self.log(f"⚠ 缓存生物信号时发生错误: {str(e)}")
            self.log(f"  错误详情: {error_info.splitlines()[-1]}")
    
    def _write_driving_data_to_csv(self):
        """将驾驶数据缓存保存到独立的CSV文件，支持合并临时批处理文件"""
        # 检查是否有临时批处理文件需要合并
        batch_files = []
        try:
            # 查找所有驾驶数据批处理文件
            for f in os.listdir(self.data_folder):
                if f.startswith('driving_batch_') and f.endswith('.csv'):
                    batch_files.append(os.path.join(self.data_folder, f))
            
            # 按文件名排序（按时间先后顺序）
            batch_files.sort()
            
            if batch_files:
                self.log(f"发现 {len(batch_files)} 个临时批处理文件需要合并")
        except Exception as e:
            self.log(f"⚠ 扫描临时批处理文件时出错: {str(e)}")
            # 继续执行，不影响主数据写入
        
        # 检查是否有数据需要处理（缓存数据或批处理文件或流式文件）
        streaming_drive_path = os.path.join(self.data_folder, "驾驶数据.csv")
        if not self.driving_data_cache and not batch_files and not os.path.exists(streaming_drive_path):
            self.log("没有驾驶数据需要保存")
            return
        
        # 使用自定义文件名或生成默认文件名
        if hasattr(self, 'custom_filename') and self.custom_filename:
            base_filename = self.custom_filename.replace('.csv', '')
            csv_filename = f"{base_filename}_driving.csv"
        else:
            csv_filename = f"driving_{self.session_timestamp}.csv"
        
        # 计算数据收集时间
        if self.actual_start_time:
            data_collection_time = time.time() - self.actual_start_time
            data_hours, data_remainder = divmod(data_collection_time, 3600)
            data_minutes, data_seconds = divmod(data_remainder, 60)
            data_time_str = f"{int(data_hours):02d}:{int(data_minutes):02d}:{int(data_seconds):02d}"
            self.log(f"准备保存驾驶数据CSV文件: 内存中 {len(self.driving_data_cache)} 条记录, 批处理文件 {len(batch_files)} 个, 数据收集时间: {data_time_str}")
        else:
            self.log(f"准备保存驾驶数据CSV文件: 内存中 {len(self.driving_data_cache)} 条记录, 批处理文件 {len(batch_files)} 个")
        
        # 计算驾驶数据的实际采样率（基于内存中的数据）
        driving_sample_rate = self._calculate_sample_rate(self.driving_data_cache, "驾驶模拟日志")
        
        retry_count = 0
        max_retries = 5
        success = False
        
        while retry_count <= max_retries and not success:
            try:
                # 创建完整的文件路径
                csv_file_path = os.path.join(self.data_folder, csv_filename)
                # 创建文件，使用utf-8-sig编码确保中文在Excel中正常显示
                with open(csv_file_path, 'w', newline='', encoding='utf-8-sig') as csv_file:
                    # 构建CSV表头：uv_/fv_版
                    csv_columns = [
                        "system_timestamp","synchronized_timestamp",
                        "marker_flag","marker_type","marker_condition","marker_target","marker_color","marker_label",
                        "uv_ID","uv_Time","uv_Model","uv_description","uv_position X","uv_position Y","uv_position Z",
                        "uv_distanceTravelled","uv_distanceAlongRoad","uv_steering","uv_throttle","uv_brake","uv_lightState",
                        "uv_automaticControl","uv_wheelBase","uv_road","uv_distanceToLeftBorder","uv_distanceToRightBorder",
                        "uv_offsetFromRoadCenter","uv_offsetFromLaneCenter","uv_laneNumber","uv_laneWidth","uv_drivingForwards",
                        "uv_speedLimit","uv_speedOver",
                        "fv_ID","fv_Time","fv_Model","fv_description","fv_position X","fv_position Y","fv_position Z",
                        "fv_distanceTravelled","fv_steering","fv_throttle","fv_brake","fv_lightState","fv_automaticControl",
                        "fv_wheelBase","fv_road","fv_distanceAlongRoad","fv_distanceToLeftBorder","fv_distanceToRightBorder",
                        "fv_offsetFromRoadCenter","fv_offsetFromLaneCenter","fv_laneNumber","fv_laneWidth","fv_drivingForwards",
                        "fv_speedLimit","fv_speedOver"
                    ]
                    
                    # 创建CSV写入器，设置 quoting 参数以确保数值正确保存
                    writer = csv.DictWriter(csv_file, fieldnames=csv_columns, quoting=csv.QUOTE_NONNUMERIC)
                    writer.writeheader()
                    
                    # 1. 先写入所有临时批处理文件中的数据
                    batch_records_count = 0
                    batch_processed_count = 0
                    
                    for batch_file in batch_files:
                        try:
                            self.log(f"正在处理批处理文件: {os.path.basename(batch_file)}")
                            with open(batch_file, 'r', encoding='utf-8-sig') as bfile:
                                reader = csv.DictReader(bfile)
                                # 跳过标题行
                                next(reader, None)
                                
                                # 写入数据行
                                batch_row_count = 0
                                for row in reader:
                                    # 处理数值转换
                                    for k in csv_columns:
                                        if k in row and row[k] != "N/A" and row[k]:
                                            try:
                                                row[k] = float(row[k])
                                            except (ValueError, TypeError):
                                                pass
                                    writer.writerow(row)
                                    batch_row_count += 1
                                    batch_records_count += 1
                                
                                self.log(f"已处理批处理文件 {os.path.basename(batch_file)}，包含 {batch_row_count} 条记录")
                            
                            batch_processed_count += 1
                            
                            # 处理完成后尝试删除临时文件
                            try:
                                os.remove(batch_file)
                                self.log(f"已删除批处理文件: {os.path.basename(batch_file)}")
                            except Exception as e:
                                self.log(f"⚠ 无法删除批处理文件 {os.path.basename(batch_file)}: {str(e)}")
                                # 继续执行，不影响后续操作
                            
                            # 每处理一个批处理文件后刷新缓冲区
                            csv_file.flush()
                            os.fsync(csv_file.fileno())
                            
                        except Exception as e:
                            error_info = traceback.format_exc()
                            self.log(f"⚠ 处理批处理文件 {os.path.basename(batch_file)} 时出错: {str(e)}")
                            self.log(f"  错误详情: {error_info.splitlines()[-1]}")
                            # 继续处理下一个文件
                    
                    if batch_processed_count > 0:
                        self.log(f"✓ 成功合并 {batch_processed_count} 个批处理文件，共 {batch_records_count} 条记录")
                    
                    # 1.5 合并流式写入的驾驶数据文件（仅当内存缓存为空时，避免与缓存重复）
                    stream_row_count = 0
                    if os.path.exists(streaming_drive_path) and len(self.driving_data_cache) == 0:
                        try:
                            self.log(f"正在合并流式驾驶数据文件: {os.path.basename(streaming_drive_path)}")
                            # 流式写入使用的是 utf-8
                            with open(streaming_drive_path, 'r', encoding='utf-8') as sfile:
                                sreader = csv.DictReader(sfile)
                                for row in sreader:
                                    # 写入行，确保列匹配
                                    row_to_write = {}
                                    for k in csv_columns:
                                        row_to_write[k] = row.get(k, "N/A")
                                    writer.writerow(row_to_write)
                                    stream_row_count += 1
                                
                                self.log(f"✓ 成功合并流式数据 {stream_row_count} 条")
                                
                                # 刷新缓冲区
                                if stream_row_count > 0:
                                    csv_file.flush()
                                    os.fsync(csv_file.fileno())
                                    
                        except Exception as e:
                            self.log(f"⚠ 合并流式驾驶数据时出错: {e}")

                    # 2. 写入当前缓存中的数据
                    row_count = 0
                    total_rows = len(self.driving_data_cache)
                    
                    # 在高压模式下（或启用了流式写入），如果已成功合并流式数据，且缓存中只有少量数据（可能是dummy record），则跳过缓存写入
                    # 避免在文件末尾写入重复的dummy数据
                    skip_cache = False
                    # 检查是否是高压模式或流式写入已启用
                    is_high_pressure = getattr(self, '_high_pressure_mode', False)
                    # 如果成功合并了流式数据，且缓存中只有1条记录（极可能是dummy），则跳过
                    if stream_row_count > 0 and total_rows <= 1:
                        # 进一步检查内容，如果是dummy record，它通常是为了过检查而添加的
                        skip_cache = True
                        self.log("检测到流式写入模式下的占位记录，已跳过重复写入")

                    if total_rows > 0 and not skip_cache:
                        self.log(f"开始写入内存中的驾驶数据...")
                        # 初始化进度百分比变量
                        progress_percent = 0.0
                        
                        # 使用较小的批次，确保写入稳定性
                        batch_size = 500
                        for i in range(0, total_rows, batch_size):
                            batch = self.driving_data_cache[i:i+batch_size]
                            for row in batch:
                                # 确保值正确处理，数值类型保持为float，非数值类型转为字符串
                                row_to_write = {}
                                numeric_params = ['system_timestamp','synchronized_timestamp',
                                                  'uv_position X','uv_position Y','uv_position Z','uv_distanceTravelled','uv_distanceAlongRoad','uv_steering','uv_throttle','uv_brake',
                                                  'uv_lightState','uv_automaticControl','uv_wheelBase','uv_distanceToLeftBorder','uv_distanceToRightBorder','uv_offsetFromRoadCenter',
                                                  'uv_offsetFromLaneCenter','uv_laneNumber','uv_laneWidth','uv_drivingForwards','uv_speedLimit','uv_speedOver',
                                                  'fv_position X','fv_position Y','fv_position Z','fv_distanceTravelled','fv_steering','fv_throttle','fv_brake','fv_lightState',
                                                  'fv_automaticControl','fv_wheelBase','fv_distanceToLeftBorder','fv_distanceToRightBorder','fv_offsetFromRoadCenter','fv_offsetFromLaneCenter',
                                                  'fv_laneNumber','fv_laneWidth','fv_drivingForwards','fv_speedLimit','fv_speedOver','fv_distanceAlongRoad']
                                
                                # 直接写入原始值，但确保数值参数在写入前被转换为float
                                row_to_write = {}
                                for k in csv_columns:  # 只处理CSV表头中定义的字段
                                    v = row.get(k)
                                    if v is None or v == "N/A":
                                        row_to_write[k] = "N/A"
                                    else:
                                        # 对于数值参数，强制尝试转换为float
                                        if k in numeric_params:
                                            try:
                                                # 确保数值被转换为float类型
                                                row_to_write[k] = float(v)
                                            except (ValueError, TypeError):
                                                # 如果转换失败，保持为字符串
                                                row_to_write[k] = str(v)
                                        else:
                                            # 非数值参数保持为字符串
                                            row_to_write[k] = str(v)
                                writer.writerow(row_to_write)
                                row_count += 1
                            # 每批次后刷新文件缓冲区并确保同步到磁盘
                        csv_file.flush()
                        os.fsync(csv_file.fileno())
                            
                        # 减少进度报告频率，每1000条记录或20%进度才报告一次
                        progress_percent = (min(row_count, total_rows) / total_rows) * 100
                        if row_count % 1000 == 0 or progress_percent % 20 < 1:
                            self.log(f"[保存进度] 驾驶数据已写入 {min(row_count, total_rows)}/{total_rows} 条记录 ({progress_percent:.1f}%)", is_debug=True)
                        
                        self.log(f"✓ 已写入 {row_count} 条内存中的驾驶数据")
                    
                    streaming_bio_path = os.path.join(self.data_folder, "生物信号数据.csv")
                    if False and os.path.exists(streaming_bio_path):
                        try:
                            self.log("正在合并流式写入的生物信号数据...")
                            with open(streaming_bio_path, 'r', encoding='utf-8') as sfile:
                                sreader = csv.DictReader(sfile)
                                for srow in sreader:
                                    dr_row = {k: "N/A" for k in csv_columns}
                                    for kk in ("system_timestamp","synchronized_timestamp","marker_flag","marker_type","marker_condition","marker_target","marker_color","marker_label"):
                                        dr_row[kk] = srow.get(kk, "N/A")
                                    writer.writerow(dr_row)
                                    row_count += 1
                            self.log("✓ 已合并流式写入文件")
                        except Exception as e:
                            self.log(f"⚠ 合并流式文件失败: {str(e)}")
                    # 流式驾驶数据已在上面 1.5 节合并，此处不再重复合并
                    csv_file.flush()
                    csv_file.flush()
                    try:
                        if os.path.exists(streaming_drive_path):
                            os.remove(streaming_drive_path)
                        self.log("✓ 已清理流式驾驶临时文件")
                    except Exception as e:
                        self.log(f"⚠ 清理临时文件失败: {str(e)}")
                
                total_all_records = batch_records_count + row_count
                self.log(f"✓ 驾驶数据CSV文件已成功保存: {csv_filename}")
                self.log(f"✓ 保存路径: {csv_file_path}")
                self.log(f"✓ 共写入 {total_all_records} 条驾驶数据记录 (批处理文件: {batch_records_count}, 内存: {row_count})")
                if os.path.exists(csv_file_path):
                    file_size = os.path.getsize(csv_file_path)
                    self.log(f"✓ 文件大小: {file_size/1024:.1f} KB")
                # 记录驾驶数据的实际采样率
                if driving_sample_rate is not None:
                    self.log(f"✓ 驾驶模拟日志实际采样率: {driving_sample_rate:.2f} Hz")
                # 生成含颜色高亮的xlsx（若库可用）
                try:
                    base, _ = os.path.splitext(csv_file_path)
                    xlsx_path = base + ".xlsx"
                    pass
                    
                except Exception as e:
                    pass
                success = True
                break
                
            except Exception as e:
                retry_count += 1
                error_info = traceback.format_exc()
                self.log(f"⚠ 保存驾驶数据CSV文件错误 (尝试 {retry_count}/{max_retries}): {str(e)}")
                self.log(f"  错误详情: {error_info.splitlines()[-1]}")
                
                if retry_count <= max_retries:
                    wait_time = retry_count * 1.0
                    self.log(f"将在 {wait_time:.1f} 秒后重试...")
                    time.sleep(wait_time)
        
        if not success:
            self.log("✗ 保存驾驶数据CSV文件失败，请检查文件权限或磁盘空间")
        
        # 清空驾驶数据缓存
        self.driving_data_cache = []
        self._update_cache_status()
    
    def _write_biosig_data_to_csv(self):
        """将生物信号数据保存到独立的CSV文件，始终读取完整流式文件并执行事件对齐"""
        # 使用Trial文件名
        if not hasattr(self, 'current_trial_filenames') or not self.current_trial_filenames.get('biosig'):
            return
            
        csv_filename = self.current_trial_filenames['biosig']
        src = os.path.join(self.data_folder, csv_filename)
        
        # 如果源文件不存在，尝试查找旧的默认文件
        if not os.path.exists(src):
            try:
                if getattr(self, 'biosig_data_cache', None):
                    csv_columns = ["system_timestamp","synchronized_timestamp","biosig_timestamp","heart_rate_bpm","gsr_uS","event_system_timestamp","event_synchronized_timestamp","marker_flag","marker_type","marker_condition","marker_target","marker_color","marker_label"]
                    with open(src, 'w', newline='', encoding='utf-8-sig') as f:
                        w = csv.DictWriter(f, fieldnames=csv_columns, quoting=csv.QUOTE_NONNUMERIC)
                        w.writeheader()
                        with self.cache_lock:
                            cache_copy = list(self.biosig_data_cache)
                        for r in cache_copy:
                            w.writerow({k: r.get(k, "N/A") for k in csv_columns})
            except Exception:
                pass
            old_src = os.path.join(self.data_folder, "生物信号数据.csv")
            if os.path.exists(old_src):
                src = old_src
            else:
                self.log(f"⚠ 无法找到源生物信号文件: {src}")
                return
        
        try:
            # 停止阶段：基于驾驶最终CSV对齐事件并生成生理最终CSV
            candidates = []
            if hasattr(self, 'current_trial_filenames') and self.current_trial_filenames.get('driving'):
                candidates.append(os.path.join(self.data_folder, self.current_trial_filenames['driving']))
            
            # 保留旧的查找逻辑作为备选
            if hasattr(self, 'custom_filename') and self.custom_filename:
                base_filename = self.custom_filename.replace('.csv', '')
                candidates.append(os.path.join(self.data_folder, f"{base_filename}_driving.csv"))
            else:
                candidates.append(os.path.join(self.data_folder, f"driving_{self.session_timestamp}.csv"))
            
            driving_csv = None
            for p in candidates:
                if os.path.exists(p):
                    driving_csv = p
                    break
            
            if not driving_csv:
                self.log("⚠ 未找到驾驶CSV文件，跳过生理事件对齐，保留生理原始文件")
                return
            
            driving_events = []
            
            def _parse_ts(ts):
                try:
                    if isinstance(ts, str):
                        s = ts.strip()
                        if '-' in s:
                            try:
                                return datetime.strptime(s, "%Y-%m-%d %H:%M:%S.%f").timestamp()
                            except Exception:
                                try:
                                    return datetime.strptime(s, "%Y-%m-%d %H:%M:%S").timestamp()
                                except Exception:
                                    return float(s)
                    return float(ts)
                except Exception:
                    return None
            
            if driving_csv and os.path.exists(driving_csv):
                with open(driving_csv, 'r', encoding='utf-8-sig') as dfile:
                    dreader = csv.DictReader(dfile)
                    for drow in dreader:
                        flag = str(drow.get('marker_flag','')).lower()
                        if flag in ('true','1','yes'):
                            ts = _parse_ts(drow.get('system_timestamp'))
                            driving_events.append({
                                'ts': ts,
                                'sys_ts_str': drow.get('system_timestamp','N/A'),
                                'sync_ts': drow.get('synchronized_timestamp','N/A'),
                                'type': drow.get('marker_type','N/A'),
                                'cond': drow.get('marker_condition','N/A'),
                                'target': drow.get('marker_target','N/A'),
                                'color': drow.get('marker_color','N/A'),
                                'label': drow.get('marker_label','N/A'),
                            })
            
            # 定义临时目标文件路径
            dst = os.path.join(self.data_folder, f"aligned_{csv_filename}")
            
            # 事件-行一次性精确匹配（秒级窗口），避免重复/遗漏
            window_s = getattr(self, 'marker_match_window_s', 0.01)
            biosig_ts_pairs = []
            with open(src, 'r', encoding='utf-8-sig') as sfile:
                reader = csv.DictReader(sfile)
                idx = 0
                for srow in reader:
                    biosig_ts_pairs.append((_parse_ts(srow.get('system_timestamp')), idx))
                    idx += 1
            used_idx = set()
            ev_to_idx = {}
            import bisect
            ts_sorted = []
            idx_sorted = []
            for t,i in sorted([p for p in biosig_ts_pairs if p[0] is not None], key=lambda x: x[0]):
                ts_sorted.append(t)
                idx_sorted.append(i)
            first_hits = 0
            second_hits = 0
            unmatched = 0
            for ev in driving_events:
                ts = ev.get('ts')
                if ts is None:
                    continue
                if not ts_sorted:
                    unmatched += 1
                    continue
                pos = bisect.bisect_left(ts_sorted, ts)
                cand = []
                if pos > 0:
                    cand.append(pos-1)
                if pos < len(ts_sorted):
                    cand.append(pos)
                best_j = None
                best_diff = 1e9
                for j in cand:
                    v = ts_sorted[j]
                    d = abs(v - ts)
                    if d <= window_s and d < best_diff and idx_sorted[j] not in used_idx:
                        best_j = j
                        best_diff = d
                if best_j is not None:
                    used_idx.add(idx_sorted[best_j])
                    ev_to_idx[idx_sorted[best_j]] = ev
                    first_hits += 1
                    continue
                best_j = None
                best_diff = 1e9
                for j in cand:
                    v = ts_sorted[j]
                    d = abs(v - ts)
                    if d <= 1.0 and d < best_diff and idx_sorted[j] not in used_idx:
                        best_j = j
                        best_diff = d
                if best_j is not None:
                    used_idx.add(idx_sorted[best_j])
                    ev_to_idx[idx_sorted[best_j]] = ev
                    second_hits += 1
                else:
                    unmatched += 1
            try:
                total_events = len([e for e in driving_events if e.get('ts') is not None])
                matched_events = len(ev_to_idx)
                self.log(f"[事件对齐] 驾驶事件 {total_events} 个，成功匹配到生理行 {matched_events} 个，未匹配 {total_events - matched_events} 个")
            except Exception:
                pass
            
            # 第二遍写出：仅在匹配行复制事件标记，保证不重复
            with open(src, 'r', encoding='utf-8-sig') as sfile, open(dst, 'w', newline='', encoding='utf-8-sig') as out:
                reader = csv.DictReader(sfile)
                csv_columns = ["system_timestamp","synchronized_timestamp","biosig_timestamp","heart_rate_bpm","gsr_uS","event_system_timestamp","event_synchronized_timestamp","marker_flag","marker_type","marker_condition","marker_target","marker_color","marker_label"]
                writer = csv.DictWriter(out, fieldnames=csv_columns, quoting=csv.QUOTE_NONNUMERIC)
                writer.writeheader()
                idx = 0
                for srow in reader:
                    out_row = {k: "N/A" for k in csv_columns}
                    for k in ("system_timestamp","synchronized_timestamp","biosig_timestamp","heart_rate_bpm","gsr_uS"):
                        out_row[k] = srow.get(k, "N/A")
                    ev = ev_to_idx.get(idx)
                    if ev:
                        out_row["marker_flag"] = True
                        out_row["marker_type"] = ev["type"]
                        out_row["marker_condition"] = ev["cond"]
                        out_row["marker_target"] = ev["target"]
                        out_row["marker_color"] = ev["color"]
                        out_row["marker_label"] = ev["label"]
                        out_row["event_system_timestamp"] = ev["sys_ts_str"]
                        out_row["event_synchronized_timestamp"] = ev["sync_ts"]
                    else:
                        out_row["marker_flag"] = False
                    writer.writerow(out_row)
                    idx += 1
            
            # 替换源文件
            try:
                # 关闭所有可能打开的文件句柄
                import gc
                gc.collect()
                time.sleep(0.5)
                
                # 如果是同一个文件（src==dst的源文件），我们需要先删除src
                if os.path.exists(src):
                    os.remove(src)
                os.rename(dst, src)
                self.log(f"✓ 生物信号数据已对齐并保存: {csv_filename}")
            except Exception as e:
                self.log(f"⚠ 替换生物信号文件失败: {str(e)}")
                # 如果替换失败，保留 aligned_... 文件
                pass
                
            return
        except Exception as e:
            self.log(f"⚠ 读取流式文件失败: {str(e)}")
        
        # 计算数据收集时间
        if self.actual_start_time:
            data_collection_time = time.time() - self.actual_start_time
            data_hours, data_remainder = divmod(data_collection_time, 3600)
            data_minutes, data_seconds = divmod(data_remainder, 60)
            data_time_str = f"{int(data_hours):02d}:{int(data_minutes):02d}:{int(data_seconds):02d}"
            self.log(f"准备保存生物信号数据CSV文件: 共 {len(self.biosig_data_cache)} 条记录，数据收集时间: {data_time_str}")
        else:
            self.log(f"准备保存生物信号数据CSV文件: 共 {len(self.biosig_data_cache)} 条记录")
        
        # 计算生物信号数据的实际采样率
        biosig_sample_rate = self._calculate_sample_rate(self.biosig_data_cache, "生理日志")
        
        retry_count = 0
        max_retries = 5
        success = False
        
        while retry_count <= max_retries and not success:
            try:
                # 创建完整的文件路径
                csv_file_path = os.path.join(self.data_folder, csv_filename)
                # 创建文件，使用utf-8-sig编码确保中文在Excel中正常显示
                with open(csv_file_path, 'w', newline='', encoding='utf-8-sig') as csv_file:
                    # 构建CSV表头：包含系统时间戳、同步时间戳、生物信号字段、事件时间戳与标记列
                    csv_columns = [
                        "system_timestamp","synchronized_timestamp","biosig_timestamp","heart_rate_bpm","gsr_uS",
                        "event_system_timestamp","event_synchronized_timestamp",
                        "marker_flag","marker_type","marker_condition","marker_target","marker_color","marker_label"
                    ]
                    candidates = []
                    if hasattr(self, 'custom_filename') and self.custom_filename:
                        base_filename = self.custom_filename.replace('.csv', '')
                        candidates.append(os.path.join(self.data_folder, f"{base_filename}_driving.csv"))
                    else:
                        candidates.append(os.path.join(self.data_folder, f"driving_{self.session_timestamp}.csv"))
                        candidates.append(os.path.join(self.data_folder, f"driving_simulator_driving_{self.session_timestamp}.csv"))
                    driving_csv = None
                    for p in candidates:
                        if os.path.exists(p):
                            driving_csv = p
                            break
                    driving_events = []
                    def _parse_ts(ts):
                        try:
                            if isinstance(ts, str):
                                s = ts.strip()
                                if '-' in s:
                                    try:
                                        return datetime.strptime(s, "%Y-%m-%d %H:%M:%S.%f").timestamp()
                                    except Exception:
                                        try:
                                            return datetime.strptime(s, "%Y-%m-%d %H:%M:%S").timestamp()
                                        except Exception:
                                            return float(s)
                            return float(ts)
                        except Exception:
                            return None
                    if driving_csv and os.path.exists(driving_csv):
                        with open(driving_csv, 'r', encoding='utf-8-sig') as dfile:
                            dreader = csv.DictReader(dfile)
                            for drow in dreader:
                                flag = str(drow.get('marker_flag','')).lower()
                                if flag in ('true','1','yes'):
                                    ts = _parse_ts(drow.get('system_timestamp'))
                                    driving_events.append({
                                        'ts': ts,
                                        'sys_ts_str': drow.get('system_timestamp','N/A'),
                                        'sync_ts': drow.get('synchronized_timestamp','N/A'),
                                        'type': drow.get('marker_type','N/A'),
                                        'cond': drow.get('marker_condition','N/A'),
                                        'target': drow.get('marker_target','N/A'),
                                        'color': drow.get('marker_color','N/A'),
                                        'label': drow.get('marker_label','N/A'),
                                    })
                    if not driving_csv:
                        writer = csv.DictWriter(csv_file, fieldnames=csv_columns, quoting=csv.QUOTE_NONNUMERIC)
                        writer.writeheader()
                        for row in self.biosig_data_cache:
                            writer.writerow({k: row.get(k, "N/A") for k in csv_columns})
                        csv_file.flush()
                        os.fsync(csv_file.fileno())
                        self.log(f"✓ 生物信号数据CSV文件已成功保存: {csv_filename}")
                        self.log(f"✓ 保存路径: {csv_file_path}")
                        self.log(f"✓ 共写入 {len(self.biosig_data_cache)} 条生物信号数据记录")
                        if os.path.exists(csv_file_path):
                            file_size = os.path.getsize(csv_file_path)
                            self.log(f"✓ 文件大小: {file_size/1024:.1f} KB")
                        if biosig_sample_rate is not None:
                            self.log(f"✓ 生理日志实际采样率: {biosig_sample_rate:.2f} Hz")
                        success = True
                        break
                    def _find_match(ts):
                        best=None; best_diff=1e9
                        for ev in driving_events:
                            if ev['ts'] is None: continue
                            diff = abs(ts-ev['ts'])
                            if diff <= 0.010 and diff < best_diff:
                                best=ev; best_diff=diff
                        return best
                    
                    # 事件-行唯一匹配（秒级窗口）用于缓存数据
                    window_s = getattr(self, 'marker_match_window_s', 0.01)
                    biosig_ts_pairs = []
                    for i,r in enumerate(self.biosig_data_cache):
                        biosig_ts_pairs.append((_parse_ts(r.get('system_timestamp')), i))
                    used_idx = set()
                    ev_to_idx = {}
                    import bisect
                    ts_sorted = []
                    idx_sorted = []
                    for t,i in sorted([p for p in biosig_ts_pairs if p[0] is not None], key=lambda x: x[0]):
                        ts_sorted.append(t)
                        idx_sorted.append(i)
                    first_hits = 0
                    second_hits = 0
                    unmatched = 0
                    for ev in driving_events:
                        ts = ev.get('ts')
                        if ts is None:
                            continue
                        if not ts_sorted:
                            unmatched += 1
                            continue
                        pos = bisect.bisect_left(ts_sorted, ts)
                        cand = []
                        if pos > 0:
                            cand.append(pos-1)
                        if pos < len(ts_sorted):
                            cand.append(pos)
                        best_j = None
                        best_diff = 1e9
                        for j in cand:
                            v = ts_sorted[j]
                            d = abs(v - ts)
                            if d <= window_s and d < best_diff and idx_sorted[j] not in used_idx:
                                best_j = j
                                best_diff = d
                        if best_j is not None:
                            used_idx.add(idx_sorted[best_j])
                            ev_to_idx[idx_sorted[best_j]] = ev
                            first_hits += 1
                        else:
                            best_j = None
                            best_diff = 1e9
                            for j in cand:
                                v = ts_sorted[j]
                                d = abs(v - ts)
                                if d <= 1.0 and d < best_diff and idx_sorted[j] not in used_idx:
                                    best_j = j
                                    best_diff = d
                            if best_j is not None:
                                used_idx.add(idx_sorted[best_j])
                                ev_to_idx[idx_sorted[best_j]] = ev
                                second_hits += 1
                            else:
                                unmatched += 1
                    
                    # 写出缓存数据，按匹配结果标记
                    total_rows = len(self.biosig_data_cache)
                    row_count = 0
                    writer = csv.DictWriter(csv_file, fieldnames=csv_columns, quoting=csv.QUOTE_NONNUMERIC)
                    writer.writeheader()
                    for i in range(total_rows):
                        row = self.biosig_data_cache[i]
                        row_to_write = {k: "N/A" for k in csv_columns}
                        for k in ("system_timestamp","synchronized_timestamp","biosig_timestamp","heart_rate_bpm","gsr_uS"):
                            row_to_write[k] = row.get(k, "N/A")
                        ev = ev_to_idx.get(i)
                        if ev:
                            row_to_write["marker_flag"] = True
                            row_to_write["marker_type"] = ev["type"]
                            row_to_write["marker_condition"] = ev["cond"]
                            row_to_write["marker_target"] = ev["target"]
                            row_to_write["marker_color"] = ev["color"]
                            row_to_write["marker_label"] = ev["label"]
                            row_to_write["event_system_timestamp"] = ev["sys_ts_str"]
                            row_to_write["event_synchronized_timestamp"] = ev["sync_ts"]
                        else:
                            row_to_write["marker_flag"] = False
                        writer.writerow(row_to_write)
                        row_count += 1
                    try:
                        total_events = len([e for e in driving_events if e.get('ts') is not None])
                        self.log(f"[事件对齐] 驾驶事件 {total_events} 个，10ms 命中 {first_hits} 个，1s 二次命中 {second_hits} 个，未命中 {unmatched} 个")
                    except Exception:
                        pass
                    
                    try:
                        total_events = len([e for e in driving_events if e.get('ts') is not None])
                        matched_events = len(ev_to_idx)
                        self.log(f"[事件对齐] 驾驶事件 {total_events} 个，成功匹配到生理行 {matched_events} 个，未匹配 {total_events - matched_events} 个")
                    except Exception:
                        pass
                    
                    csv_file.flush()
                    os.fsync(csv_file.fileno())
                
                self.log(f"✓ 生物信号数据CSV文件已成功保存: {csv_filename}")
                self.log(f"✓ 保存路径: {csv_file_path}")
                self.log(f"✓ 共写入 {len(self.biosig_data_cache)} 条生物信号数据记录")
                if os.path.exists(csv_file_path):
                    file_size = os.path.getsize(csv_file_path)
                    self.log(f"✓ 文件大小: {file_size/1024:.1f} KB")
                # 记录生物信号数据的实际采样率
                if biosig_sample_rate is not None:
                    self.log(f"✓ 生理日志实际采样率: {biosig_sample_rate:.2f} Hz")
                # 生成含颜色高亮的xlsx（若库可用）
                pass
                success = True
                break
                
            except Exception as e:
                retry_count += 1
                error_info = traceback.format_exc()
                self.log(f"⚠ 保存生物信号数据CSV文件错误 (尝试 {retry_count}/{max_retries}): {str(e)}")
                self.log(f"  错误详情: {error_info.splitlines()[-1]}")
                
                if retry_count <= max_retries:
                    wait_time = retry_count * 1.0
                    self.log(f"将在 {wait_time:.1f} 秒后重试...")
                    time.sleep(wait_time)
        
        if not success:
            self.log("✗ 保存生物信号数据CSV文件失败，请检查文件权限或磁盘空间")
        
        # 清空生物信号数据缓存
        self.biosig_data_cache = []
        self._update_cache_status()

    def _calculate_sample_rate(self, data_cache, data_type):
        """计算数据的实际采样率
        
        Args:
            data_cache: 数据缓存列表
            data_type: 数据类型名称（用于日志）
            
        Returns:
            float: 采样率（Hz），如果无法计算则返回None
        """
        if not data_cache or len(data_cache) < 2:
            self.log(f"⚠ 数据量不足，无法计算{data_type}的采样率")
            return None
        
        try:
            # 尝试使用synchronized_timestamp（如果存在且有效）
            # 否则使用system_timestamp
            valid_timestamps = []
            
            for record in data_cache:
                # 优先使用synchronized_timestamp
                if 'synchronized_timestamp' in record and record['synchronized_timestamp'] not in (None, "N/A"):
                    try:
                        ts = float(record['synchronized_timestamp'])
                        if ts > 0:  # 确保时间戳有效
                            valid_timestamps.append(ts)
                    except (ValueError, TypeError):
                        continue
                # 其次使用system_timestamp
                elif 'system_timestamp' in record and record['system_timestamp'] not in (None, "N/A"):
                    try:
                        # 检查是否为datetime字符串格式
                        if isinstance(record['system_timestamp'], str) and len(record['system_timestamp']) > 10:
                            # 解析datetime字符串为时间戳
                            ts = datetime.strptime(record['system_timestamp'], "%Y-%m-%d %H:%M:%S.%f").timestamp()
                        else:
                            ts = float(record['system_timestamp'])
                        if ts > 0:  # 确保时间戳有效
                            valid_timestamps.append(ts)
                    except (ValueError, TypeError):
                        continue
            
            if len(valid_timestamps) < 2:
                self.log(f"⚠ 有效时间戳不足，无法计算{data_type}的采样率")
                return None
            
            # 按时间戳排序
            valid_timestamps.sort()
            
            # 计算时间范围（秒）
            time_range = valid_timestamps[-1] - valid_timestamps[0]
            
            if time_range <= 0:
                self.log(f"⚠ 时间范围无效，无法计算{data_type}的采样率")
                return None
            
            # 计算采样率（记录数/时间范围）
            sample_rate = len(valid_timestamps) / time_range
            
            # 记录详细信息用于调试
            self.log(f"[采样率计算] {data_type}: 记录数={len(valid_timestamps)}, "
                   f"时间范围={time_range:.3f}秒, 采样率={sample_rate:.2f}Hz", is_debug=True)
            
            return sample_rate
            
        except Exception as e:
            self.log(f"⚠ 计算{data_type}采样率时出错: {str(e)}")
            return None

    def _generate_xlsx_with_markers(self, csv_path: str, xlsx_path: str):
        try:
            import openpyxl
            from openpyxl.styles import PatternFill
            # 读取CSV
            with open(csv_path, 'r', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                headers = reader.fieldnames or []
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.append(headers)
                for row in reader:
                    values = [row.get(h, "") for h in headers]
                    ws.append(values)
                    color = (row.get('marker_color') or '').lower()
                    if color in ('yellow','blue'):
                        fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid') if color=='yellow' else PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')
                        for cell in ws[ws.max_row]:
                            cell.fill = fill
                wb.save(xlsx_path)
                return
        except Exception:
            pass
        # 尝试使用xlsxwriter
        try:
            import xlsxwriter
            with open(csv_path, 'r', encoding='utf-8-sig') as f:
                reader = csv.reader(f)
                rows = list(reader)
            wb = xlsxwriter.Workbook(xlsx_path)
            ws = wb.add_worksheet()
            fmt_yellow = wb.add_format({'bg_color': '#FFFF00'})
            fmt_blue = wb.add_format({'bg_color': '#0000FF', 'font_color': '#FFFFFF'})
            fmt_normal = wb.add_format()
            # 找到 marker_color 列索引
            marker_idx = -1
            if rows:
                headers = rows[0]
                for i, h in enumerate(headers):
                    if h == 'marker_color':
                        marker_idx = i
                        break
            for r, row in enumerate(rows):
                use_fmt = fmt_normal
                if r > 0 and marker_idx >= 0 and len(row) > marker_idx:
                    color = (row[marker_idx] or '').lower()
                    if color == 'yellow':
                        use_fmt = fmt_yellow
                    elif color == 'blue':
                        use_fmt = fmt_blue
                for c, val in enumerate(row):
                    ws.write(r, c, val, use_fmt)
            wb.close()
        except Exception:
            # 若无可用库则静默失败，由调⽤处记录日志
            raise
            
    def log(self, message, is_debug=False):
        """记录日志消息。不入队 after(0)，改为放入 _log_queue，由主线程 _drain_log_queue 统一排出，避免点击卡死。"""
        timestamp = self.__class__.datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
        log_message = f"[{timestamp}] {message}"
        
        try:
            if getattr(self, '_log_queue', None) is not None:
                try:
                    self._log_queue.put_nowait(log_message)
                except queue.Full:
                    pass
            # 同时打印到控制台
            print(log_message)
            
            # 记录到文件日志 - 保存到驾驶指标保存文件夹
            try:
                if not hasattr(self, "_last_log_file_time_ms"):
                    self._last_log_file_time_ms = 0
                if not hasattr(self, "log_file_interval_ms"):
                    self.log_file_interval_ms = 500
                now_ms_file = int(time.time() * 1000)
                if now_ms_file - self._last_log_file_time_ms >= self.log_file_interval_ms:
                    self._last_log_file_time_ms = now_ms_file
                    log_file_path = os.path.join(self.data_folder, 'driving_simulator_sync.log')
                    with open(log_file_path, 'a', encoding='utf-8-sig') as log_file:
                        log_file.write(log_message + "\n")
            except:
                # 如果写入文件失败，继续运行不影响主程序
                pass
                
        except Exception as e:
            # 确保日志功能的错误不会影响主程序
            try:
                print(f"日志记录错误: {str(e)}")
            except:
                pass
                
    def _display_biosig_params(self):
        """90分钟实验专用：生物信号显示优化"""
        # 90分钟实验：大幅降低生物信号UI更新频率
        now_ms = int(time.time() * 1000)
        # 生物信号更新间隔延长到5秒（1000Hz数据用）
        biosig_ui_interval = 5000
        if now_ms - getattr(self, 'last_ui_update_time_biosig', 0) < biosig_ui_interval:
            return
        self.last_ui_update_time_biosig = now_ms

        def update_display():
            # 使用生物信号锁读取参数
            with self.biosig_lock:
                params_copy = self.current_biosig_params.copy()
                sample_count = self._biosig_sample_count
                connected = self.biosig_connected
            
            self.biosig_text.delete(1.0, tk.END)
            
            # 显示连接状态
            if connected:
                self.biosig_text.insert(tk.END, "✓ 生物信号传输: 已连接\n", "connected")
            else:
                self.biosig_text.insert(tk.END, "✗ 生物信号传输: 未连接\n", "disconnected")
            
            # 显示样本计数
            self.biosig_text.insert(tk.END, f"总样本数: {sample_count}\n")
            
            # 显示系统时间戳，使用完整的年月日时分秒毫秒格式
            current_time_str = self.__class__.datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
            self.biosig_text.insert(tk.END, f"系统时间戳: {current_time_str}\n")
            
            # 显示生物信号参数
            if params_copy:
                # 显示生物信号时间戳
                if 'biosig_timestamp' in params_copy:
                    # 尝试将LSL时间戳转换为年月日时分秒格式
                    biosig_ts = params_copy['biosig_timestamp']
                    try:
                        # 如果启用了时间同步，尝试转换LSL时间戳
                        if hasattr(self, 'time_sync_enabled') and self.time_sync_enabled and isinstance(biosig_ts, (int, float)):
                            # 使用LSL时间戳加上偏移量得到系统时间
                            from datetime import datetime, timedelta
                            # pylsl的时间戳是从1970年1月1日开始的秒数
                            ts_dt = datetime.fromtimestamp(float(biosig_ts) + getattr(self, 'time_offset', 0))
                            formatted_ts = ts_dt.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
                            self.biosig_text.insert(tk.END, f"生物信号时间戳: {formatted_ts}\n", "blue")
                        else:
                            self.biosig_text.insert(tk.END, f"生物信号时间戳: {biosig_ts}\n")
                    except:
                        self.biosig_text.insert(tk.END, f"生物信号时间戳: {biosig_ts}\n")
                    
                # 显示心率(原始数据)，确保实时显示具体数值
                if 'heart_rate_bpm' in params_copy:
                    hr_value = params_copy['heart_rate_bpm']
                    try:
                        hr_float = float(hr_value)
                        # 实时显示精确的心率原始值，使用3位小数确保精度
                        # 按要求：不识别数据大小做是否合规，只需要如实的实时显示
                        # 对所有心率值使用统一的显示颜色
                        hr_color = "darkgreen"  # 统一使用绿色显示心率原始值
                        # 使用更突出的格式显示心率
                        hr_text = f"💓 心率原始值: {hr_float:.3f}\n"
                        self.biosig_text.insert(tk.END, hr_text, hr_color)
                        # 更新单独的心率显示标签，实时显示具体数值
                        if hasattr(self, 'heart_rate_display_var') and hasattr(self, 'heart_rate_display'):
                            self.heart_rate_display_var.set(f"{hr_float:.1f}")
                            # 对所有心率值使用统一的显示颜色
                            self.heart_rate_display.configure(foreground='#27AE60')  # 统一使用绿色显示
                    except (ValueError, TypeError):
                        self.biosig_text.insert(tk.END, f"💓 心率原始值: {hr_value}\n", "warning")
                        if hasattr(self, 'heart_rate_display_var'):
                            self.heart_rate_display_var.set(f"{hr_value}")
                
                # 显示GSR(皮肤电导率)，确保实时显示具体数值
                if 'gsr_uS' in params_copy:
                    gsr_value = params_copy['gsr_uS']
                    try:
                        gsr_float = float(gsr_value)
                        # 实时显示精确的皮肤电导率数值，使用6位小数确保高精度
                        # 使用紫色和加粗字体突出显示皮电数据
                        gsr_text = f"⚡ 皮肤电导率: {gsr_float:.6f} μS\n"
                        self.biosig_text.insert(tk.END, gsr_text, "purple")
                        # 更新单独的皮电显示标签，实时显示具体数值
                        if hasattr(self, 'gsr_display_var') and hasattr(self, 'gsr_display'):
                            self.gsr_display_var.set(f"{gsr_float:.3f} μS")
                    except (ValueError, TypeError):
                        self.biosig_text.insert(tk.END, f"⚡ 皮肤电导率: {gsr_value} μS (无法解析)\n", "warning")
                        if hasattr(self, 'gsr_display_var'):
                            self.gsr_display_var.set("-- μS")
                
                # 显示其他可能存在的生理信号参数
                for key, value in params_copy.items():
                    # 跳过已经显示过的参数
                    if key not in ['system_timestamp', 'biosig_timestamp', 'heart_rate_bpm', 'gsr_uS']:
                        try:
                            # 尝试将值转换为浮点数以显示具体数值
                            float_value = float(value)
                            self.biosig_text.insert(tk.END, f"{key}: {float_value:.6f}\n")
                        except (ValueError, TypeError):
                            self.biosig_text.insert(tk.END, f"{key}: {value}\n")
            
            # 确保滚动条在底部
            self.biosig_text.yview_moveto(1.0)
        
        # 在主线程中更新UI
        self.root.after(0, update_display)
        
        # 配置文本标签颜色（只需要在初始化时配置一次，但放在这里确保始终有效）
        self.biosig_text.tag_config("connected", foreground="green", font=("Arial", 9, "bold"))
        self.biosig_text.tag_config("disconnected", foreground="red", font=("Arial", 9, "bold"))
        self.biosig_text.tag_config("blue", foreground="blue")
        self.biosig_text.tag_config("red", foreground="red")
        self.biosig_text.tag_config("purple", foreground="purple", font=("Arial", 9, "bold"))
        self.biosig_text.tag_config("warning", foreground="orange", font=("Arial", 8, "italic"))
    
    def _receive_lsl_data(self):
        """增强的LSL生物信号数据接收方法 - 添加详细调试和更强的错误恢复"""
        # 初始化状态
        self.biosig_inlet = None
        with self.biosig_lock:
            self._biosig_sample_count = 0
            self.biosig_connected = False
            self.biosig_last_timestamp = None
        
        # 初始化显示
        self._display_biosig_params()
        
        # 清空队列（90分钟 1000Hz：约5秒缓冲，减少主线程短时卡顿导致的丢包）
        self.biosig_batch_queue = queue.Queue(maxsize=5000)
        
        # 添加调试计数器
        retry_count = 0
        max_retries = 5
        data_counter = 0
        last_debug_print = time.time()
        
        # 创建LSL客户端实例
        self.lsl_client = None
        if PYLSL_AVAILABLE:
            try:
                self.log("🔍 正在初始化OpenSignals LSL客户端...")
                self.lsl_client = OpenSignalsLSLClient(mac_address="")
                self.log("✓ 初始化OpenSignals LSL客户端成功")
            except Exception as e:
                self.log(f"⚠ 初始化OpenSignals LSL客户端失败: {str(e)}")
                import traceback
                traceback.print_exc()
        else:
            self.log("❌ PYLSL库不可用，无法接收LSL数据")
        
        while self.running and PYLSL_AVAILABLE:
            try:
                # 尝试连接LSL流，使用指数退避重试
                if not self.lsl_client or not self.lsl_client.find_and_connect():
                    retry_count += 1
                    wait_time = min(retry_count * 0.5, 3.0)  # 指数退避，最大3秒
                    self.log(f"⚠ 无法连接到LSL流，将在{wait_time:.1f}秒后重试... (尝试 {retry_count}/{max_retries})")
                    time.sleep(wait_time)
                    if retry_count >= max_retries:
                        self.log("🔄 达到最大重试次数，重置连接状态")
                        retry_count = 0
                        # 重新创建LSL客户端实例
                        if PYLSL_AVAILABLE:
                            try:
                                self.lsl_client = OpenSignalsLSLClient(mac_address="")
                                self.log("✓ 已重新初始化LSL客户端")
                            except Exception as e:
                                self.log(f"⚠ 重新初始化LSL客户端失败: {str(e)}")
                    continue
                
                # 连接成功后重置重试计数器
                retry_count = 0
                
                with self.biosig_lock:
                    self.biosig_connected = True
                self.log("✅ 已成功连接到LSL流")
                
                # 显示连接信息
                if hasattr(self.lsl_client, 'found_stream_info') and self.lsl_client.found_stream_info:
                    info = self.lsl_client.found_stream_info
                    self.log(f"📊 已连接流信息: 名称='{info.name()}', 类型='{info.type()}', 通道数={info.channel_count()}")
                    self.log(f"📊 固定通道映射: 心率=索引1, 皮电=索引2")
                
                self._display_biosig_params()
                
                # 持续接收数据
                while self.running:
                    try:
                        # 增加超时时间，从0.1秒增加到0.5秒
                        # self.log("⏳ 尝试获取LSL样本数据...")
                        # 只传递timeout参数，不传递blocking参数
                        data_result = self._lsl_receive_batch(timeout=self.lsl_receive_timeout, max_samples=self.lsl_max_samples_per_pull)
                        
                        # 处理批量样本列表
                        if isinstance(data_result, list):
                            batch_size = len(data_result)
                            if batch_size > 0:
                                # 批量处理所有样本
                                for sample, timestamp in data_result:
                                    data_counter += 1
                                    
                                    # 减少调试信息，避免影响1000Hz高频采集
                                    if data_counter % 1000 == 0 or (time.time() - last_debug_print) >= 10:
                                        last_debug_print = time.time()
                                        self.log(f"📈 接收到样本 #{data_counter}: 通道数={len(sample)}, 批量大小={batch_size}")
                                    
                                    # 增加样本计数
                                    with self.biosig_lock:
                                        self._biosig_sample_count += 1
                                        current_count = self._biosig_sample_count
                                    
                                    # 更新最后时间戳
                                    self.biosig_last_timestamp = timestamp
                                    
                                    # 显示第一个样本的详细信息
                                    if current_count == 1:
                                        self.log(f"🎉 接收到第一个样本，包含 {len(sample)} 个通道")
                                        self.log(f"   通道数据预览: {sample}")
                                    
                                    # 直接使用V9版本的固定通道映射
                                    heart_rate_bpm = "N/A"
                                    gsr_uS = "N/A"
                                    
                                    try:
                                        # 通道2(索引1)为心率，通道3(索引2)为皮电
                                        if len(sample) > 1:
                                            heart_rate_bpm = float(sample[1])
                                        if len(sample) > 2:
                                            gsr_uS = float(sample[2])
                                    except (ValueError, TypeError):
                                        # 完全静默处理错误，避免影响高频采集
                                        pass
                                    
                                    # 构建数据字典
                                    biosig_data = {
                                        "biosig_timestamp": timestamp,  # 直接使用原始LSL时间戳
                                        "heart_rate_bpm": heart_rate_bpm,
                                        "gsr_uS": gsr_uS
                                    }
                                    
                                    # 添加到处理队列
                                    try:
                                        self.biosig_batch_queue.put_nowait(biosig_data)
                                        # 进一步减少队列状态日志
                                        if data_counter % 1000 == 0:
                                            self.log(f"✅ 数据已添加到队列，当前队列大小: {self.biosig_batch_queue.qsize()}")
                                    except queue.Full:
                                        # 静默丢弃，避免日志开销
                                        pass
                                
                                # 由主线程/调度器处理批次，避免阻塞LSL接收线程
                                pass
                        # 向后兼容：处理单样本返回格式
                        elif data_result and isinstance(data_result, tuple) and len(data_result) == 2:
                            sample, timestamp = data_result
                            if sample and timestamp:
                                data_counter += 1
                                # 处理单个样本的逻辑（与上面类似但简化）
                                with self.biosig_lock:
                                    self._biosig_sample_count += 1
                                self.biosig_last_timestamp = timestamp
                                
                                # 通道映射
                                heart_rate_bpm = "N/A"
                                gsr_uS = "N/A"
                                try:
                                    if len(sample) > 1:
                                        heart_rate_bpm = float(sample[1])
                                    if len(sample) > 2:
                                        gsr_uS = float(sample[2])
                                except:
                                    pass
                                
                                # 添加到队列
                                try:
                                    self.biosig_batch_queue.put_nowait({
                                        "biosig_timestamp": timestamp,
                                        "heart_rate_bpm": heart_rate_bpm,
                                        "gsr_uS": gsr_uS
                                    })
                                except queue.Full:
                                    pass
                                
                                if data_counter % 100 == 0:
                                    self._process_biosig_batch()
                        else:
                            # 最小化休眠时间，确保高频采集
                            time.sleep(0.001)
                                
                    except Exception as e:
                        # 增强错误处理
                        self.log(f"⚠ LSL数据接收错误: {str(e)}")
                        import traceback
                        traceback.print_exc()
                        time.sleep(0.1)
                        break  # 跳出内部循环，重新尝试连接
            except Exception as e:
                self.log(f"✗ LSL流处理错误: {str(e)}")
                import traceback
                traceback.print_exc()
                time.sleep(0.5)
        
        # 处理剩余数据
        if not self.biosig_batch_queue.empty():
            self._process_biosig_batch()
        
        # 关闭连接
        with self.biosig_lock:
            self.biosig_connected = False
        
        self._display_biosig_params()
        
        if hasattr(self, 'lsl_client') and self.lsl_client:
            try:
                self.lsl_client.close()
                self.log("✓ LSL流已关闭")
            except Exception as e:
                self.log(f"⚠ 关闭LSL流时发生错误: {str(e)}")
                
    def _process_biosig_batch(self):
        """高效批量处理生物信号数据，支持1000Hz高频采样"""
        # 处理队列中的数据
        try:
            # 优先处理队列中的数据
            batch_data = []
            
            # 从队列中获取所有可用数据
            while not self.biosig_batch_queue.empty():
                try:
                    data = self.biosig_batch_queue.get_nowait()
                    batch_data.append(data)
                    self.biosig_batch_queue.task_done()
                except queue.Empty:
                    break
            
            # 如果队列为空，尝试从缓冲区获取数据
            if not batch_data:
                with self.cache_lock:
                    if len(self.biosig_batch_buffer) > 0:
                        batch_data = self.biosig_batch_buffer.copy()
                        self.biosig_batch_buffer.clear()
                    else:
                        return
            
            # 添加日志频率控制，避免影响1000Hz高频采样
            if len(batch_data) > 0:
                if not hasattr(self, '_process_batch_log_counter'):
                    self._process_batch_log_counter = 0
                self._process_batch_log_counter += 1
                # 每50次处理记录一次日志
                if self._process_batch_log_counter % 50 == 0:
                    self.log(f"📋 正在处理 {len(batch_data)} 个生物信号样本")
            
            if not batch_data:
                return
            
            # 标记生物信号数据已收到
            if not self.biosig_received:
                self.biosig_received = True
                # 尝试激活时间同步
                self._synchronize_timestamps()
            
            # 批量处理数据并准备写入
            processed_data = []
            
            for biosig_data in batch_data:
                try:
                    # 基于LSL时间戳计算系统时间，确保与驾驶事件在同一时间域对齐
                    if self.time_sync_enabled and biosig_data.get("biosig_timestamp") not in (None, "N/A"):
                        try:
                            system_ts_sec = float(biosig_data["biosig_timestamp"]) - float(self.time_offset)
                            system_timestamp_str = self.__class__.datetime.fromtimestamp(system_ts_sec).strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
                        except Exception:
                            system_timestamp_str = self.__class__.datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
                    else:
                        system_timestamp_str = self.__class__.datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
                    # 获取同步后的显示与LSL格式
                    original_timestamp, synchronized_timestamp, lsl_time_str = self._get_synchronized_timestamp(system_timestamp_str)
                    
                    # 创建数据行
                    # 初始化时间差为N/A
                    time_diff_lsl = "N/A"
                    
                    # 当时间同步启用且有有效的LSL时间数据时，计算两个输出流的时间差
                    if self.time_sync_enabled and lsl_time_str != "N/A" and biosig_data["biosig_timestamp"] != "N/A":
                        try:
                            # 将系统时间的LSL格式和生物信号时间戳转换为浮点数
                            system_lsl_time = float(lsl_time_str)
                            biosig_lsl_time = float(biosig_data["biosig_timestamp"])
                            
                            # 计算时间差：系统LSL时间 - 生物信号LSL时间
                            # 转换为毫秒级精度，使显示更直观
                            time_diff_seconds = system_lsl_time - biosig_lsl_time
                            time_diff_lsl = time_diff_seconds * 1000  # 转换为毫秒
                            
                            # 大幅减少时间差日志输出，仅在时间差较大或周期性输出
                            if hasattr(self, 'log'):
                                # 添加静态计数器用于控制日志频率
                                if not hasattr(self, '_time_diff_log_counter'):
                                    self._time_diff_log_counter = 0
                                self._time_diff_log_counter += 1
                                
                                # 暂时完全禁用时间差计算日志，避免影响1000Hz高频采样性能
                                # if abs(time_diff_lsl) > 100 or self._time_diff_log_counter % 100 == 0:
                                #     self.log(f"ℹ 时间差计算: 系统LSL时间={system_lsl_time:.6f}, 生物信号LSL时间={biosig_lsl_time:.6f}, 差值={time_diff_lsl:.3f} ms")
                        except Exception as e:
                            if hasattr(self, 'log'):
                                self.log(f"⚠ 计算时间差失败: {str(e)}")
                    
                    row = {
                        "system_timestamp": original_timestamp,
                        "synchronized_timestamp": synchronized_timestamp if self.time_sync_enabled else "N/A",
                        "event_system_timestamp": "N/A",
                        "event_synchronized_timestamp": "N/A",
                        "system_lsl_time": lsl_time_str if self.time_sync_enabled else "N/A",
                        "biosig_timestamp": biosig_data["biosig_timestamp"],
                        "time_diff_lsl": time_diff_lsl,
                        "heart_rate_bpm": float(biosig_data["heart_rate_bpm"]) if biosig_data["heart_rate_bpm"] != "N/A" else "N/A",
                        "gsr_uS": float(biosig_data["gsr_uS"]) if biosig_data["gsr_uS"] != "N/A" else "N/A",
                        "marker_flag": False,
                        "marker_type": "N/A",
                        "marker_condition": "N/A",
                        "marker_target": "N/A",
                        "marker_color": "N/A",
                        "marker_label": "N/A"
                    }
                    try:
                        eye_latest = getattr(self, 'eye_latest', {}) or {}
                        gp = eye_latest.get('gp')
                        row["Gaze_TS"] = eye_latest.get('ts', '')
                        if isinstance(gp, (list, tuple)) and len(gp) >= 2:
                            row["Gaze_X"] = gp[0]
                            row["Gaze_Y"] = gp[1]
                        elif self._last_gaze_xy and (time.time() - self._last_gaze_time) <= 1.0:
                            row["Gaze_X"] = self._last_gaze_xy[0]
                            row["Gaze_Y"] = self._last_gaze_xy[1]
                        else:
                            row["Gaze_X"] = ''
                            row["Gaze_Y"] = ''
                    except Exception:
                        row["Gaze_TS"] = ''
                        row["Gaze_X"] = ''
                        row["Gaze_Y"] = ''
                    try:
                        matched = False
                        if hasattr(self, '_marker_event_queue') and self._marker_event_queue:
                            # 解析当前行的系统时间为epoch秒
                            row_ts_epoch = None
                            try:
                                row_ts_epoch = datetime.strptime(original_timestamp, "%Y-%m-%d %H:%M:%S.%f").timestamp()
                            except Exception:
                                row_ts_epoch = None
                            win = getattr(self, 'marker_match_window_s', 0.01)
                            if row_ts_epoch is not None:
                                # 在窗口内找到最早的匹配事件并打标，确保一对一
                                for idx_ev, ev in enumerate(list(self._marker_event_queue)):
                                    et = ev.get('time', 0)
                                    if (time.time() - et) > 5.0:
                                        try:
                                            self._marker_event_queue.remove(ev)
                                        except Exception:
                                            pass
                                        continue
                                    if abs(row_ts_epoch - et) <= win:
                                        row['marker_flag'] = True
                                        row['marker_type'] = ev.get('type', 'N/A')
                                        row['marker_condition'] = ev.get('cond', 'N/A')
                                        row['marker_target'] = ev.get('target', 'N/A')
                                        row['marker_color'] = ev.get('color', 'N/A')
                                        row['marker_label'] = ev.get('label', 'N/A')
                                        row['event_system_timestamp'] = ev.get('event_system_timestamp', 'N/A')
                                        row['event_synchronized_timestamp'] = ev.get('event_synchronized_timestamp', 'N/A')
                                        try:
                                            self._marker_event_queue.remove(ev)
                                        except Exception:
                                            pass
                                        matched = True
                                        break
                    except Exception:
                        pass
                    processed_data.append(row)
                    
                    # 使用缓存锁来添加到缓存列表（高压模式下不再累积到内存）
                    if not self._high_pressure_mode:
                        with self.cache_lock:
                            self.biosig_data_cache.append(row)
                    
                    # 更新当前生物信号参数，用于GUI实时显示
                    with self.biosig_lock:
                        self.current_biosig_params = {
                            "biosig_timestamp": biosig_data["biosig_timestamp"],
                            "heart_rate_bpm": biosig_data["heart_rate_bpm"],
                            "gsr_uS": biosig_data["gsr_uS"]
                        }
                        
                except Exception as e:
                    # 静默处理单个样本错误，避免影响批量处理
                    pass
            
            # 批量写入CSV文件，减少I/O操作
            if processed_data:
                biosig_file_path = os.path.join(self.data_folder, "生物信号数据.csv")
                file_exists = os.path.exists(biosig_file_path)
                
                try:
                    self._biosig_stream_write_queue.put_nowait(processed_data)
                except queue.Full:
                    try:
                        _ = self._biosig_stream_write_queue.get_nowait()
                    except Exception:
                        pass
                    try:
                        self._biosig_stream_write_queue.put_nowait(processed_data)
                    except Exception:
                        pass
                
                # 同步写入到磁盘，确保数据安全（在流式写入时降低频率）
                # 这里不强制fsync，避免频繁磁盘阻塞
            
            # 一次性更新缓存状态
            self._update_cache_status()
            
            # 更新批量处理时间
            self._last_biosig_batch_time = time.time()
            
            # 立即更新UI显示，确保生物信号数据实时可见
            self._display_biosig_params()
            
        except queue.Empty:
            # 队列为空，正常情况
            pass
        except Exception as e:
            self.log(f"批量处理生物信号数据时发生错误: {str(e)}", "错误")
        
    def _check_sampling_rate(self):
        """检查采样率状态，确保达到目标采样率"""
        current_time = time.time()
        
        # 定期检查，避免频繁计算
        if current_time - self.biosig_last_check_time < self.biosig_check_interval:
            return
        
        # 计算实际采样率
        with self.biosig_lock:
            total_samples = self._biosig_sample_count
            missed_samples = self.biosig_missed_samples
        
        if self.actual_start_time and total_samples > 0:
            elapsed_time = current_time - self.actual_start_time
            actual_rate = total_samples / elapsed_time
            
            # 更新检查时间
            self.biosig_last_check_time = current_time
            
            # 如果采样率低于目标的90%或检测到大量掉包，记录警告
            if actual_rate < self.biosig_sampling_rate_target * 0.9 or missed_samples > total_samples * 0.01:
                self.log(f"⚠ 采样率警告: 实际采样率 {actual_rate:.1f} Hz (目标: {self.biosig_sampling_rate_target} Hz)，丢失样本数: {missed_samples} ({missed_samples/total_samples*100:.2f}%)")
            elif total_samples % 5000 == 0:  # 每5000个样本记录一次正常状态
                self.log(f"✓ 采样率正常: {actual_rate:.1f} Hz，丢失样本率: {missed_samples/total_samples*100:.2f}%")

    def _lsl_receive_batch(self, timeout, max_samples):
        try:
            cli = getattr(self, 'lsl_client', None)
            if cli is None:
                return None
            if hasattr(cli, 'receive_chunk'):
                samples, timestamps = cli.receive_chunk(max_samples=max_samples, timeout=timeout)
                if samples and timestamps and len(samples) > 0:
                    return [(samples[i], timestamps[i]) for i in range(len(samples))]
                return None
            else:
                result = cli.receive_data(timeout=timeout, max_samples=max_samples)
                if isinstance(result, list) and len(result) > 0:
                    return result
                if isinstance(result, tuple) and len(result) == 2 and result[0] is not None:
                    return [result]
                return None
        except TypeError:
            try:
                result = cli.receive_data(timeout=timeout, max_samples=max_samples)
                if isinstance(result, tuple) and len(result) == 2 and result[0] is not None:
                    return [result]
                if isinstance(result, list) and len(result) > 0:
                    return result
                return None
            except Exception as e:
                try:
                    self.log(f"⚠ LSL数据接收错误: {str(e)}")
                except Exception:
                    pass
                return None
        except Exception as e:
            try:
                self.log(f"⚠ LSL数据接收错误: {str(e)}")
            except Exception:
                pass
            return None

    def run_sampling_rate_test(self, duration_seconds=10):
        start = time.time()
        count = 0
        while time.time() - start < duration_seconds:
            res = self._lsl_receive_batch(timeout=self.lsl_receive_timeout, max_samples=self.lsl_max_samples_per_pull)
            if isinstance(res, list):
                count += len(res)
            elif isinstance(res, tuple) and len(res) == 2:
                count += 1
            else:
                time.sleep(0.001)
        rate = count / duration_seconds
        try:
            self.log(f"采样率测试: {rate:.1f} Hz / 目标 {self.biosig_sampling_rate_target} Hz")
        except Exception:
            pass
        return rate

    def _biosig_stream_writer(self):
        csv_columns = ["system_timestamp","synchronized_timestamp","system_lsl_time","biosig_timestamp","heart_rate_bpm","gsr_uS","event_system_timestamp","event_synchronized_timestamp","marker_flag","marker_type","marker_condition","marker_target","marker_color","marker_label"]
        file_path = None
        try:
            filename = None
            if hasattr(self, 'current_trial_filenames') and self.current_trial_filenames.get('biosig'):
                filename = self.current_trial_filenames['biosig']
            elif hasattr(self, 'session_timestamp'):
                filename = f"biosig_{self.session_timestamp}.csv"
            else:
                filename = f"biosig_{self.__class__.datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            file_path = os.path.join(self.data_folder, filename)
            created = not os.path.exists(file_path)
            self._biosig_stream_writer_file = open(file_path, 'a', newline='', encoding='utf-8-sig')
            self._biosig_stream_writer_csv = csv.DictWriter(self._biosig_stream_writer_file, fieldnames=csv_columns, quoting=csv.QUOTE_NONNUMERIC)
            if created:
                try:
                    self._biosig_stream_writer_csv.writeheader()
                except Exception:
                    pass
        except Exception:
            self._biosig_stream_writer_file = None
            self._biosig_stream_writer_csv = None

        flush_counter = 0
        while not self._biosig_stream_writer_stop:
            try:
                item = self._biosig_stream_write_queue.get(timeout=0.5)
            except queue.Empty:
                continue
            except Exception:
                continue
            rows = item if isinstance(item, list) else [item]
            if not self._biosig_stream_writer_csv:
                continue
            for row in rows:
                if not isinstance(row, dict):
                    continue
                try:
                    self._biosig_stream_writer_csv.writerow({k: row.get(k, "N/A") for k in csv_columns})
                    flush_counter += 1
                except Exception:
                    pass
            if flush_counter >= 500:
                try:
                    self._biosig_stream_writer_file.flush()
                except Exception:
                    pass
                flush_counter = 0
        
        # 清理资源
        if self._biosig_stream_writer_file:
            try:
                self._biosig_stream_writer_file.close()
            except Exception:
                pass
            self._biosig_stream_writer_file = None
        self._biosig_stream_writer_csv = None

    def _marker_stream_writer(self):
        csv_columns = ["system_timestamp","synchronized_timestamp","lsl_timestamp","marker_flag","marker_type","marker_condition","marker_target","marker_color","marker_label","marker_stream_name"]
        try:
            filename = None
            if hasattr(self, 'current_trial_filenames') and self.current_trial_filenames.get('markers'):
                filename = self.current_trial_filenames['markers']
            elif hasattr(self, 'session_timestamp'):
                filename = f"markers_{self.session_timestamp}.csv"
            else:
                filename = f"markers_{self.__class__.datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            file_path = os.path.join(self.data_folder, filename)
            self._marker_stream_writer_file = open(file_path, 'a', newline='', encoding='utf-8-sig')
            self._marker_stream_writer_csv = csv.DictWriter(self._marker_stream_writer_file, fieldnames=csv_columns, quoting=csv.QUOTE_NONNUMERIC)
            if not getattr(self, 'trial_files_initialized', {}).get('markers', False):
                try:
                    self._marker_stream_writer_csv.writeheader()
                except Exception:
                    pass
                try:
                    self.trial_files_initialized['markers'] = True
                except Exception:
                    pass
        except Exception:
            self._marker_stream_writer_file = None
            self._marker_stream_writer_csv = None
            return

        flush_counter = 0
        while not self._marker_stream_writer_stop:
            try:
                item = self._marker_stream_write_queue.get(timeout=0.5)
            except queue.Empty:
                continue
            except Exception:
                continue
            if not self._marker_stream_writer_csv:
                continue
            if not isinstance(item, dict):
                continue
            try:
                self._marker_stream_writer_csv.writerow({k: item.get(k, "N/A") for k in csv_columns})
                flush_counter += 1
                try:
                    self._marker_rows_written = getattr(self, '_marker_rows_written', 0) + 1
                except Exception:
                    pass
            except Exception:
                pass
            if flush_counter >= 200:
                try:
                    self._marker_stream_writer_file.flush()
                except Exception:
                    pass
                flush_counter = 0

        try:
            while hasattr(self, '_marker_stream_write_queue') and self._marker_stream_write_queue.qsize() > 0:
                try:
                    item = self._marker_stream_write_queue.get_nowait()
                except Exception:
                    break
                if isinstance(item, dict) and self._marker_stream_writer_csv:
                    try:
                        self._marker_stream_writer_csv.writerow({k: item.get(k, "N/A") for k in csv_columns})
                    except Exception:
                        pass
        except Exception:
            pass

        if self._marker_stream_writer_file:
            try:
                self._marker_stream_writer_file.flush()
            except Exception:
                pass
            try:
                self._marker_stream_writer_file.close()
            except Exception:
                pass
        self._marker_stream_writer_file = None
        self._marker_stream_writer_csv = None

    def _driving_stream_writer(self):
        csv_columns = [
            "system_timestamp","synchronized_timestamp","lsl_timestamp",
            "marker_flag","marker_type","marker_condition","marker_target","marker_color","marker_label",
            "uv_ID","uv_Time","uv_Model","uv_description","uv_position X","uv_position Y","uv_position Z",
            "uv_distanceTravelled","uv_distanceAlongRoad","uv_steering","uv_throttle","uv_brake","uv_lightState",
            "uv_automaticControl","uv_wheelBase","uv_road","uv_distanceToLeftBorder","uv_distanceToRightBorder",
            "uv_offsetFromRoadCenter","uv_offsetFromLaneCenter","uv_laneNumber","uv_laneWidth","uv_drivingForwards",
            "uv_speedLimit","uv_speedOver",
            "fv_ID","fv_Time","fv_Model","fv_description","fv_position X","fv_position Y","fv_position Z",
            "fv_distanceTravelled","fv_steering","fv_throttle","fv_brake","fv_lightState","fv_automaticControl",
            "fv_wheelBase","fv_road","fv_distanceAlongRoad","fv_distanceToLeftBorder","fv_distanceToRightBorder",
            "fv_offsetFromRoadCenter","fv_offsetFromLaneCenter","fv_laneNumber","fv_laneWidth","fv_drivingForwards",
            "fv_speedLimit","fv_speedOver"
        ]

        flush_counter = 0
        try:
            if not hasattr(self, 'current_trial_filenames') or not self.current_trial_filenames.get('driving'):
                return
            p = os.path.join(self.data_folder, self.current_trial_filenames['driving'])
            self._driving_stream_writer_file = open(p, 'a', newline='', encoding='utf-8-sig')
            self._driving_stream_writer_csv = csv.DictWriter(self._driving_stream_writer_file, fieldnames=csv_columns, quoting=csv.QUOTE_NONNUMERIC)
            if not getattr(self, 'trial_files_initialized', {}).get('driving', False):
                try:
                    self._driving_stream_writer_csv.writeheader()
                except Exception:
                    pass
                try:
                    self.trial_files_initialized['driving'] = True
                except Exception:
                    pass
        except Exception:
            self._driving_stream_writer_file = None
            self._driving_stream_writer_csv = None
            return

        while not self._driving_stream_writer_stop:
            try:
                item = self._driving_stream_write_queue.get(timeout=0.5)
            except queue.Empty:
                continue
            except Exception:
                continue

            if not self._driving_stream_writer_csv:
                continue
            if not isinstance(item, dict):
                continue
            try:
                self._driving_stream_writer_csv.writerow({k: item.get(k, "N/A") for k in csv_columns})
                flush_counter += 1
                try:
                    self._driving_rows_written = getattr(self, '_driving_rows_written', 0) + 1
                except Exception:
                    pass
            except Exception:
                pass

            if flush_counter >= 500:
                try:
                    self._driving_stream_writer_file.flush()
                except Exception:
                    pass
                flush_counter = 0

        try:
            while hasattr(self, '_driving_stream_write_queue') and self._driving_stream_write_queue.qsize() > 0:
                try:
                    item = self._driving_stream_write_queue.get_nowait()
                except Exception:
                    break
                if isinstance(item, dict) and self._driving_stream_writer_csv:
                    try:
                        self._driving_stream_writer_csv.writerow({k: item.get(k, "N/A") for k in csv_columns})
                    except Exception:
                        pass
        except Exception:
            pass

        if self._driving_stream_writer_file:
            try:
                self._driving_stream_writer_file.flush()
            except Exception:
                pass
            try:
                self._driving_stream_writer_file.close()
            except Exception:
                pass
        self._driving_stream_writer_file = None
        self._driving_stream_writer_csv = None
    def _biosig_batch_processing(self):
        """调度批量处理任务，确保高频数据稳定处理"""
        if self.running:
            self._process_biosig_batch()
            self.root.after(self.queue_process_interval, self._biosig_batch_processing)

    def _schedule_sampling_rate_check(self):
        if self.running:
            self._check_sampling_rate()
            self.root.after(1000, self._schedule_sampling_rate_check)



if __name__ == "__main__":
    print("启动驾驶模拟器数据同步系统")
    print(f"Python版本: {platform.python_version()}")
    print(f"操作系统: {platform.system()} {platform.release()}")
    print("正在监控的驾驶参数:")
    for param in [
        "Model", "ID", "description", "laneWidth", "fv_distanceTravelled", "speedInKmPerHour", 
        "speedInMetresPerSecond", "distanceTravelled", "steering", 
        "brake", "lightState", "distanceAlongRoad", 
        "fv_ID", "fv_Time", "fv_Model", "fv_description", "fv_distanceToLeftBorder",
        "fv_distanceToRightBorder", "fv_speedInKmPerHour", "fv_steering", "fv_throttle", "fv_brake", "fv_lightState",
        "fv_automaticControl", "fv_wheelBase", "fv_road", "fv_distanceAlongRoad", "fv_laneNumber",
        "distanceToLeftBorder", "distanceToRightBorder", "offsetFromRoadCenter", "fv_position X", "fv_position Y", "fv_position Z",
        "offsetFromLaneCenter"
    ]:
        print(f"  - {param}")
    

    
    # 创建主窗口
    root = tk.Tk()
    
    # 创建应用实例
    app = CustomDrivingSyncSystem(root)
    
    # 设置关闭事件处理
    def on_closing():
        if app.running:
            app._stop_system()
        root.destroy()
    
    root.protocol("WM_DELETE_WINDOW", on_closing)
    
    # 运行主循环 - 系统现在只会在点击启动按钮时开始运行
    root.mainloop()
